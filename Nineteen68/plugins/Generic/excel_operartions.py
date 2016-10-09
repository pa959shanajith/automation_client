#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      sushma.p
#
# Created:     05-10-2016
# Copyright:   (c) sushma.p 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import logger
import generic_constants
import xlwt
import openpyxl
import xlsxwriter
import xlutils
import datetime
import xlrd
import os
import sys
from xlrd import open_workbook
from openpyxl import load_workbook
import Exceptions


class ExcelFile:

    def __init__(self):

        self.excel_path=''
        self.sheetname=''
        self.dict={
        'delete_row_.xls':self.delete_row_xls,
        'delete_row_.xlsx':self.delete_row_xlsx,
        'read_cell_.xls':self.read_cell_xls,
        'read_cell_.xlsx':self.read_cell_xlsx,
        'write_cell_.xls':self.write_cell_xls,
        'write_cell_.xlsx':self.write_cell_xlsx,
        'clear_cell_.xls':self.clear_cell_xls,
        'clear_cell_.xlsx':self.clear_cell_xlsx,
        'get_rowcount_.xls':self.get_rowcount_xls,
        'get_rowcount_.xlsx':self.get_rowcount_xlsx,
        'get_colcount_.xls':self.get_colcount_xls,
        'get_colcount_.xlsx':self.get_colcount_xlsx,
        }
        self.cell_type={'string' : 's',
        'formula': 'f',
        'numeric': 'n',
        'date':'n',
        'bool': 'b',
        'null':'n',
        'inline': 'inlineStr',
        'error': 'e',
        'formula_cache': 'str'}

    def __get_ext(self,input_path):
        try:
            filename,file_ext=os.path.splitext(input_path)
            if file_ext in generic_constants.FILE_TYPES:
                status=True
                logger.log('File type is'+file_ext)
                return file_ext,status
            else:
                logger.log(generic_constants.INVALID_FILE_FORMAT)
        except Exception as e:
            Exceptions.error(e)
        return '',False

    def delete_row(self,row):
        try:
            input_path=input_path.strip()
            file_ext,status=self.__get_ext(input_path)
            if status == True:
                status=self.dict['delete_row_'+file_ext](row)
                return status
        except Exception as e:
            Exceptions.error(e)
        return False


    def read_cell(self,row,col,*args):
        try:
            file_ext,status=self.__get_ext(self.excel_path)
            if status == True:
                status=self.dict['read_cell_'+file_ext](int(row),int(col),*args)
                return status
        except Exception as e:
            Exceptions.error(e)
        return False


    def write_cell(self,row,col,value):
        try:
            file_ext,status=self.__get_ext(self.excel_path)
            if status == True:
                status=self.dict['write_cell_'+file_ext](int(row),int(col),value)
                return status
        except Exception as e:
            Exceptions.error(e)
        return False


    def clear_cell(self,row,col):
        try:
            file_ext,status=self.__get_ext(self.excel_path)
            if status == True:
                status=self.dict['clear_cell_'+file_ext](row,col)
                return status
        except Exception as e:
            Exceptions.error(e)
        return False


    def get_rowcount(self,*args):
        try:
            file_ext,status=self.__get_ext(self.excel_path)
            if status == True:
                status=self.dict['get_rowcount_'+file_ext](*args)
                logger.log('Row count is:'+str(status))
                return status
        except Exception as e:
            Exceptions.error(e)
        return False



    def get_colcount(self,*args):
        try:
            file_ext,status=self.__get_ext(self.excel_path)
            if status == True:
                status=self.dict['get_colcount_'+file_ext](*args)
                logger.log('Column count is:'+str(status))
                return status
        except Exception as e:
            Exceptions.error(e)
        return False



    def get_linenumber_xls(self,input_path,sheetname,content):
        value=None
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        book = open_workbook(self.excel_path)
        sheet = book.sheet_by_name(self.sheetname)
        for col in range(sheet.ncols):
##            col_values=sheet.col_values(col)
            indices = [i for i, x in enumerate(sheet.col_values(col)) if x == content]
            value=indices
        return value


    def get_linenumber_xlsx(self,input_path,sheetname,content):
        value=None
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        book = load_workbook(self.excel_path,data_only=True)
        sheet = book.get_sheet_by_name(self.sheetname)
        i=0
        for row in sheet.iter_rows():
            i+=1
            for cell in row:
                if cell.internal_value==content:
                    value.append(i)
        return value


    def delete_row_xls(self,row):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        from win32com.client.gencache import EnsureDispatch
        excel = EnsureDispatch("Excel.Application")
        excel_file = excel.Workbooks.Open(inputpath)
        sheet = excel.Sheets(sheetname)
        sheet.Rows(row).Delete()
        excel_file.Close(True)

    def delete_row_xlsx(self,row):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        #same functionality as compare_content_xlsx
        self.delete_row_xls(row)


    def compare_content_xls(self,input_path1,sheetname1,input_path2,sheetname2,*args):
        import pandas as pd
        file1=pd.ExcelFile(input_path1)
        data1=file1.parse(sheetname1)
        logger.log('File content1'+data1)
        file2=pd.ExcelFile(input_path2)
        data2=file2.parse(sheetname2)
        logger.log('File content2'+data2)
        return list(data1)==list(data2)

    def compare_content_xlsx(self,input_path1,sheetname1,input_path2,sheetname2,*args):
        logger.log(generic_constants.INPUT_IS+input_path1+' '+input_path2)
        #same functionality as compare_content_xlsx
        self.compare_content_xls(input_path1,sheetname1,input_path2,sheetname2)

    def replace_content_xls(self,input_path,sheetname,existingcontent,replacecontent,*args):
        logger.log(generic_constants.INPUT_IS+input_path+' '+sheetname)
        from xlutils.copy import copy
        book = open_workbook(self.excel_path)
        sheet = book.sheet_by_name(self.sheetname)
        workbook = copy(book)
        sheets=book.sheet_names()
        sheetnum=sheets.index(sheetname)
        s = workbook.get_sheet(sheetnum)
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                if sheet.cell(row,col).value==existingcontent:
                    s.write(row,col,replacecontent)
                    workbook.save(input_path)

    def replace_content_xlsx(self,input_path,sheetname,existingcontent,replacecontent,*args):
        logger.log(generic_constants.INPUT_IS+input_path+' '+sheetname)
        book = load_workbook(input_path,data_only=True)
        sheet = book.get_sheet_by_name(sheetname)
        i=0
        for row in sheet.iter_rows():
            i+=1
            for cell in row:
                if cell.internal_value==existingcontent:
                    cell.value=replacecontent
        book.save(input_path)

    def verify_content_xls(self,input_path,sheetname,content,*args):
        logger.log(generic_constants.INPUT_IS+input_path+' '+sheetname)
        if self.get_linenumber_xls(input_path,sheetname,content) != []:
            return True
        else:
            return False

    def verify_content_xlsx(self,input_path,sheetname,content,*args):
        logger.log(generic_constants.INPUT_IS+input_path+' '+sheetname)
        if self.get_linenumber_xlsx(input_path,sheetname,content) != []:
            return True
        else:
            return False

    def set_excel_path(self,input_path,sheetname):
        logger.log(generic_constants.INPUT_IS+input_path+' '+sheetname)
        self.excel_path=input_path
        self.sheetname=sheetname
        return True

    def clear_content_xlsx(self,inputpath,sheetname):
        logger.log(generic_constants.INPUT_IS+inputpath+' '+sheetname)
        book=load_workbook(self.inputpath)
        sheet=book.get_sheet_by_name(self.sheetname)
        book.remove(sheet)
        book.active=len(book.sheetnames)-1
        book.save(self.inputpath)
        return True

    def clear_content_xls(self,inputpath,sheetname):
        logger.log(generic_constants.INPUT_IS+inputpath+' '+sheetname)
        from win32com.client.gencache import EnsureDispatch
        excel = EnsureDispatch("Excel.Application")
        excel_file = excel.Workbooks.Open(inputpath)
        sheet = excel.Sheets(sheetname)
        sheet.Delete()
        excel_file.Close(True)


    def clear_excel_path(self):
        self.excel_path=''
        self.sheetname=''
        return True

    def read_cell_xls(self,row,col,*args):
        value=None
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        from xlrd import open_workbook
        book = open_workbook(self.excel_path)
        sheet = book.sheet_by_name(self.sheetname)
        if row<sheet.nrows and col<sheet.ncols:
            cell=sheet.cell(row-1,col-1)
            if cell.ctype==3:
                value=datetime.datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))
                value=value.date()
            else:
                value=cell.value
            logger.log(value)
        else:
            logger.log(generic_constants.INDEX_EXCEEDS)

        return value


    def read_cell_xlsx(self,row,col,*args):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        wb=openpyxl.load_workbook(self.excel_path,read_only=True,data_only=True)
        sheet=wb.get_sheet_by_name(self.sheetname)
##        sheet.get_index()
        cell=sheet.cell(row=row,column=col,value=None)
        print cell.internal_value
        print cell.value
        return cell.value

    def __write_to_cell_xls(self,input_path,book,sheetname,row,col,value):
        from xlutils.copy import copy
        workbook = copy(book)
        sheets=book.sheet_names()
        try:
            sheetnum=sheets.index(sheetname)
            s = workbook.get_sheet(sheetnum)
            s.write(int(row),int(col),value)
            workbook.save(input_path)
        except Exception as e:
            Exceptions.error(e)

    def __load_workbook_xls(self,inputpath,sheetname):
        book = open_workbook(inputpath,formatting_info=True)
        sheet=book.sheet_by_name(sheetname)
        if sheet.nrows==0:
            last_row=0
            last_col=0
        else:
            last_row=sheet.nrows-1
            row=sheet.row(last_row)
            last_col=len(row)

        return book,sheet,last_row,last_col,sheet.nrows

    def write_to_file_xls(self,input_path,sheetname,content,*args):
       logger.log(generic_constants.INPUT_IS+input_path+' '+sheetname)
       #loads the xls workbook
       workbook_info=self.__load_workbook_xls(input_path,sheetname)
       row=workbook_info[2]
       col=workbook_info[3]
       if len(args)>0 and args[0] == 'newline' and workbook_info[3] !=0:
            row=row+1
            col=0
       #writes to the cell in given row,col
       self.__write_to_cell_xls(input_path,workbook_info[0],sheetname,row,col,content,)


    def write_cell_xls(self,row,col,value):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        #loads the xls workbook
        workook_info=self.__load_workbook_xls(self.excel_path,self.sheetname)
        #writes to the cell in given row,col
        self.__write_to_cell_xls(self.excel_path,workook_info[0],self.sheetname,row,col,value)

    def __write_to_cell_xlsx(self,input_path,book,sheetname,row,col,value,*args):
        try:
           sheet=book.get_sheet_by_name(sheetname)
           cell=sheet.cell(row=row,column=col)
           cell.value=value
           book.save(input_path)
        except Exception as e:
            Exceptions.error(e)
            book.save(input_path)

    def __load_workbook_xlsx(self,inputpath,sheetname):
        book=load_workbook(inputpath)
        sheet=book.get_sheet_by_name(sheetname)
        if sheet.max_row==1:
            last_row=1
            last_col=1
        else:
            last_row=sheet.max_column
            row=list(sheet.iter_rows())[last_row-1]
            last_col=len(row)

        return book,sheet,last_row,last_col,sheet.max_row

    def write_cell_xlsx(self,row,col,value,*args):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        #loads the xls workbook
        workbook_info=self.__load_workbook_xlsx(self.excel_path,self.sheetname)
        #writes to the cell in given row,col
        self.__write_to_cell_xlsx(self.excel_path,workbook_info[0],self.sheetname,int(row),int(col),value,*args)

    def write_to_file_xlsx(self,input_path,sheetname,content,*args):
        logger.log(generic_constants.INPUT_IS+input_path+' '+sheetname)
        #loads the xls workbook
        workbook_info=self.__load_workbook_xlsx(input_path,sheetname)
        row=workbook_info[2]
        col=workbook_info[3]
        if len(args)>0 and args[0] == 'newline' and workbook_info[3] !=0:
            print row
            row=row+1
            col=1

        #writes to the cell in given row,col
        self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,content,)



    def clear_cell_xls(self,row,col):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        from xlutils.copy import copy
        book = open_workbook(self.excel_path)
        workbook = copy(book)
        sheets=book.sheet_names()
        try:
            sheetnum=sheets.index(self.sheetname)
            s = workbook.get_sheet(sheetnum)
            s.write(row,col,'')
            workbook.save(input_path)
        except ValueError:
            Exception.message('Value Error occured')

    def clear_cell_xlsx(self,row,col):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        workbook=openpyxl.load_workbook(self.excel_path)
        sheet=workbook.get_sheet_by_name(self.sheetname)
        cell=sheet.cell(row=row,column=col,value=None)
        cell.value=None
        workbook.save(self.excel_path)


    def get_rowcount_xls(self,*args):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        book = open_workbook(self.excel_path)
        sheet=book.sheet_by_name(self.sheetname)
        rowcount=sheet.nrows
        return rowcount

    def get_rowcount_xlsx(self,*args):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        book=openpyxl.load_workbook(self.excel_path)
        sheet=book.get_sheet_by_name(self.sheetname)
        rowcount=sheet.max_row
        return rowcount

    def get_colcount_xls(self,*args):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        book = open_workbook(self.excel_path)
        sheet=book.sheet_by_name(self.sheetname)
        colcount=sheet.ncols
        return colcount

    def get_colcount_xlsx(self,*args):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        book=openpyxl.load_workbook(self.excel_path)
        sheet=book.get_sheet_by_name(self.sheetname)
        colcount=sheet.max_column
        return colcount
