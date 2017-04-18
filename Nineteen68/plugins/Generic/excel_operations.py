#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      sushma.p
#
# Created:     05-10-2016
# Copyright:   (c) sushma.p 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import logger
import generic_constants
import xlwt
import openpyxl
import xlsxwriter
import xlutils
import datetime
import xlrd
import os
import sys
from xlrd import open_workbook
from openpyxl import load_workbook
from constants import *
import logging
from constants import *
log = logging.getLogger('excel_operations.py')


class ExcelFile:

    """The instantiation operation __init__ creates an empty object of the class ExcelFile when it is instantiated"""
    def __init__(self):

        self.excel_path=None
        self.sheetname=None
        self.xls_obj=ExcelXLS()
        self.xlsx_obj=ExcelXLSX()

        self.dict={
        'delete_row_.xls':self.xls_obj.delete_row_xls,
        'read_cell_.xls':self.xls_obj.read_cell_xls,
        'write_cell_.xls':self.xls_obj.write_cell_xls,
        'clear_cell_.xls':self.xls_obj.clear_cell_xls,
        'get_rowcount_.xls':self.xls_obj.get_rowcount_xls,
        'get_colcount_.xls':self.xls_obj.get_colcount_xls,

        'delete_row_.xlsx':self.xlsx_obj.delete_row_xlsx,
        'read_cell_.xlsx':self.xlsx_obj.read_cell_xlsx,
        'write_cell_.xlsx':self.xlsx_obj.write_cell_xlsx,
        'clear_cell_.xlsx':self.xlsx_obj.clear_cell_xlsx,
        'get_rowcount_.xlsx':self.xlsx_obj.get_rowcount_xlsx,
        'get_colcount_.xlsx':self.xlsx_obj.get_colcount_xlsx,

        }

    def open_and_save_file(self,input_path):
        #win32 part opening and closing of file to claculate and save values of the formula
        #This is to support the feature of   simulataneous writing and reading to a file
        from win32com.client.gencache import EnsureDispatch
        excel = EnsureDispatch("Excel.Application")
        excel.DisplayAlerts = False
        excel_file = excel.Workbooks.Open(input_path)
        excel_file.Close(True)



    def __get_ext(self,input_path):
        """
        def : __get_ext(private method)
        purpose : returns the extension/type of the given file and checks if it is valid type
        param : input_path
        return : Returns Status and file type

        """
        err_msg=None
        try:
            filename,file_ext=os.path.splitext(input_path)
            status=True
            log.info('File type is :')
            log.info(file_ext)
            return file_ext,status,err_msg

        except Exception as e:
            log.error(e)
            err_msg=generic_constants.INVALID_FILE_FORMAT
##            logger.print_on_console(err_msg)
        return '',False,err_msg

    def set_excel_path(self,input_path,sheetname):
        """
        def : set_excel_path
        purpose : sets the given excel file path and sheet name for excel operations
        param : input_path,sheetname
        return : bool

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        output = OUTPUT_CONSTANT
        err_msg = None
        logger.print_on_console(generic_constants.INPUT_IS+input_path+' '+sheetname)
        if input_path != None and sheetname != None and input_path != '' and sheetname != '':
            if os.path.isfile(input_path):
                file_ext,res,err_msg=self.__get_ext(input_path)
                if file_ext in generic_constants.EXCEL_TYPES:
                    log.info('File type is :'+str(file_ext))
                    logger.print_on_console('File type is :'+str(file_ext))
                    self.excel_path=input_path
                    self.sheetname=sheetname
                    status=TEST_RESULT_PASS
                    methodoutput=TEST_RESULT_TRUE
            else:
                err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_FOUND_EXCEPTION']
        else:
            err_msg=generic_constants.INVALID_INPUT
        if err_msg != None:
            log.error(err_msg)
            logger.print_on_console(err_msg)

        return status,methodoutput,output,err_msg


    def delete_row(self,row):
        """
        def : delete_row
        purpose : calls the respective method to delete the given row of excel
                  file set by the user
        param : row
        return : Returns Bool

        """
        output = OUTPUT_CONSTANT
        err_msg = None
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        try:
            if self.excel_path != None and self.sheetname != None:
                file_ext,res,err_msg=self.__get_ext(self.excel_path)
                if res:
                    info_msg=generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname+' and the row is '+str(row)
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    myfile = open(self.excel_path, "a+") # or "a+", whatever you need
                    myfile.close()
                    res,err_msg=self.dict['delete_row_'+file_ext](int(row),self.excel_path,self.sheetname)
                    if res:
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
            else:
                err_msg=generic_constants.FILE_PATH_NOT_SET
        except IOError:
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
        except ValueError as e:
            err_msg=ERROR_CODE_DICT['ERR_NUMBER_FORMAT_EXCEPTION']
        except Exception as e:
            err_msg='Error occured in deleting row'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output,err_msg


    def read_cell(self,row,col,*args):
        """
        def : read_cell
        purpose : calls the respective method to read the given cell of excel
                  file set by the user
        param : row,col
        return : Returns cell value

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        output=None
        err_msg = None
        try:
            if self.excel_path != None and self.sheetname != None:
                file_ext,res,err_msg=self.__get_ext(self.excel_path)
                if res:
                    info_msg=generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    info_msg='Row is '+str(row)+' and col is '+str(col)
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    row=int(row)
                    col=int(col)
                    if row>0 and col>0:
                        res,value,err_msg=self.dict['read_cell_'+file_ext](row,col,self.excel_path,self.sheetname,*args)
                        info_msg='cell value is '+str(value)
                        logger.print_on_console(info_msg)
                        log.info(info_msg)
                        if res:
                            status=TEST_RESULT_PASS
                            methodoutput=TEST_RESULT_TRUE
                            output=value
                    else:
                        err_msg=ERROR_CODE_DICT["ERR_ROW_COLUMN"]
            else:
                err_msg=generic_constants.FILE_PATH_NOT_SET
        except ValueError as e:
            err_msg=ERROR_CODE_DICT['ERR_NUMBER_FORMAT_EXCEPTION']
        except Exception as e:
            err_msg='Error occured in reading excel cell value'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output,err_msg


    def write_cell(self,row,col,value,*args):
        """
        def : write_cell
        purpose : calls the respective method to write to the given cell of excel
                  file set by the user
        param : row,col,value
        return : Returns Bool

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        output = OUTPUT_CONSTANT
        err_msg = None
        try:
            if self.excel_path != None and self.sheetname != None:
                file_ext,res,err_msg=self.__get_ext(self.excel_path)
                if res:
                    info_msg=generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname
                    log.info(info_msg)
                    info_msg='Row is '+str(row)+' col is '+str(col)+' and Value: '+str(value)
                    log.info(info_msg)
##                    logger.print_on_console(info_msg)
                    row=int(row)
                    col=int(col)
                    if row>0 and col>0:
                        res,err_msg=self.dict['write_cell_'+file_ext](row,col,value,self.excel_path,self.sheetname,*args)
                        if res:
                            status=TEST_RESULT_PASS
                            methodoutput=TEST_RESULT_TRUE
                    else:
                        err_msg=ERROR_CODE_DICT["ERR_ROW_COLUMN"]
            else:
                err_msg=generic_constants.FILE_PATH_NOT_SET
        except ValueError as e:
            err_msg=ERROR_CODE_DICT['ERR_NUMBER_FORMAT_EXCEPTION']
        except IOError:
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
        except Exception as e:
            err_msg='Error occured in writing excel cell value'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output,err_msg


    def clear_cell(self,row,col):
        """
        def : clear_cell
        purpose : calls the respective method to clear the given cell of excel
                  file set by the user
        param : row,col
        return : Returns Bool

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        output = OUTPUT_CONSTANT
        err_msg = None
        try:
            if self.excel_path != None and self.sheetname != None:
                file_ext,res,err_msg=self.__get_ext(self.excel_path)
                if res:
                    info_msg=generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    info_msg='Row is '+str(row)+' and col is '+str(col)
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    row=int(row)
                    col=int(col)
                    if row>0 and col>0:
                        res,err_msg=self.dict['clear_cell_'+file_ext](row,col,self.excel_path,self.sheetname)
                        if res:
                            status=TEST_RESULT_PASS
                            methodoutput=TEST_RESULT_TRUE
                    else:
                        err_msg=ERROR_CODE_DICT["ERR_ROW_COLUMN"]
            else:
                err_msg=generic_constants.FILE_PATH_NOT_SET
        except ValueError as e:
            err_msg=ERROR_CODE_DICT['ERR_NUMBER_FORMAT_EXCEPTION']
        except IOError:
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
        except Exception as e:
            err_msg='Error occured in clearing excel cell value'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output,err_msg


    def get_rowcount(self,*args):
        """
        def : get_rowcount
        purpose : calls the respective method to get the number of rows of excel
                  file set by the user
        param : None
        return : Returns row count(int)

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        row_count=None
        err_msg = None
        try:
            if self.excel_path != None and self.sheetname != None:
                file_ext,res,err_msg=self.__get_ext(self.excel_path)
                if res:
                    info_msg=generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    res,row_count,err_msg=self.dict['get_rowcount_'+file_ext](self.excel_path,self.sheetname,*args)
                    logger.print_on_console('Row count is:'+str(row_count))
                    if res:
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
            else:
                err_msg=generic_constants.FILE_PATH_NOT_SET
        except Exception as e:
            err_msg='Error occured in getting row count of excel'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,row_count,err_msg



    def get_colcount(self,*args):
        """
        def : get_colcount
        purpose : calls the respective method to get the number of columns of excel
                  file set by the user
        param : None
        return : Returns col count(int)

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        col_count=None
        err_msg = None
        try:
            if self.excel_path != None and self.sheetname != None:
                file_ext,res,err_msg=self.__get_ext(self.excel_path)
                if res:
                    info_msg=generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    res,col_count,err_msg=self.dict['get_colcount_'+file_ext](self.excel_path,self.sheetname,*args)
                    logger.print_on_console('Column count is:'+str(col_count))
                    if res:
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
            else:
                err_msg=generic_constants.FILE_PATH_NOT_SET
        except Exception as e:
            err_msg='Error occured in getting column count of excel'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,col_count,err_msg

    def clear_excel_path(self,*args):
        """
        def : clear_excel_path
        purpose : celars the excel path
        param : None
        return : bool

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        output = OUTPUT_CONSTANT
        err_msg = None
        if self.excel_path != None and self.excel_path != '' and self.sheetname != None and self.sheetname != '':
            self.excel_path=''
            self.sheetname=''
            status=TEST_RESULT_PASS
            methodoutput=TEST_RESULT_TRUE
        else:
            err_msg=generic_constants.FILE_PATH_NOT_SET
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output,err_msg


class ExcelXLS:

    def __init__(self):
        self.excel_obj=''
        self.cell_type={'string' : 'General',
        'formula': 'f',
        'number': '0.00',
        'date':'n',
        'boolean': 'b',
        'null':'n',
        'inline': 'inlineStr',
        'error': 'e',
        'formula_cache': 'str'}


    def __load_workbook_xls(self,inputpath,sheetname):
        """
        def : __load_workbook_xls(private method)
        purpose : loads the given .xls file
        param : inputpath,sheetname
        return : list

        """
        book = open_workbook(inputpath,formatting_info=True)
        sheet=book.sheet_by_name(sheetname)
        last_row=-1
        last_col=-1
        if sheet.nrows>0:
            last_row=sheet.nrows-1
            row=sheet.row(last_row)
            last_col=len(row)-1

        return book,sheet,last_row,last_col,sheet.nrows



    def __write_to_cell_xls(self,input_path,book,sheetname,row,col,value,*args):
        """
        def : __write_to_cell_xls(private method)
        purpose : writes to the the cell in the given row and column of .xls file
        param : input_path,book,sheetname,row,col,value
        return : bool

        """
        from xlutils.copy import copy
        from xlwt import easyxf,XFStyle
        workbook = copy(book)
        sheets=book.sheet_names()
        status=False
        err_msg=None
        flag=False
        self.excel_obj=ExcelFile()
        try:
            sheetnum=sheets.index(sheetname)
            s = workbook.get_sheet(sheetnum)
            if len(args)>0 and args[0] is not None:
                type=args[0].lower()
                if type in self.cell_type.keys():
                    type=self.cell_type[type]
                    if type=='0.00':
                        flag=True
                    	#Setting cell format to number
                        style=XFStyle()
                        style.num_format_str=type
                        value=int(value)
                        s.write(int(row),int(col),value,style)
                    elif type in ['f','b']:
                        flag=True
                    	#Evaluating formula
                        value=xlwt.ExcelFormula.Formula(value)
##                        value='='+value
                        s.write(int(row),int(col),value)
                    else:
                        flag=True
                        value=str(value)
                        s.write(int(row),int(col),value)

                    workbook.save(input_path)
                    status= True
                else:
                    logger.print_on_console('Cell Type not supported')
                    log.info('Cell Type not supported')
            else:
                flag=True
                status= True
                value=str(value)
                s.write(int(row),int(col),value)


            workbook.save(input_path)

        except IOError:
            err_msg=err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
            log.error(err_msg)
        except Exception as e:
            log.error(e)
            err_msg='Error writing to excel cell of .xls file'
        workbook.save(input_path)
        if flag and status:
            self.excel_obj.open_and_save_file(input_path)
        return status,err_msg

    def get_linenumber_xls(self,input_path,sheetname,content):
        """
        def : get_linenumber_xls
        purpose : get the line number where content is present in given .xls file
        param  : input_path,sheetname,content
        return : Returns line number (list)

        """
        status=False
        line_number=None
        err_msg=None
        try:
            logger.print_on_console(generic_constants.INPUT_IS+input_path+' '+sheetname)
            log.info(generic_constants.INPUT_IS+input_path+' '+sheetname)
            book = open_workbook(input_path)
            sheet = book.sheet_by_name(sheetname)
            for col in range(sheet.ncols):
                indices = [i for i, x in enumerate(sheet.col_values(col)) if x == content]
                line_number=indices
                if line_number != None:
                    log.debug('line numbers are '+''.join(str(line_number)))
                log.debug(line_number)
                status=True
        except Exception as e:
           err_msg='Error getting line number in .xlsx'
           log.error(e)
           logger.print_on_console(err_msg)
        log.info('Status is '+str(status))
        return status,line_number,err_msg

    def replace_content_xls(self,input_path,sheetname,existingcontent,replacecontent,*args):
        """
        def : replace_content_xls
        purpose : replace the 'existingcontent' by 'replacecontent' in given sheet of
                  given .xls file
        param : input_path,sheetname,existingcontent,replacecontent
        return : bool

        """
        logger.print_on_console(generic_constants.INPUT_IS+input_path+' '+sheetname)
        log.info(generic_constants.INPUT_IS+input_path+' '+sheetname)
        status=False
        err_msg=None
        log.debug('Replacing content of .xls files')
        try:
            from xlutils.copy import copy
            book = open_workbook(input_path)
            sheet = book.sheet_by_name(sheetname)
            workbook = copy(book)
            sheets=book.sheet_names()
            sheetnum=sheets.index(sheetname)
            s = workbook.get_sheet(sheetnum)
            log.debug('Existing content '+str(existingcontent))
            log.debug('Replacing content '+str(replacecontent))
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    if existingcontent == sheet.cell(row,col).value:
                        log.debug('Replacing at '+str(col))
                        s.write(row,col,replacecontent)
            workbook.save(input_path)
            status=True
        except Exception as e:
           err_msg='Error occured in replacing content of .xls files'
           log.error(e)
        log.info('Status is '+str(status))
        return status,err_msg

    def delete_row_xls(self,row,excel_path,sheetname):
        """
        def : delete_row_xls
        purpose : deletes the given row of both .xls and .xlsx files
        param : row
        return : bool

        """
        status=False
        err_msg=None
        self.excel_obj=ExcelFile()
        log.debug('Deleting a row of .xls file')
        excel_file=None
        try:
            import pythoncom
            pythoncom.CoInitialize()
            from win32com.client.gencache import EnsureDispatch
            excel = EnsureDispatch("Excel.Application")
            excel.DisplayAlerts = False
            excel_file = excel.Workbooks.Open(excel_path)

            if not(excel_file.ReadOnly):
                sheet = excel.Sheets(sheetname)
                last_row=sheet.UsedRange.Rows.Count
                if row<=last_row:
                    sheet.Rows(row).Delete()
                    status=True
                else:
                    err_msg=ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
            else:
                err_msg='Excel is Read only'
        except Exception as e:
            err_msg='Error occured in deleting row of excel file'
            log.error(e)
        finally:
            if excel_file!=None:
                excel_file.Close(True)


        if err_msg!=None:
            log.error(err_msg)
        log.info('Status is '+str(status))
        return status,err_msg

    def compare_content_xls(self,input_path1,sheetname1,input_path2,sheetname2,*args):
        """
        def : compare_content_xls
        purpose : compares the data of given sheets of 2 different excel files
        param : input_path1,input_path2,Sheet1,Sheet2
        return : bool

        """
        status=False
        err_msg=None
        log.debug('Comparing content of .xls files')
        try:
##            import pandas as pd
##            file1=pd.ExcelFile(input_path1)
##            data1=file1.parse(sheetname1)
##            log.debug('File content1'+str(data1))
##            file2=pd.ExcelFile(input_path2)
##            data2=file2.parse(sheetname2)
##            log.debug('File content2'+str(data2))
##            status= (str(data1)==str(data2))
            from itertools import izip_longest
            book1 = open_workbook(input_path1)
            book2 = open_workbook(input_path2)
            sheet1 = book1.sheet_by_name(sheetname1)
            sheet2 = book2.sheet_by_name(sheetname2)

            for rownum in range(max(sheet1.nrows, sheet2.nrows)):
                if rownum < sheet1.nrows and rownum < sheet2.nrows:
                    row_rb1 = sheet1.row_values(rownum)
                    row_rb2 = sheet2.row_values(rownum)

                    for colnum, (c1, c2) in enumerate(izip_longest(row_rb1, row_rb2)):
                        if c1 != c2:
                            log.debug('Row ',str(rownum+1),' Col ',str(colnum+1),' cell value 1 ',c1,' cell value 2 ',c2)
                            break
                    status=True
                else:
                    log.error('Row ',str(rownum+1),' is missing')
                    err_msg='File contents are not same'
                    status=False
                    break


        except Exception as e:
            import traceback
            traceback.print_exc()
            err_msg='Error occured in compare content of .xls files'
            log.error(e)
        log.info('Status is '+str(status))
        return status,err_msg

    def verify_content_xls(self,input_path,sheetname,content,*args):
        """
        def : verify_content_xls
       purpose : verify whether the 'existingcontent' is present in given sheet of
                  given .xls file
        param : input_path,sheetname,content
        return : bool

        """
        status=False
        err_msg=None
        res,line_number,err_msg=self.get_linenumber_xls(input_path,sheetname,content)
        if res and line_number != None:
            status=True
        return status,err_msg

    def clear_content_xls(self,inputpath,filename,sheetname):
        """
        def : clear_content_xls
        purpose : removes the given sheet from the given .xls file
        param : input_path,sheetname
        return : bool

        """
        status=False
        err_msg=None
        info_msg=generic_constants.INPUT_IS+inputpath+' filename: '+filename+' Sheetname: '+sheetname
        logger.print_on_console(info_msg)
        log.info(info_msg)
        inputpath=inputpath+'/'+filename
        try:
            from win32com.client.gencache import EnsureDispatch
            excel = EnsureDispatch("Excel.Application")
            excel.DisplayAlerts = False
            excel_file = excel.Workbooks.Open(inputpath)
            sheet = excel.Sheets(sheetname)
            if excel_file.Sheets.Count>1:
                sheet.Delete()
            else:
                err_msg='Cannot clear the content if single sheet is present in Excel'
                log.error(err_msg)
            excel_file.Close(True)
            status=True
        except Exception as e:
            err_msg='File/Sheet does not exist to clear'
            log.error(e)
        return status,err_msg

    def read_cell_xls(self,row,col,excel_path,sheetname,*args):
        """
        def : read_cell_xls
        purpose : reads the cell value in the given row and column of .xls file
        param : input_path,sheetname
        return : cell value

        """
        status=False
        value=None
        err_msg=None
        try:
            log.debug('Writing to the file'+generic_constants.INPUT_IS+excel_path+' '+sheetname)
            book = open_workbook(excel_path)
            sheet = book.sheet_by_name(sheetname)
            if row<=sheet.nrows:
                if col<=sheet.ncols:
                    cell=sheet.cell(row-1,col-1)
                    value=cell.value
                    if cell.ctype==3:
                        value=datetime.datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))
                        value=value.date()
                    status=True
                    log.info(value)
                else:
                    err_msg=ERROR_CODE_DICT["ERR_COL_DOESN'T_EXIST"]
            else:
                err_msg=ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
        except Exception as e:
            err_msg='Error occured in read cell of .xls file'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
        return status,value,err_msg



    def clear_cell_xls(self,row,col,excel_path,sheetname):
        """
        def : clear_cell_xls
        purpose : clears the given cell of .xls file set by the user
        param : row,col
        return : bool

        """
        status=False
        err_msg=None
        log.debug(generic_constants.INPUT_IS+excel_path+' '+sheetname)
        from xlutils.copy import copy
        book = open_workbook(excel_path)
        workbook = copy(book)
        sheets=book.sheet_names()
        try:
            if row>0 and col>0:
                sheetnum=sheets.index(sheetname)
                s = workbook.get_sheet(sheetnum)
                s.write(row-1,col-1,'')
                workbook.save(excel_path)
                status=True
            else:
                logger.print_on_console('Row/Col should be greater than 0')
        except Exception as e:
            err_msg='Error occurred in clearing cell of .xlsx file'
        workbook.save(excel_path)
        return status


    def write_cell_xls(self,row,col,value,excel_path,sheetname,*args):
        """
        def : write_cell_xls
        purpose : calls the private methods to write given value to the cell of given
                row or col of .xls file
        param : row,col,value
        return : bool

        """
        #loads the xls workbook
        workook_info=self.__load_workbook_xls(excel_path,sheetname)
        #writes to the cell in given row,col
        status=False
        if(int(row)>0 and int(col)>0):
            row=int(row)-1
            col=int(col)-1
            status,err_msg=self.__write_to_cell_xls(excel_path,workook_info[0],sheetname,row,col,value,*args)
        return status,err_msg

    def get_rowcount_xls(self,excel_path,sheetname,*args):
        """
        def : get_rowcount_xls
        purpose : get the number of rows of .xls excel file set by the user
        param : None
        return : Returns row count(int)

        """
        status=False
        row_count=None
        err_msg=None
        try:
            log.debug('Fetching row count of '+generic_constants.INPUT_IS+excel_path+' '+sheetname)
            book = open_workbook(excel_path)
            sheet=book.sheet_by_name(sheetname)
            row_count=sheet.nrows
            status=True
        except Exception as e:
            err_msg='Error getting the row count of .xls file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,row_count,err_msg

    def get_colcount_xls(self,excel_path,sheetname,*args):
        """
        def : get_colcount_xls
        purpose : get the number of columns of .xls excel file set by the user
        param : None
        return : Returns row count(int)

        """
        status=False
        col_count=None
        err_msg=None
        log.debug('Fetching col count of '+generic_constants.INPUT_IS+excel_path+' '+sheetname)
        try:
            book = open_workbook(excel_path)
            sheet=book.sheet_by_name(sheetname)
            col_count=sheet.ncols
            status=True
        except Exception as e:
            err_msg='Error getting the col count of .xls file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,col_count,err_msg

    def write_to_file_xls(self,input_path,sheetname,content,*args):
        """
        def : write_to_file_xls
        purpose : writes the given content to the given .xls file
        param : inputpath,sheetname,content,('newline'-optional param)
        return : bool

        """
        self.excel_obj=ExcelFile()
        status=False
        err_msg=None
        log.debug(generic_constants.INPUT_IS+input_path+' '+sheetname)
        try:
            #loads the xls workbook
            workbook_info=self.__load_workbook_xls(input_path,sheetname)
            row=workbook_info[2]
            col=workbook_info[3]
            if len(args)>0 and args[0].lower() == 'newline'  :
                col=0
                if workbook_info[3] != -1:
                    row=row+1
                    col=workbook_info[3]
                else:
                    row=0
            elif workbook_info[3]==-1:
                row=col=0
            else:
                col+=1

            #writes to the cell in given row,col
            status,err_msg=self.__write_to_cell_xls(input_path,workbook_info[0],sheetname,row,col,content)
        except Exception as e:
            err_msg='Error occurred writing to .xls file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,err_msg



class ExcelXLSX:

    def __init__(self):


        self.excel_obj=''

        self.cell_type={'string' : 'General',
        'formula': 'f',
        'number': '0.00',
        'date':'n',
        'boolean': 'b',
        'null':'n',
        'inline': 'inlineStr',
        'error': 'e',
        'formula_cache': 'str'}

        self.date_formats={'[$-14009]dd\ mmmm\ yyyy;@':'%d %B %Y',
            'mm-dd-yy':'%d-%m-%Y',
            '[$-14009]d\.m\.yy;@':'%d.%m.%y',
            'dd\/mm\/yyyy':'%d/%m/%y',
            '[$-F800]dddd\,\ mmmm\ dd\,\ yyyy':'%d %B %Y',
            '[$-14009]dd/mm/yyyy;@':'%d-%m-%Y',
            '[$-14009]dd/mm/yy;@':'%d-%m-%y',
            '[$-14009]d/m/yy;@':'%d-%m-%y',
            '[$-14009]yyyy/mm/dd;@':'%Y-%m-%d',
            '[$-14009]d\ mmmm\ yyyy;@':'%d %B %Y',
            'mmm/yyyy':'%b-%Y',
            }

    def __load_workbook_xlsx(self,inputpath,sheetname):
        """
        def : __load_workbook_xlsx(private method)
        purpose : loads the given .xlsx file
        param : inputpath,sheetname
        return : list

        """
        book=load_workbook(inputpath)
        sheet=book.get_sheet_by_name(sheetname)
        row_list=list(sheet.iter_rows())
        row=row_list[0]
        cell=row[0]
        if sheet.max_row==1 and cell.value is None:
            last_row=1
            last_col=1
        else:
            last_row=sheet.max_row
            row=row_list[len(row_list)-1]
            last_col=len(row)+1

        return book,sheet,last_row,last_col,sheet.max_row


    def __write_to_cell_xlsx(self,input_path,book,sheetname,row,col,value,*args):
        """
        def : __write_to_cell_xlsx(private method)
        purpose : writes to the the cell in the given row and column of .xlsx file
        param : input_path,book,sheetname,row,col,value
        return : bool

        """
        status=False
        self.excel_obj=ExcelFile()
        err_msg=None
        flag=False
        try:
           sheet=book.get_sheet_by_name(sheetname)
           cell=sheet.cell(row=row,column=col)
           if len(args)>0 and args[0] is not None:
            type=args[0].lower()
            if type in self.cell_type.keys():
                type=self.cell_type[type]
                if type=='0.00':
                    cell.number_format=type
                    value=int(value)
                    flag=True
                elif type in ['f','b']:
                    flag=True
                    value='='+value
                else:
                    flag=True
                    value=str(value)
                cell.value=value
                status=True
            else:
                logger.print_on_console('Cell Type not supported')
           else:
                cell.value=value
                status=True
                flag = True
        except IOError:
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
            log.error(err_msg)
        except Exception as e:
            log.error(e)
            err_msg='Error writing to excel cell of .xlsx file'
            logger.print_on_console(err_msg)
        book.save(input_path)
        if flag and status:
            self.excel_obj.open_and_save_file(input_path)
        return status,err_msg

    def __get_formatted_date(self,val,fmt):
        from datetime import datetime
        if fmt in self.date_formats.keys():
            return val.strftime(self.date_formats[fmt])
        return None

    def get_linenumber_xlsx(self,input_path,sheetname,content):
        """
        def : get_linenumber_xlsx
        purpose : get the line number where content is present in given .xlsx file
        param  : input_path,sheetname,content
        return : Returns line number (list)

        """
        status=False
        line_number=None
        err_msg=None
        logger.print_on_console(generic_constants.INPUT_IS+input_path+' '+sheetname)
        try:
            book = load_workbook(input_path,data_only=True)
            sheet = book.get_sheet_by_name(sheetname)
            i=0
            value=[]
            for row in sheet.iter_rows():
                i+=1
                for cell in row:
                    if  cell.internal_value is not None and content == cell.internal_value:
                        value.append(i)
            line_number=value
            status=True
        except Exception as e:
           err_msg='Error getting line number in .xls'
           log.error(e)
           logger.print_on_console(err_msg)
        return status,line_number,err_msg

    def replace_content_xlsx(self,input_path,sheetname,existingcontent,replacecontent,*args):
        """
        def : replace_content_xlsx
        purpose : replace the 'existingcontent' by 'replacecontent' in given sheet of
                  given .xlsx file
        param : input_path,sheetname,existingcontent,replacecontent
        return : bool

        """
        logger.print_on_console(generic_constants.INPUT_IS+input_path+' '+sheetname)
        status=False
        try:
            book = load_workbook(input_path,data_only=True)
            sheet = book.get_sheet_by_name(sheetname)
            for row in sheet.iter_rows():
                for cell in row:
                    if existingcontent ==  str(cell.value):
                        cell.value=replacecontent
                        status=True
            book.save(input_path)

        except Exception as e:
           log.error(e)
           logger.print_on_console(e)
        return status




    def delete_row_xlsx(self,row,excel_path,sheetname):
        """
        def : delete_row_xlsx
        purpose : calls delete_row_xls to delete a row in given .xlsx file
        param : row
        return : bool

        """
        #same functionality as compare_content_xlsx
        obj=ExcelXLS()
        return obj.delete_row_xls(row,excel_path,sheetname)




    def compare_content_xlsx(self,input_path1,sheetname1,input_path2,sheetname2,*args):
        """
        def : compare_content_xlsx
        purpose : calls compares the data of given sheets of 2 different excel files
        param : input_path1,input_path2,Sheet1,Sheet2
        return : bool

        """
        logger.print_on_console(generic_constants.INPUT_IS+input_path1+' '+input_path2)
        obj=ExcelXLS()
        #same functionality as compare_content_xlsx
        status=obj.compare_content_xls(input_path1,sheetname1,input_path2,sheetname2)
        return status






    def verify_content_xlsx(self,input_path,sheetname,content,*args):
        """
        def : verify_content_xlsx
        purpose : verify whether the 'existingcontent' is present in given sheet of
                  given .xlsx file
        param : input_path,sheetname,content
        return : bool

        """
        status=False
        res,line_number,err_msg=self.get_linenumber_xlsx(input_path,sheetname,content)
        if res and line_number != None:
            status=True
        return status,err_msg



    def clear_content_xlsx(self,inputpath,filename,sheetname):
        """
        def : clear_content_xlsx
        purpose : removes the given sheet from the given .xlsx file
        param : input_path,sheetname
        return : bool

        """
        status=False
        err_msg=None
        info_msg=generic_constants.INPUT_IS+inputpath+' filename: '+filename+' Sheetname: '+sheetname
        logger.print_on_console(info_msg)
        log.info(info_msg)
        inputpath=inputpath+'/'+filename
        try:
            book=load_workbook(inputpath)
            sheet=book.get_sheet_by_name(sheetname)
            book.remove(sheet)
            book.active=len(book.sheetnames)-1
            book.save(inputpath)
            status=True
        except Exception as e:
            err_msg='Error occured in clearing the content of excel sheet'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,err_msg



    def read_cell_xlsx(self,row,col,excel_path,sheetname,*args):
        """
        def : read_cell_xlsx
        purpose : reads the cell value in the given row and column of .xls file
        param : input_path,sheetname
        return : cell value

        """
        status=False
        value=None
        err_msg=None
        wb=None
        try:
            log.debug('Writing to the file'+generic_constants.INPUT_IS+excel_path+' '+sheetname)
            wb=openpyxl.load_workbook(excel_path,data_only=True)
            sheet=wb.get_sheet_by_name(sheetname)
            value=None
            if row<=sheet.max_row:
                if col<=sheet.max_column:
                    cell=sheet.cell(row=row,column=col)
                    value=cell.value
                    if cell.is_date == True and self.__get_formatted_date(value,cell.number_format) is not None :
                        value= self.__get_formatted_date(value,cell.number_format)
                    status=True
                else:
                    err_msg=ERROR_CODE_DICT["ERR_COL_DOESN'T_EXIST"]
            else:
                err_msg=ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
        except Exception as e:
            err_msg='Error occured in read cell of .xlsx file'
            log.error(e)

        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,value,err_msg

    def clear_cell_xlsx(self,row,col,excel_path,sheetname):
        """
        def : clear_cell_xls
        purpose : clears the given cell of .xls file set by the user
        param : row,col
        return : bool

        """
        status=False
        err_msg=None
        log.debug(generic_constants.INPUT_IS+excel_path+' '+sheetname)
        workbook=openpyxl.load_workbook(excel_path)
        try:
            sheet=workbook.get_sheet_by_name(sheetname)
            cell=sheet.cell(row=row,column=col,value=None)
            cell.value=None
            workbook.save(excel_path)
            status=True
        except Exception as e:
            err_msg='Error occurred in clearing cell of .xlsx file'
        workbook.save(excel_path)
        return status,err_msg




    def write_cell_xlsx(self,row,col,value,excel_path,sheetname,*args):
        """
        def : write_cell_xlsx
        purpose : calls the private methods to write given value to the cell of given
                row or col of .xlsx file
        param : row,col,value
        return : bool

        """
        status=False
        err_msg=None

        try:
            #loads the xls workbook
            workbook_info=self.__load_workbook_xlsx(excel_path,sheetname)
            #writes to the cell in given row,col
            if not(workbook_info[0].read_only):
                status,err_msg=self.__write_to_cell_xlsx(excel_path,workbook_info[0],sheetname,int(row),int(col),value,*args)
            else:
                logger.print_on_console('Excel is readonly')
                log.info('Excel is readonly')
        except Exception as e:
            log.error(e)
            logger.print_on_console(e)
        return status,err_msg



    def get_rowcount_xlsx(self,excel_path,sheetname,*args):
        """
        def : get_rowcount_xlsx
        purpose : get the number of rows of .xlsx excel file set by the user
        param : None
        return : Returns row count(int)

        """
        status=False
        row_count=None
        err_msg=None
        try:
            log.debug('Fetching row count of '+generic_constants.INPUT_IS+excel_path+' '+sheetname)
            book=openpyxl.load_workbook(excel_path)
            sheet=book.get_sheet_by_name(sheetname)
            row_count=sheet.max_row
            status=True
        except Exception as e:
            err_msg='Error getting the row count of .xlsx file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,row_count,err_msg



    def get_colcount_xlsx(self,excel_path,sheetname,*args):
        """
        def : get_colcount_xls
        purpose : get the number of columns of .xls excel file set by the user
        param : None
        return : Returns row count(int)

        """
        status=False
        col_count=None
        err_msg=None
        try:
            log.debug('Fetching col count of '+generic_constants.INPUT_IS+excel_path+' '+sheetname)
            book=openpyxl.load_workbook(excel_path)
            sheet=book.get_sheet_by_name(sheetname)
            col_count=sheet.max_column
            status=True
        except Exception as e:
            err_msg='Error getting the col count of .xlsx file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,col_count,err_msg

    def write_to_file_xlsx(self,input_path,sheetname,content,*args):
        """
        def : write_to_file_xlsx
        purpose : writes the given content to the given .xlsx file
        param : inputpath,sheetname,content,('newline'-optional param)
        return : bool

        """

        status=False
        err_msg=None
        log.debug(generic_constants.INPUT_IS+input_path+' '+sheetname)
        try:
            #loads the xls workbook
            workbook_info=self.__load_workbook_xlsx(input_path,sheetname)
            row=workbook_info[2]
            col=workbook_info[3]
            if len(args)>0 and args[0].lower() == 'newline':
                row=row+1
                col=1

            #writes to the cell in given row,col
            status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,content)
        except Exception as e:
            log.error(e)
            err_msg='Error occurred writing to .xlsx file '
            logger.print_on_console(err_msg)
        return status,err_msg


