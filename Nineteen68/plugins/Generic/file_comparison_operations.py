#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      sushma.p
#
# Created:     29-09-2016
# Copyright:   (c) sushma.p 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import logger
import generic_constants

class ExcelFile:


    import xlwt
    import openpyxl
    import xlsxwriter
    import xlutils

    def __init__(self):

        self.excel_path=''
        self.sheetname=''

    def get_linenumber_xls(self,input_path,content):
        print 'pass'

    def get_linenumber_xlsx(self,input_path,content):
        print 'pass'


    def write_to_file_xls(self,input_path,content):
       print 'xls'

    def write_to_file_xlsx(self,input_path,content):
        print 'xlsx'

    def delete_row_xls(self,row):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)


    def delete_row_xlsx(self,row):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)


    def compare_content_xls(self,input_path1,input_path2,*args):
        logger.log(generic_constants.INPUT_IS+input_path1+' '+input_path2)

    def compare_content_xlsx(self,input_path1,input_path2,*args):
        logger.log(generic_constants.INPUT_IS+input_path1+' '+input_path2)

    def set_excel_path(self,input_path,sheetname):
        logger.log(generic_constants.INPUT_IS+input_path+' '+sheetname)
        self.excel_path=input_path
        self.sheetname=sheetname
        return True

    def clear_excel_path(self):
        self.excel_path=''
        self.sheetname=''
        return True

    def read_cell_xls(self,row,col,*args):
        import datetime
        import xlrd
        value=None
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        from xlrd import open_workbook
        book = open_workbook(self.excel_path)
        sheet = book.sheet_by_name(self.sheetname)
        if row<sheet.nrows and col<sheet.ncols:
            cell=sheet.cell(row-1,col-1)
            if cell.ctype==3:
                value=datetime.datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))
                value=value.date()
            else:
                value=cell.value
            logger.log(value)
        else:
            logger.log(generic_constants.INDEX_EXCEEDS)

        return value



    def read_cell_xlsx(self,row,col,*args):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        wb=openpyxl.load_workbook(self.excel_path,read_only=true,data_only=True)
        sheet=wb.get_sheet_by_name(self.sheetname)
        sheet.get_index()
        cell=sheet.cell(row=row,column=col,value=None)
        print cell.internal_value
        print cell.value
        return cell.value


    def write_cell_xls(self,row,col,value):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        from xlutils.copy import copy
        book = open_workbook(self.excel_path)
        workbook = copy(book)
        sheets=book.sheet_names()
        try:
            sheetnum=sheets.index(self.sheetname)
            s = workbook.get_sheet(sheetnum)
            s.write(row,col,value)
            workbook.save(input_path)
        except ValueError:
            Exception.message('Value Error occured')

    def write_cell_xlsx(self,row,col,value):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        workbook=openpyxl.load_workbook(self.excel_path)
        sheet=workbook.get_sheet_by_name(self.sheetname)
        cell=sheet.cell(row=row,column=col,value=None)
        cell.value=value
        workbook.save(self.excel_path)


    def clear_cell_xls(self,row,col):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        from xlutils.copy import copy
        book = open_workbook(self.excel_path)
        workbook = copy(book)
        sheets=book.sheet_names()
        try:
            sheetnum=sheets.index(self.sheetname)
            s = workbook.get_sheet(sheetnum)
            s.write(row,col,'')
            workbook.save(input_path)
        except ValueError:
            Exception.message('Value Error occured')

    def clear_cell_xlsx(self,row,col):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        workbook=openpyxl.load_workbook(self.excel_path)
        sheet=workbook.get_sheet_by_name(self.sheetname)
        cell=sheet.cell(row=row,column=col,value=None)
        cell.value=None
        workbook.save(self.excel_path)


    def get_rowcount_xls(self,*args):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        book = open_workbook(self.excel_path)
        sheet=book.sheet_by_name(self.sheetname)
        rowcount=sheet.nrows
        return rowcount

    def get_rowcount_xlsx(self,*args):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        book=openpyxl.load_workbook(self.excel_path)
        sheet=book.get_sheet_by_name(self.sheetname)
        rowcount=sheet.max_row
        return rowcount

    def get_colcount_xls(self,*args):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        book = open_workbook(self.excel_path)
        sheet=book.sheet_by_name(self.sheetname)
        colcount=sheet.ncols
        return colcount

    def get_colcount_xlsx(self,*args):
        logger.log(generic_constants.INPUT_IS+self.excel_path+' '+self.sheetname)
        book=openpyxl.load_workbook(self.excel_path)
        sheet=book.get_sheet_by_name(self.sheetname)
        colcount=sheet.max_column
        return colcount


class PdfFile:

    def verify_content(self,input_path,pagenumber,content):
        pdf_content=self.get_content(input_path,pagenumber,'_internal_verify_content')
        if pdf_content == content:
            return True
        return False

    def compare_content(self,input_path1,input_path2):
        print 'pass'


    def get_content(self,input_path,pagenumber,output_file,*args):
        from PyPDF2 import PdfFileReader, PdfFileWriter
        try:
             reader=PdfFileReader(open(input_path,'rb'))
             pagenumber=int(pagenumber)-1
             if pagenumber<reader.getNumPages():
                page=reader.getPage(pagenumber)
                content=page.extractText()
                content=content.encode('utf-8')
                logger.log('Content is:',content)
                if output_file=='_internal_verify_content':
                    return content
                if len(args) == 2:
                    start=args[0].strip()
                    end=args[1].strip()

                else:
                    with open(output_file,'w') as file:
                        file.write(content)
                return content
             else:
                logger.log(generic_constants.INVALID_INPUT)

        except ValueError as e:
            Exception.message(generic_constants.INVALID_INPUT)



class TextFile:

    def verify_content(self,input_path,content):
        with open(input_path) as myFile:
            for num, line in enumerate(myFile, 1):
                print line
                if content in line:
                    print 'found at line:'
                    break;


    def compare_content(self,input_path1,input_path2):
        content1=self.get_content(input_path1)
        content2=self.get_content(input_path2)
        logger.log('File1 content is '+content1)
        logger.log('File2 content is '+content2)
        if content1==content2:
            print 'pass'
            return True
        return False


    def clear_content(self,input_path,content):
        with open(input_path) as myFile:
            myFile.write('')
            myFile.close()


    def get_content(self,input_path):
        str=""
        with open(input_path) as myFile:
            str=myFile.read()
        logger.log(str)
        return str

    def get_linenumber(self,input_path,content):
        line_numbers=[]
        with open(input_path) as myFile:
            for num, line in enumerate(myFile, 1):
                if content in line:
                    print 'found at line:', num
                    line_numbers.append(num)
        logger.log(line_numbers)



    def replace_content(self,input_path,existing_content,replace_content):
        filecontent=''
        try:
            with open(input_path,'r') as myFile:
                filecontent=myFile.read()
                filecontent=filecontent.replace(existing_content,replace_content)
            with open(input_path, 'w') as file:
                    file.write(filecontent)
                    file.close()
            return True
        except:
            Exception.message('Error occurred')
            return False



    def write_to_file(self,input_path,content):
        logger.log('Writing to text file')
        try:
            with open(input_path, 'a') as file:
                file.write(content)
                file.close()
        except (OSError, IOError) as e:
            Exception.message('Cannot open file')


class XML:
    def write_to_file(self,input_path,content):
        logger.log('Writing to XML file')
        try:
            with open(input_path, 'a') as file:
                file.write(content)
        except (OSError, IOError) as e:
            Exception.message('Cannot open file')

class CSV:
    print 'pass'













