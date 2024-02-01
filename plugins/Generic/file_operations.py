#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      sushma.p
#
# Created:     29-09-2016
# Copyright:   (c) sushma.p 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import sys
import os
import logger
import xlwt
from xlutils.copy import copy as xl_copy
import openpyxl
from openpyxl.utils import get_column_letter
import generic_constants
from constants import *
import folder_operations
from file_comparison_operations import TextFile,XML,JSON
from file_operations_pdf import FileOperationsPDF
import excel_operations
import core_utils
import urllib.request, urllib.error, urllib.parse
import xml.dom.minidom
import json
import difflib
from xmldiff import main, formatting
import readconfig
from xlrd import open_workbook
import shutil
import pathlib
import dynamic_variable_handler
import constant_variable_handler
import logging
from itertools import islice
import re
import zipfile
import platform
import time
import docx
import urllib.request
import pyautogui
if SYSTEM_OS=='Windows':
    import win32gui
    from pywinauto.application import Application 


log = logging.getLogger('file_operations.py')


class FileOperations:

    """The instantiation operation __init__ creates an empty object of the class FileOperations when it is instantiated"""
    def __init__(self):
        self.txt=TextFile()
        self.pdf=FileOperationsPDF()
        self.xml=XML()
        self.json=JSON()
        self.folder=folder_operations.FolderOperations()
        self.xls_obj=excel_operations.ExcelXLS()
        self.xlsx_obj=excel_operations.ExcelXLSX()
        self.csv_obj=excel_operations.ExcelCSV()
        self.convert=excel_operations.ExcelFile()
        self.DV = dynamic_variable_handler.DynamicVariables()
        self.CV = constant_variable_handler.ConstantVariables()

        """Mapping of keywords to its respective methods"""
        self.dict={'.txt_write_to_file':self.txt.write_to_file,
              '.xls_write_to_file':self.xls_obj.write_to_file_xls,
              '.xlsx_write_to_file':self.xlsx_obj.write_to_file_xlsx,
              '.xml_write_to_file':self.xml.write_to_file,
              '.json_write_to_file':self.json.write_to_file,

              '.xls_verify_content':self.xls_obj.verify_content_xls,
              '.xlsx_verify_content':self.xlsx_obj.verify_content_xlsx,
              '.txt_verify_content':self.txt.verify_content,
              '.pdf_verify_content':self.pdf.verify_content,

              '.txt_compare_content':self.txt.compare_content,
              '.xls_compare_content':self.xls_obj.compare_content_xls,
              '.xlsx_compare_content':self.xlsx_obj.compare_content_xlsx,
              '.csv_compare_content':self.csv_obj.compare_content_csv,
              '.json_compare_content':self.json.compare_content,

              '.txt_clear_content':self.txt.clear_content,
              '.xls_clear_content':self.xls_obj.clear_content_xls,
              '.xlsx_clear_content':self.xlsx_obj.clear_content_xlsx,
              '.xml_clear_content':self.xml.clear_content,

              '.txt_get_content':self.txt.get_content,
              '.pdf_get_content':self.pdf.get_content,

              '.txt_replace_content':self.txt.replace_content,
              '.xls_replace_content':self.xls_obj.replace_content_xls,
              '.xlsx_replace_content':self.xlsx_obj.replace_content_xlsx,

              '.txt_get_line_number':self.txt.get_linenumber,
              '.xls_get_line_number':self.xls_obj.get_linenumber_xls,
              '.xlsx_get_line_number':self.xlsx_obj.get_linenumber_xlsx,

              '.xlsx_create_file':self.xlsx_obj.create_file_xlsx,
              '.xls_create_file':self.xls_obj.create_file_xls,

              '.pdf_get_page_count':self.pdf.get_page_count
              }

    def create_file(self,inputpath,file_name):
        """
        def : create_file
        purpose : creates the file in the given input path
        param : inputpath,file_name
        return : bool

        """
        try:

            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            err_msg=None
            output_res=OUTPUT_CONSTANT
            log.debug('reading the inputs')
            inputpath=inputpath.strip()
            log.debug(generic_constants.INPUT_IS+inputpath+' File name '+file_name)

            if not (input is None and input is '') and self.folder.validateFolderName(file_name) :
                if not os.path.isfile(inputpath + '/' + file_name):
                    log.debug('creating the file')
                    fullpath = inputpath + '/' + file_name
                    file_extension,status_get_ext = self.__get_ext(fullpath)
                    if status_get_ext and file_extension is not None and file_extension in generic_constants.EXCEL_TYPES:
                        status_excel_create_file,e_msg = self.dict[file_extension + '_create_file'](fullpath,'Sheet1')
                        if status_excel_create_file and e_msg is None:
                            status=TEST_RESULT_PASS
                            methodoutput=TEST_RESULT_TRUE
                        else:
                            err_msg = e_msg
                    else:
                        open(fullpath, 'w').close()
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
                else:
                    err_msg=generic_constants.FILE_EXISTS
            else:
                err_msg=generic_constants.INVALID_INPUT
        except Exception as e:
            log.error(e)
            err_msg=generic_constants.ERR_MSG1+'Creating'+generic_constants.ERR_MSG2
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output_res,err_msg


    def verify_file_exists(self,inputpath,file_name):
        """
        def : verify_file_exists
        purpose : verifies if file exists in the given input path
        param : inputpath,file_name
        return : bool

        """
        try:
            coreutilsobj=core_utils.CoreUtils()
            inputpath=coreutilsobj.get_UTF_8(inputpath)
            file_name=coreutilsobj.get_UTF_8(file_name)
            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            err_msg=None
            output_res=OUTPUT_CONSTANT
            log.debug('reading the inputs')
            inputpath=inputpath.strip()
            info_msg=generic_constants.INPUT_IS+inputpath
            log.info(info_msg)
            if inputpath!= None and inputpath!='':
                if file_name!= None and file_name!='':
                    log.info(' and File name is:'+file_name)
                    inputpath=inputpath+os.sep+file_name
                if os.path.isfile(inputpath):
                    log.debug(generic_constants.FILE_EXISTS)
                    status=TEST_RESULT_PASS
                    methodoutput=TEST_RESULT_TRUE
                else:
                    err_msg=ERROR_CODE_DICT['ERR_NO_SUCH_FILE_EXCEPTION']
            else:
                err_msg=generic_constants.INVALID_INPUT
        except (IOError,WindowsError):
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
        except Exception as e:
            err_msg=generic_constants.ERR_MSG1+'Verifying'+generic_constants.ERR_MSG2
            log.error(e)
        if err_msg:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output_res,err_msg

    def rename_file(self,inputpath,file_name,rename_file):
        """
        def : rename_file
        purpose : renames the file in the 'inputpath' by 'renamepath'
        param : inputpath,file_name
        return : bool

        """
        try:

            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            err_msg=None
            output_res=OUTPUT_CONSTANT
            log.debug('reading the inputs')
            inputpath=inputpath.strip()
            log.debug(generic_constants.INPUT_IS+inputpath+' File name '+file_name)
           #logger.print_on_console(generic_constants.INPUT_IS+inputpath+' File name '+file_name)
            if not (inputpath is None and inputpath is ''):
                rename_path=inputpath+'/'+rename_file
                inputpath=inputpath+'/'+file_name
                if os.path.isfile(inputpath):
                    log.debug('renaming the file')
                    os.rename(inputpath,rename_path)
                    status=TEST_RESULT_PASS
                    methodoutput=TEST_RESULT_TRUE
                else:
                    err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_FOUND_EXCEPTION']
            else:
                err_msg=generic_constants.INVALID_INPUT
        except (IOError,WindowsError):
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
        except Exception as e:
            err_msg=generic_constants.ERR_MSG1+'Renaming'+generic_constants.ERR_MSG2
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output_res,err_msg

    def saveFileAs(self,source_path,destination_path,opt,extension):
        import csv
        import pythoncom
        pythoncom.CoInitialize()
        status = TEST_RESULT_FAIL
        result = TEST_RESULT_FALSE
        err_msg = None
        fileFlag = False
        try:
            file1 = os.path.basename(source_path)
            if(not extension.startswith('.')): extension = '.'+extension    
            file2 = os.path.dirname(destination_path)+'\\'+file1.split('.')[0]+extension               
            res = os.path.isfile(file2)
            filename = os.path.basename(file2)
            input_ext = os.path.splitext(source_path)[1]
            ext_list=['.txt','.csv','.xlsx','.xls','.doc','.docx','.pdf','.zip','.unzip']
            if(extension in ext_list):
                if((input_ext in ['.xlsx', '.xls']) or (not res or opt=='1')): 
                    if(input_ext=='.txt' and extension=='.csv'):
                        with open(source_path, 'r') as in_file:
                            stripped = (line.strip() for line in in_file)
                            lines = (line.split(",") for line in stripped if line)
                            with open(file2, 'w',newline='') as out_file:
                                writer = csv.writer(out_file)
                                writer.writerows(lines)
                    
                    elif(input_ext=='.csv' and extension=='.txt'):
                        with open(file2, "w") as my_output_file:
                            with open(source_path, "r") as my_input_file:
                                [ my_output_file.write(",".join(row)+'\n') for row in csv.reader(my_input_file)]
                            my_output_file.close()

                    elif(input_ext=='.xlsx' and extension=='.csv'):
                        wb = openpyxl.load_workbook(source_path)
                        sh = wb.worksheets
                        for sheet in sh:
                            if(not len(sh)==1):
                                file2=destination_path.split('.')[0]+'_'+sheet.title+'.csv'
                            if(not os.path.isfile(file2) or opt=='1'):
                                with open(file2, 'w', newline="") as f:
                                    c = csv.writer(f)
                                    for r in sheet.rows:
                                        c.writerow([cell.value for cell in r])
                                wb.close()
                            else:
                                fileFlag=True
                        if(fileFlag):
                            err_msg = 'One or more files already exists with the same name in the destination path.'

                    elif(input_ext=='.xls' and extension=='.csv'):
                        with open_workbook(source_path) as wb:
                            sheet_name=wb.sheet_names()
                            for i in range(len(sheet_name)):
                                if(not len(sheet_name)==1):
                                    file2=destination_path.split('.')[0]+'_'+sheet_name[i]+'.csv'
                                if(not os.path.isfile(file2) or opt=='1'):
                                    sheet = wb.sheet_by_index(i)
                                    with open(file2, 'w', newline="") as f:
                                        c = csv.writer(f)
                                        for r in range(sheet.nrows):
                                            c.writerow(sheet.row_values(r))
                                else:
                                    fileFlag=True
                            if(fileFlag):
                                err_msg = 'One or more files already exists with the same name in the destination path.'

                    elif(input_ext=='.csv' and extension=='.xlsx'):
                        workbook = openpyxl.Workbook()
                        workbook.active.title = file1.split('.')[0]
                        worksheet = workbook.active
                        with open(source_path, 'rt', encoding='utf8') as f:
                            reader = csv.reader(f)
                            for r, row in enumerate(reader):
                                for c, col in enumerate(row):
                                    worksheet.cell(r+1, c+1, col)
                        workbook.save(os.path.join(os.path.dirname(file2), filename))
                        workbook.close()
                    
                    elif(input_ext=='.csv' and extension=='.xls'):
                        with open(source_path, 'rt') as f:
                            reader = csv.reader(f)
                            workbook = xlwt.Workbook()
                            sheet = workbook.add_sheet(file1.split('.')[0])
                            for rowi, row in enumerate(reader):
                                for coli, value in enumerate(row):
                                    sheet.write(rowi,coli,value)
                            workbook.save(file2)

                    elif((input_ext=='.xlsx' and extension=='.xls') or (input_ext=='.xls' and extension=='.xlsx')):
                        if SYSTEM_OS == 'Windows':
                            import win32com.client
                            excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
                            wb = excel.Workbooks.Open(source_path)
                            if(extension=='.xls'):
                                wb.SaveAs(file2, FileFormat = 56)
                            else:
                                wb.SaveAs(file2, FileFormat = 51)
                            wb.Close()
                            excel.Application.Quit()
                        else:
                            err_msg = 'File conversion support is not given for ' + str(input_ext) + ' to ' + str(extension) + ' in Mac OS'

                    elif((input_ext=='.docx' or input_ext=='.doc') and (extension=='.pdf' or extension=='.docx' or extension=='.doc')):
                        if SYSTEM_OS == 'Windows':
                            import win32com.client
                            word = win32com.client.Dispatch('Word.Application')
                            doc = word.Documents.Open(source_path)
                            if(extension=='.docx'):
                                doc.SaveAs(file2, FileFormat = 16)
                            elif(extension=='.doc'):
                                doc.SaveAs(file2, FileFormat = 0)
                            else:
                                doc.SaveAs(file2, FileFormat = 17)
                            doc.Close()
                            word.Quit()
                        else:
                            err_msg = 'File conversion support is not given for ' + str(input_ext) + ' to ' + str(extension) + ' in Mac OS'

                    elif(input_ext=='.pdf' and (extension=='.docx' or extension=='.doc')):
                        from pdf2docx import parse
                        parse(source_path, file2)

                    elif(extension=='.zip'):
                        zipf = zipfile.ZipFile(file2, 'w', zipfile.ZIP_DEFLATED)
                        if(input_ext):
                            zipf.write(source_path, file1)
                        else:
                            for root, dirs, files in os.walk(source_path):
                                for f in files:
                                    if(filename==f):
                                        continue
                                    else:
                                        zip_path1 = os.path.join(source_path, '..')
                                        zip_path2 = os.path.join(root, f)
                                        zipf.write(zip_path2, os.path.relpath(zip_path2, zip_path1))
                        zipf.close()

                    elif(input_ext=='.zip'):
                        unzip_path=os.path.dirname(destination_path)
                        with zipfile.ZipFile(source_path, 'r') as zip_ref:
                            zip_ref.extractall(unzip_path)
                        zip_ref.close()

                    else:
                        err_msg = 'File conversion support is not given for ' + str(input_ext) + ' to ' + str(extension)
                else:
                    err_msg = 'File/Folder already exists in the destination path'
                if(not err_msg):
                    log.debug('Successfully converted ' + str(source_path) + ' to ' + str(extension)+' extension.') 
                    logger.print_on_console('Successfully converted ' + str(source_path) + ' to ' + str(extension)+' extension.')
                    status=TEST_RESULT_PASS
                    result=TEST_RESULT_TRUE
            else:
                err_msg = 'Supported extensions for conversions are '+ str(ext_list)
        except Exception as e:
            log.error(e)
            err_msg = 'Error occurred while file conversion in copyFileFolder'
        return status, result, err_msg
            
    def copyFileFolder(self, source_path, destination_path, opt = 0, extension = 'null'):
        """
        def: copyFileFolder
        purpose: Copy from input path 1, to input path 2, optional(0(default)-don't overwrite/1- overwrite), extension(optional)
        param: <path of source file/folder>,<path of destination(directory path only)>,bool,extension(to convert file to other formats)
        return: bool
        """
        status = TEST_RESULT_FAIL
        result = TEST_RESULT_FALSE
        err_msg = None
        output_res = OUTPUT_CONSTANT
        try:
            if(extension.strip() == ''): extension='null'
            if(not(isinstance(opt, int)) and opt.strip() == ''): opt=0
            if( source_path and destination_path ):
                source_path=os.path.normpath(source_path)
                destination_path=os.path.normpath(destination_path)
                try:
                    if( os.path.splitext(source_path)[1] and not os.path.splitext(destination_path)[1] ):
                        #file ops
                        if ( os.path.isfile(source_path) ):
                            if( os.path.isdir(destination_path) ):# #destination should be a folder only
                                if(extension=='null'):
                                    destination_path = destination_path + '/' + os.path.basename(source_path)
                                elif(extension.startswith('.')):
                                    destination_path = destination_path + '/' + os.path.basename(source_path).split('.')[0]+extension
                                else:
                                    destination_path = destination_path + '/' + os.path.basename(source_path).split('.')[0]+'.'+extension
                                if ( shutil.disk_usage(os.path.dirname(destination_path)).free / 1048576 > os.path.getsize(source_path) / 1048576):
                                    if( int(opt) == 0 ):
                                        #dont-overwrite
                                        if ( not os.path.isfile(destination_path) or extension!='null'): #already exists
                                            logger.print_on_console( "Copying the file to destination. Please wait..." )
                                            if( extension!= 'null'):
                                                status,result,err_msg=self.saveFileAs(source_path,destination_path,opt,extension)
                                            else:
                                                shutil.copyfile(source_path,destination_path)
                                                log.debug( 'File successfully copied from ' + str(source_path) + ' to ' + str(destination_path) )                                             
                                                status=TEST_RESULT_PASS
                                                result=TEST_RESULT_TRUE
                                        else:
                                            err_msg = 'File already exists in the directory'
                                    elif( int(opt) == 1 ):
                                        #overwrite
                                        logger.print_on_console( "Copying the file to destination. Please wait..." )
                                        if( extension!= 'null'):
                                            status,result,err_msg=self.saveFileAs(source_path,destination_path,opt,extension)
                                        else:
                                            shutil.copyfile(source_path,destination_path)
                                            log.debug( 'File successfully copied from ' + str(source_path) + ' to ' + str(destination_path) )
                                            status=TEST_RESULT_PASS
                                            result=TEST_RESULT_TRUE
                                    else:
                                        err_msg = 'Invalid option, Please provide 1 or 0'
                                else:
                                    log.error('Available space on destination directory : ' + str(shutil.disk_usage(os.path.dirname(destination_path)).free / 1048576) + 'MB and file size of source path : ' + str(os.path.getsize(source_path) / 1048576) + 'MB')
                                    err_msg = 'Destination directory does not have sufficient storage space'
                            else:
                                err_msg = 'Destination directory does not exist, Please provide a valid destination directory'
                        else:
                            err_msg = 'Source file does not exist, Please provide a valid source file'
                    elif( not os.path.splitext(source_path)[1] and not os.path.splitext(destination_path)[1] ):
                        #folder ops
                        if( os.path.isdir(source_path) ):
                            if (os.path.isdir(destination_path)):
                                if(extension=='null'):
                                    destination_path = destination_path + '/' + os.path.basename(source_path)
                                elif(extension.startswith('.')):
                                    destination_path = destination_path + '/' + os.path.basename(source_path).split('.')[0]+extension
                                else:
                                    destination_path = destination_path + '/' + os.path.basename(source_path).split('.')[0]+'.'+extension
                                if ( shutil.disk_usage(os.path.dirname(destination_path)).free / 1000000 > os.path.getsize(source_path)):
                                    if( int(opt) == 0 ):
                                        #dont - overwrite
                                        if( not os.path.exists(destination_path) ):
                                            logger.print_on_console( "Copying the directory to destination. Please wait..." )
                                            if( extension!= 'null'):
                                                status,result,err_msg=self.saveFileAs(source_path,destination_path,opt,extension)
                                            else:
                                                shutil.copytree(source_path, destination_path)
                                                log.debug( 'Folder successfully copied to ' + str(source_path) + ' from ' + str(destination_path) )
                                                status=TEST_RESULT_PASS
                                                result=TEST_RESULT_TRUE
                                        else:
                                            err_msg = 'Folder already exists in the directory'
                                    elif( int(opt) == 1 ):
                                        #overwrite
                                            logger.print_on_console( "Copying the directory to destination. Please wait..." )
                                            if( extension!= 'null'):
                                                status,result,err_msg=self.saveFileAs(source_path,destination_path,opt,extension)
                                            else:
                                                shutil.copytree(source_path, destination_path)
                                                log.debug( 'Folder successfully copied to ' + str(source_path) + ' from ' + str(destination_path) )
                                                status=TEST_RESULT_PASS
                                                result=TEST_RESULT_TRUE
                                    else:
                                        err_msg = 'Invalid option, Please provide 1 or 0'
                                else:
                                    log.error('Available space on destination directory : ' + str(shutil.disk_usage(os.path.dirname(destination_path)).free / 1000000) + 'MB and folder size of source path : ' + str(os.path.getsize(source_path)) + 'MB')
                                    err_msg = 'Destination directory does not have sufficient storage space'
                            else:
                                err_msg = 'Destination directory does not exist, Please provide a valid destination directory'
                        else:
                            err_msg = 'Source directory does not exist, Please provide a valid source directory'
                    else:
                        err_msg = generic_constants.INVALID_INPUT
                except FileExistsError as e:
                    print(e)
                    err_msg = 'Source and destination represents the same folder.'
                except NotADirectoryError:
                    err_msg = 'Invalid directory'
                except IsADirectoryError:
                    err_msg = 'Path is a directory'
                except shutil.SameFileError:
                    err_msg = 'Source and destination represents the same file.'
                except PermissionError:
                    err_msg = 'Permission denied. Check: 1.If file has valid permission 2.Source path should not end with "\\"'
                except FileNotFoundError as e:
                    if ( ']' in str(e) ): err_msg = str(e)[str(e).index(']')+2:]
                    else: err_msg = 'FileNotFoundError : ' + str(e)
                except IOError as e:
                    if ('[Errno 28] No space left on device' in str(e)):
                        err_msg = 'No space left on device'
                        if( os.path.isdir( destination_path ) ): shutil.rmtree( destination_path )
                        elif( os.path.isfile( destination_path ) ): os.remove( destination_path )
                    else: err_msg = 'IOError : ' + str(e)
            else:
                err_msg = generic_constants.INVALID_INPUT
        except ValueError:
            err_msg = 'Invalid option, Please provide 1 or 0'
        except Exception as e:
            err_msg = 'Error occurred in copyFileFolder'
            log.error(e)
        if err_msg:
            log.error( err_msg )
            logger.print_on_console( err_msg )
        del source_path, destination_path, opt #deleting variables
        return status, result, output_res, err_msg

    def moveFileFolder(self, source_path, destination_path, opt = 0, extension = 'null'):
        """
        def: moveFileFolder
        purpose: Copy from input path 1, to input path 2, optional(0(default)- overwrite/1- overwrite ), extension(optional)
        param: <path of source file>,<path of destination(full file path)>,bool,extension(to convert file to other formats)
        return: bool
        """
        status = TEST_RESULT_FAIL
        result = TEST_RESULT_FALSE
        err_msg = None
        output_res = OUTPUT_CONSTANT
        try:
            if(extension.strip() == ''): extension='null'
            if(not(isinstance(opt,int)) and opt.strip() == ''): opt=0
            if( source_path and destination_path ):
                try:
                    if( os.path.splitext(source_path)[1] and not os.path.splitext(destination_path)[1] ):
                        #file ops
                        if ( os.path.isfile(source_path) ):
                            if( os.path.isdir(destination_path) ):# #destination should be a folder only
                                if(extension=='null'):
                                    destination_path = destination_path + '/' + os.path.basename(source_path)
                                elif(extension.startswith('.')):
                                     destination_path = destination_path + '/' + os.path.basename(source_path).split('.')[0]+extension
                                else:
                                    destination_path = destination_path + '/' + os.path.basename(source_path).split('.')[0]+'.'+extension
                                if( int(opt) == 0 ):
                                    if ( not os.path.isfile(destination_path) or extension!='null'):
                                        logger.print_on_console( "Moving the file to destination. Please wait..." )
                                        if(extension!='null'):
                                            status,result,err_msg=self.saveFileAs(source_path,destination_path,opt,extension)
                                        else:
                                            shutil.move(source_path,destination_path)
                                            log.debug('File successfully moved from ' + str(source_path) + ' to ' + str(destination_path))
                                            status=TEST_RESULT_PASS
                                            result=TEST_RESULT_TRUE
                                        if ( os.path.isfile(source_path) ):#delete the file if still present in path(shutil only copies if location is on another disk)
                                            pathlib.Path(source_path).unlink()
                                    else:
                                        err_msg = 'File already exists in the directory'
                                elif( int(opt) == 1 ):
                                    #----- if file exists then overwrite
                                    if os.path.isfile(destination_path):
                                        log.info( 'File already exists in destination path, performing overwrite by deletion' )
                                        os.remove(destination_path)
                                    #----- if file exists then overwrite
                                    logger.print_on_console( "Moving the file to destination. Please wait..." )
                                    if(extension!='null'):
                                        status,result,err_msg=self.saveFileAs(source_path,destination_path,opt,extension)
                                    else:
                                        shutil.move(source_path,destination_path)
                                        log.debug('File successfully moved from ' + str(source_path) + ' to ' + str(destination_path))
                                        status=TEST_RESULT_PASS
                                        result=TEST_RESULT_TRUE
                                    if (os.path.isfile(source_path)):#delete the file if still present in path(shutil only copies if location is on another disk)
                                        pathlib.Path(source_path).unlink()
                                else:
                                    err_msg = 'Invalid option, Please provide 1 or 0'
                            else:
                                err_msg = 'Destination directory does not exist, Please provide a valid destination directory'
                        else:
                            err_msg = 'Source file does not exist, Please provide a valid source file'
                    elif( not os.path.splitext(source_path)[1] and not os.path.splitext(destination_path)[1] ):
                        #folder ops
                        if( os.path.isdir(source_path) ):
                            if (os.path.isdir(destination_path)):
                                if(extension=='null'):
                                    destination_path = destination_path + '/' + os.path.basename(source_path)
                                elif(extension.startswith('.')):
                                    destination_path = destination_path + '/' + os.path.basename(source_path).split('.')[0]+extension
                                else:
                                    destination_path = destination_path + '/' + os.path.basename(source_path).split('.')[0]+'.'+extension
                                if( int(opt) == 0 ):
                                    if ( not os.path.exists(destination_path) ):
                                        logger.print_on_console( "Moving the folder to destination. Please wait..." )
                                        if(extension!='null'):
                                            status,result,err_msg=self.saveFileAs(source_path,destination_path,opt,extension)
                                        else:
                                            shutil.move(source_path,destination_path)
                                            log.debug('Folder successfully moved to ' + str(source_path) + ' from ' + str(source_path))
                                            status=TEST_RESULT_PASS
                                            result=TEST_RESULT_TRUE
                                        if (os.path.isdir(source_path)):#delete the folder if still present in path(shutil only copies if location is on another disk)
                                            shutil.rmtree(source_path)
                                    else:
                                        err_msg = 'Folder already exists in the directory'
                                elif( int(opt) == 1 ):
                                    if ( os.path.isdir(destination_path)):
                                        shutil.rmtree(destination_path)
                                    logger.print_on_console( "Moving the folder to destination. Please wait..." )
                                    if(extension!='null'):
                                        status,result,err_msg=self.saveFileAs(source_path,destination_path,opt,extension)
                                    else:
                                        shutil.move(source_path,destination_path)
                                        log.debug('Folder successfully moved to ' + str(source_path) + ' from ' + str(source_path))
                                        status=TEST_RESULT_PASS
                                        result=TEST_RESULT_TRUE
                                    if (os.path.isdir(source_path)):#delete the folder if still present in path(shutil only copies if location is on another disk)
                                        shutil.rmtree(source_path)
                                else:
                                    err_msg = 'Invalid option, Please provide 1 or 0'
                            else:
                                err_msg = 'Destination directory does not exist, Please provide a valid destination directory'
                        else:
                            err_msg = 'Source directory does not exist, Please provide a valid source directory'
                    else:
                        err_msg = generic_constants.INVALID_INPUT
                except FileExistsError as e:
                    print(e)
                    err_msg = 'Source and destination represents the same folder.'
                except NotADirectoryError:
                    err_msg = 'Invalid directory'
                except IsADirectoryError:
                    err_msg = 'Path is a directory'
                except shutil.SameFileError:
                    err_msg = 'Source and destination represents the same file.'
                except PermissionError:
                    err_msg = 'Permission denied.'
                except FileNotFoundError as e:
                    if ( ']' in str(e) ): err_msg = str(e)[str(e).index(']')+2:]
                    else: err_msg = 'FileNotFoundError : ' + str(e)
                except IOError as e:
                    if ('[Errno 28] No space left on device' in str(e)):
                        err_msg = 'No space left on device'
                        if( os.path.isdir( destination_path ) ): shutil.rmtree( destination_path )
                        elif( os.path.isfile( destination_path ) ): os.remove( destination_path )
                    else: err_msg = 'IOError : ' + str(e)
            else:
                err_msg = generic_constants.INVALID_INPUT
        except ValueError:
            err_msg = 'Invalid option, Please provide 1 or 0'
        except Exception as e:
            err_msg = 'Error occurred in moveFileFolder'
            log.error(e)
        if err_msg:
            logger.print_on_console( err_msg )
            log.error(err_msg)
        del source_path, destination_path, opt #deleting variables
        return status, result, output_res, err_msg

    def delete_file(self,inputpath,file_name):
        """
        def : delete_file
        purpose : deletes the file in the input path
        param : inputpath,file_name
        return : bool

        """
        try:

            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            err_msg=None
            output_res=OUTPUT_CONSTANT
            log.debug('reading the inputs')
            log.debug(generic_constants.INPUT_IS+inputpath+' File name '+file_name)
            #logger.print_on_console(generic_constants.INPUT_IS+inputpath+' File name '+file_name)
            if not (input is None and input is ''):
                if os.path.isfile(inputpath+'/'+file_name):
                    log.debug('removing the file')
                    os.remove(inputpath+'/'+file_name)
                    status=TEST_RESULT_PASS
                    methodoutput=TEST_RESULT_TRUE
                else:
                    err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_FOUND_EXCEPTION']
            else:
                err_msg=generic_constants.INVALID_INPUT
        except (IOError,WindowsError):
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
        except Exception as e:
            err_msg=generic_constants.ERR_MSG1+'Deleting'+generic_constants.ERR_MSG2
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output_res,err_msg


    def verify_content(self,input_path,content,*args):
        """
        def : verify_content
        purpose : calls the respective method to verify the given content based on file type
        param : input_path,content
        return : bool

        """
        try:

            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            err_msg=None
            output_res=OUTPUT_CONSTANT
            log.debug('reading the inputs')
            params=self.__split(input_path,content,*args)
            result=self.verify_file_exists(params[0],'')
            if result[1] == TEST_RESULT_TRUE:
                file_ext,res=self.__get_ext(params[0])
                if res == True:
                    log.debug('verifying the contents')
                    res,err_msg= self.dict[file_ext+'_verify_content'](*params)
                    if res:
                        log.debug('content present in the given file')
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
            else:
                err_msg=result[3]
        except Exception as e:
            err_msg=generic_constants.ERR_MSG1+'Verifying content of'+generic_constants.ERR_MSG2
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output_res,err_msg


    def __get_ext(self,input_path):
        """
        def : __get_ext
        purpose : returns the file type and verifies if it is valid file type
        param : input_path
        return : bool,filetype

        """
        try:
            status=False
            filename,file_ext=os.path.splitext(input_path)
            if file_ext.lower() in generic_constants.FILE_TYPES:
                status=True
            else:
                log.info(generic_constants.INVALID_FILE_FORMAT)
                err_msg=generic_constants.INVALID_FILE_FORMAT
            return file_ext,status
        except Exception as e:
            log.error(e)
            logger.print_on_console(e)
            err_msg=INPUT_ERROR
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return '',status

    def compare_content(self,input_path1,input_path2,*args):
        """
        def : verify_content
        purpose : calls the respective method to compare the content of 2 files based on file type
        param : input_path1,input_path2
        return : bool

        """
        try:

            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            err_msg=None
            output_res=OUTPUT_CONSTANT
            log.debug('reading the inputs')
            input_path1 = str(input_path1)
            input_path2 = str (input_path2)
            input_path1 = input_path1.replace(',',';')
            input_path2 = input_path2.replace(',',';')
            params=self.__split(input_path1,input_path2,*args)
            path2=params[1]
            if len(params)>2:
                path2=params[2]
            log.debug('verifying whether the files exists')
            result1=self.verify_file_exists(params[0],'')
            result2=self.verify_file_exists(path2,'') if path2!='' else TEST_RESULT_FAIL
            if result1[1] == TEST_RESULT_TRUE and result2[1]==TEST_RESULT_TRUE :
                file_ext1,status1=self.__get_ext(params[0])
                file_ext2,status2=self.__get_ext(path2)
                log.debug('comparing the contents')
                if status1 == True and status2==True and file_ext1==file_ext2:
                    res,err_msg=self.dict[file_ext1+'_compare_content'](*params)
                    if res:
                        log.info('files contents are same')
                        logger.print_on_console('files contents are same')
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
                else:
                    err_msg=ERROR_CODE_DICT['ERR_FILE_EXT_MISMATCH']
            else:
                err_msg=result1[3]
                if err_msg==None:
                    err_msg=result2[3]
            if type(result2) != list and result2==TEST_RESULT_FAIL:
                err_msg=generic_constants.INVALID_INPUT
        except TypeError as e:
            err_msg=ERROR_CODE_DICT['ERR_INDEX_OUT_OF_BOUNDS_EXCEPTION']
        except Exception as e:
            err_msg=generic_constants.ERR_MSG1+'Comapring content of'+generic_constants.ERR_MSG2
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output_res,err_msg

    def clear_content(self,input_path,file_name,*args):
        """
        def : clear_content
        purpose : calls the respective method to clear the content of given file based on file type
        param : input_path
        return : bool

        """
        try:

            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            err_msg=None
            output_res=OUTPUT_CONSTANT
            log.debug('reading the inputs')
            params=self.__split(input_path,file_name,*args)
            log.debug('verifying whether the files exist')
            result=self.verify_file_exists(params[0],params[1])
            if result[1]==TEST_RESULT_TRUE:
                file_ext,res=self.__get_ext(params[1])
                if res == True:
                    log.debug('clearing the content of file')
                    res,err_msg=self.dict[file_ext+'_clear_content'](*params)
                    if res:
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
        except TypeError as e:
            err_msg=ERROR_CODE_DICT['ERR_INDEX_OUT_OF_BOUNDS_EXCEPTION']
        except Exception as e:
            err_msg=generic_constants.ERR_MSG1+'Clearing'+generic_constants.ERR_MSG2
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output_res,err_msg

    def get_content(self,input_path,*args):
        """
        def : get_content
        purpose : calls the respective method to get the content of given file based on file type
        param : input_path
        return :string

        """
        try:

            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            content=None
            err_msg=None

            log.debug('reading the inputs')
            params=self.__split(input_path,*args)
            log.debug('verifying whether the files exists')
            result=self.verify_file_exists(params[0],'')
            if result[1]==TEST_RESULT_TRUE:
                file_ext,res=self.__get_ext(params[0])
                if file_ext not in ['.xls','.xlsx']:
                    if res == True:
                        log.debug('clearing the content of file')
                        if len(params)==1:
                            params.append("getfullcontent")
                            res,content,err_msg=self.dict[file_ext+'_get_content'](*params)
                        else:
                            if params[1] in ["", " "]:
                                params[1] = "nolinenumber"
                                res,content,err_msg=self.dict[file_ext+'_get_content'](*params)
                            else:
                                res,content,err_msg=self.dict[file_ext+'_get_content'](*params)
                        if res:
                            status=TEST_RESULT_PASS
                            methodoutput=TEST_RESULT_TRUE
                    else:
                        if len(args) > 0:
                            row = int(args[0])
                        else:
                            row = None
                        if row != None and row != '':
                            with open(input_path, "r") as file1:
                                content = file1.readlines()
                            if row < len(content):
                                content = content[row]
                            else:
                                err_msg = 'row length is out of index.'
                        else:
                            with open(input_path, "r") as file1:
                                content = file1.read()
                        if content is not None:
                            status=TEST_RESULT_PASS
                            methodoutput=TEST_RESULT_TRUE
                else:
                    err_msg=generic_constants.NOT_SUPPORTED
            else:
                err_msg=result[3]
        except TypeError as e:
            err_msg=ERROR_CODE_DICT['ERR_INDEX_OUT_OF_BOUNDS_EXCEPTION']
        except Exception as e:
             log.error(e)
             err_msg=generic_constants.ERR_MSG1+'Fetching content of'+generic_constants.ERR_MSG2
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,content,err_msg

    def replace_content(self,input_path,existing_content,replace_content,*args):
        """
        def : replace_content
        purpose : calls the respective method to replace the content of given file based on file type
        param : input_path,existing_content,replace_content
        return : bool

        """
        try:

            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            err_msg=None
            output_res=OUTPUT_CONSTANT
            log.debug('reading the inputs')
            params=self.__split(input_path,existing_content,replace_content,*args)
            log.debug('verifying whether the files exist')
            result=self.verify_file_exists(params[0],'')
            if result[1]:
                file_ext,res=self.__get_ext(params[0])
                if res == True:
                    log.debug('replacing the content of file')
                    res1=self.dict[file_ext+'_replace_content'](*params)
                    if res1:
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
            else:
                err_msg=result[3]
        except TypeError as e:
            err_msg=ERROR_CODE_DICT['ERR_INDEX_OUT_OF_BOUNDS_EXCEPTION']
        except Exception as e:
            err_msg=err_msg=generic_constants.ERR_MSG1+'Replacing content of'+generic_constants.ERR_MSG2
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output_res,err_msg

    def write_to_file(self,input_path,content,*args):
        """
        def : replace_content
        purpose : calls the respective method to write the content to given file based on file type
        param : input_path,content
        return : bool

        """
        # Defect #872 Special Language Support using unicode to address the issue (Himanshu)
        coreutilsobj=core_utils.CoreUtils()
        try:
            input_path=coreutilsobj.get_UTF_8(input_path)
            content=coreutilsobj.get_UTF_8(content)
            if SYSTEM_OS == 'Windows':
                import win32com.client
            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            err_msg=None
            output_res=OUTPUT_CONSTANT
            log.debug('reading the inputs')
            params=[]
            log.debug('verifying whether the file exists')
            result=self.verify_file_exists(input_path,'')
            if result[1]==TEST_RESULT_TRUE:
                file_ext,res=self.__get_ext(input_path)
                if file_ext.lower()=='.json':
                    params.append(input_path)
                    params.append(content)
                else:
                    params=self.__split(input_path,content,*args)
                if res == True:
                    log.debug('writing to the file')
                    res,err_msg = self.dict[file_ext+'_write_to_file'](*params)
                    if res:
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
            else:
                err_msg=result[3]
        except TypeError as e:
            err_msg=ERROR_CODE_DICT['ERR_INDEX_OUT_OF_BOUNDS_EXCEPTION']
        except Exception as e:
            err_msg=generic_constants.ERR_MSG1+'Writing to'+generic_constants.ERR_MSG2
            log.error(e)
        if err_msg:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output_res,err_msg

    def get_line_number(self,input_path,content,*args):
        """
        def : get_line_number
        purpose : calls the respective method to get the line number where the content is present in
                  given file based on file type
        param : input_path,content
        return : linenumbers [list]

        """
        Tflag= False
        try:
            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            linenumbers=None
            err_msg=None
            output_res=OUTPUT_CONSTANT
            log.debug('reading the inputs')
            data=content
            try:
                if content != '':
                    params=self.__split(input_path,content,*args)
                    result=self.verify_file_exists(params[0],'')
                    if result[1]==TEST_RESULT_TRUE:
                        file_ext,res=self.__get_ext(params[0])
                        if file_ext == '.xlsx' or file_ext=='.xls':
                            sheet_name=content
                            data=args[0]
                        if res == True:
                            res,linenumbers,err_msg= self.dict[file_ext+'_get_line_number'](*params)
        ##                    logger.print_on_console(linenumbers)
                            if linenumbers is not None and len(linenumbers) > 0:
                                status=TEST_RESULT_PASS
                                methodoutput=TEST_RESULT_TRUE
                                Tflag=True
                    else:
                        err_msg=result[3]
            except:
                params=[]
                params.append(input_path)
                params.append(content)
                result=self.verify_file_exists(params[0],'')
                if result[1]==TEST_RESULT_TRUE:
                    file_ext,res=self.__get_ext(params[0])
                    try:
                        if res == True:
                            res,linenumbers,err_msg= self.dict[file_ext+'_get_line_number'](*params)
        ##                    logger.print_on_console(linenumbers)
                            if linenumbers is not None and len(linenumbers) > 0:
                                status=TEST_RESULT_PASS
                                methodoutput=TEST_RESULT_TRUE
                                Tflag=True
                    except:
                        for src, dest in list(self.cp1252.items()):
                            if src in content:
                                content = content.replace(src, dest)
                        content = content.encode('raw_unicode_escape')
                        params=self.__split(input_path,content,*args)
                        result=self.verify_file_exists(params[0],'')
                        if result[1]==TEST_RESULT_TRUE:
                            file_ext,res=self.__get_ext(params[0])
                            if res == True:
                                res,linenumbers,err_msg= self.dict[file_ext+'_get_line_number'](*params)
            ##                    logger.print_on_console(linenumbers)
                                if linenumbers is not None and len(linenumbers) > 0:
                                    status=TEST_RESULT_PASS
                                    methodoutput=TEST_RESULT_TRUE
                                    Tflag=True
                                else:
                                    content=data.encode('cp1252')
                                    params=self.__split(input_path,content,*args)
                                    result=self.verify_file_exists(params[0],'')
                                    if result[1]==TEST_RESULT_TRUE:
                                        file_ext,res=self.__get_ext(params[0])
                                        if res == True:
                                            res,linenumbers,err_msg= self.dict[file_ext+'_get_line_number'](*params)
                        ##                    logger.print_on_console(linenumbers)
                                            if linenumbers is not None and len(linenumbers) > 0:
                                                status=TEST_RESULT_PASS
                                                methodoutput=TEST_RESULT_TRUE
                                                Tflag=True
                                    else:
                                        err_msg=result[3]
                        else:
                            err_msg=result[3]
            if Tflag!=True:
                ##content=data.encode('cp1252')
                for src, dest in list(self.cp1252.items()):
                    if src in content:
                        content = content.replace(src, dest)
                params=self.__split(input_path,content,*args)        
                content = params[-1].encode('raw_unicode_escape')
                result=self.verify_file_exists(params[0],'')
                if result[1]==TEST_RESULT_TRUE:
                    file_ext,res=self.__get_ext(params[0])
                    if res == True:
                        res,linenumbers,err_msg= self.dict[file_ext+'_get_line_number'](*params)
    ##                    logger.print_on_console(linenumbers)
                        if linenumbers is not None and len(linenumbers) > 0:
                            status=TEST_RESULT_PASS
                            methodoutput=TEST_RESULT_TRUE
                            Tflag=True
                        else:
                            params=self.__split(input_path,content,*args)
                            content = data.encode('cp1252')
                            if file_ext=='.xlsx' or file_ext=='.xls':
                                params[1] = sheet_name
                            params[-1]=content
                            result=self.verify_file_exists(params[0],'')
                            if result[1]==TEST_RESULT_TRUE:
                                file_ext,res=self.__get_ext(params[0])
                                if res == True:
                                    res,linenumbers,err_msg= self.dict[file_ext+'_get_line_number'](*params)
                ##                    logger.print_on_console(linenumbers)
                                    if linenumbers is not None and len(linenumbers) > 0:
                                        status=TEST_RESULT_PASS
                                        methodoutput=TEST_RESULT_TRUE
                                        Tflag=True
                                    else:
                                        err_msg = ERROR_CODE_DICT["ERR_INPUT_NOT_PRESENT"]
                                        linenumbers = OUTPUT_CONSTANT
                            else:
                                err_msg=result[3]
                else:
                    err_msg=result[3]

        except TypeError as e:
            err_msg=ERROR_CODE_DICT['ERR_INDEX_OUT_OF_BOUNDS_EXCEPTION']
        except Exception as e:
            err_msg=err_msg=generic_constants.ERR_MSG1+'getting line number of'+generic_constants.ERR_MSG2
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,linenumbers,err_msg

    def save_file(self, folder_path, file_path,url=None):
        """
        def : save_file
        purpose : Saving a file in windows
        param :folder_path,file_path,url
        return :

        """
        try:
            configvalues = readconfig.configvalues
            delay_stringinput = float(configvalues['delay_stringinput'])
            status = TEST_RESULT_FAIL
            methodoutput = TEST_RESULT_FALSE
            err_msg = None
            output_res = OUTPUT_CONSTANT
            log.debug('reading the inputs')
            folder_path = str(folder_path)
            file_path = str(file_path)
            log.debug('Folder path is ' + folder_path + ' and File is ' + file_path)
            if (not (folder_path is None or folder_path == '' or file_path is None or file_path == '') and os.path.exists(
                folder_path)):

                if url is not None:
                    log.debug('saving the file')
                    path=str(folder_path + "\\" + file_path)  
                    urllib.request.urlretrieve(url,path)
                else:
                    log.debug('saving the file')
                    #Wait for the Save As dialog to load. Might need to increase the wait time on slower machinesy:
                    waiting_time=0 
                    while waiting_time<5:
                        if win32gui.FindWindow(None,"Save As"):
                            break
                        time.sleep(1)
                        waiting_time+=1
                    FILE_NAME = folder_path+'\\'+ file_path
                    #connect used when you are attempting to automate a running process. 
                    # You can pass in the process id, handle, title or path of the program.
                    app = Application().connect(title="Save As") 
                    # Type the file path and name is Save AS dialog
                    app.SaveAs.edit.SetText(str(FILE_NAME))
                    app.SaveAs.Save.click_input()
                status = TEST_RESULT_PASS
                methodoutput = TEST_RESULT_TRUE
                status_message="File has been saved"
                log.info('File has been saved')   
        except (IOError, WindowsError):
            err_msg = ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
        except Exception as e:
            log.error(e)
            err_msg = generic_constants.ERR_MSG1 + 'Saving' + generic_constants.ERR_MSG2
        if err_msg != None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        
        return status,status_message, methodoutput, output_res, err_msg

    def json_compare_content(self,input_path1,input_path2):
        """
        def : json_compare_content
        purpose : calls the respective method to compare the content of 2 Json
        param : input_path1,input_path2
        return : bool

        """

        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        err_msg=None
        output_res=OUTPUT_CONSTANT
        status1=False
        status2=False
        file_ext1=None
        file_ext2=None
        try:
            try:
                if (os.path.exists(input_path1) and os.path.exists(input_path2)):
                    file_ext1,status1=self.__get_ext(input_path1)
                    file_ext2,status2=self.__get_ext(input_path2)
                else:
                    log.info("Provided inputs as a string")
            except Exception as e:
                log.debug("Files not provided {}".format(e))

            if status1 == True and status2==True and file_ext1==file_ext2:
                log.debug('verifying whether the files exists')
                result1=self.verify_file_exists(input_path1,'')
                result2=self.verify_file_exists(input_path2,'')
                if result1[1] == TEST_RESULT_TRUE and result2[1]==TEST_RESULT_TRUE :
                    content1 = json.dumps(input_path1)
                    with open(input_path1) as file_open:
                        content1=json.dumps((json.load(file_open)))
                        file_open.close()

                    with open(input_path2) as file_open:
                        content2=json.dumps((json.load(file_open)))
                        file_open.close()
                    params=[content1,content2]
                    res,err_msg=self.dict[file_ext1.lower()+'_compare_content'](*params)
                    if res:
                        log.info('files contents are same')
                        logger.print_on_console('files contents are same')
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
                else:
                    err_msg=result1[3]
                    if err_msg==None:
                        err_msg=result2[3]

            else:
                err_msg=ERROR_CODE_DICT['ERR_FILE_EXT_MISMATCH']
                try:
                    try:
                        json.loads(input_path1)
                        json.loads(input_path2)
                    except Exception as e:
                        logger.print_on_console("Input error: Invalid json")
                        log.error(e)
                    params=[input_path1,input_path2]
                    res,err_msg=self.dict['.json_compare_content'](*params)
                    if res:
                        log.info('files contents are same')
                        logger.print_on_console('files contents are same')
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
                except Exception as e:
                    if err_msg==None:
                        err_msg="Failed to load the JSON"
                    log.error(e)
        except Exception as e:
            err_msg=generic_constants.ERR_MSG1+'Comapring content of'+generic_constants.ERR_MSG2
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            #logger.print_on_console(err_msg)
        return status,methodoutput,output_res,err_msg

    def __split(self,*args):
        """
        def : __split
        purpose : splits each argumnet by ';' and add it to list and returns
        param : variable number of args
        return : list

        """
        params=[]
        for x in args:
            if isinstance(x, str):
                x.strip()
                val=x.split(';')
                for y in val:
                    params.append(y)
            else:
                params.append(x)
        return params

    cp1252 = {
        # from http://www.microsoft.com/typography/unicode/1252.htm
        "\\u20AC": "\x80", # EURO SIGN
        "\\u201A": "\x82", # SINGLE LOW-9 QUOTATION MARK
        "\\u0192": "\x83", # LATIN SMALL LETTER F WITH HOOK
        "\\u201E": "\x84", # DOUBLE LOW-9 QUOTATION MARK
        "\\u2026": "\x85", # HORIZONTAL ELLIPSIS
        "\\u2020": "\x86", # DAGGER
        "\\u2021": "\x87", # DOUBLE DAGGER
        "\\u02C6": "\x88", # MODIFIER LETTER CIRCUMFLEX ACCENT
        "\\u2030": "\x89", # PER MILLE SIGN
        "\\u0160": "\x8A", # LATIN CAPITAL LETTER S WITH CARON
        "\\u2039": "\x8B", # SINGLE LEFT-POINTING ANGLE QUOTATION MARK
        "\\u0152": "\x8C", # LATIN CAPITAL LIGATURE OE
        "\\u017D": "\x8E", # LATIN CAPITAL LETTER Z WITH CARON
        "\\u2018": "\x91", # LEFT SINGLE QUOTATION MARK
        "\\u2019": "\x92", # RIGHT SINGLE QUOTATION MARK
        "\\u201C": "\x93", # LEFT DOUBLE QUOTATION MARK
        "\\u201D": "\x94", # RIGHT DOUBLE QUOTATION MARK
        "\\u2022": "\x95", # BULLET
        "\\u2013": "\x96", # EN DASH
        "\\u2014": "\x97", # EM DASH
        "\\u02DC": "\x98", # SMALL TILDE
        "\\u2122": "\x99", # TRADE MARK SIGN
        "\\u0161": "\x9A", # LATIN SMALL LETTER S WITH CARON
        "\\u203A": "\x9B", # SINGLE RIGHT-POINTING ANGLE QUOTATION MARK
        "\\u0153": "\x9C", # LATIN SMALL LIGATURE OE
        "\\u017E": "\x9E", # LATIN SMALL LETTER Z WITH CARON
        "\\u0178": "\x9F", # LATIN CAPITAL LETTER Y WITH DIAERESIS
    }

    def write_result_file(self,outputFilePath,content,sheetname):
        """
        def : write_result_file
        purpose : Writing the compare results into the specified file, supporting xls and xlsx extension.
        param : outputFilePath;content;sheetname
        return : bool
        """
        status = False
        err_msg = None
        result3 = None
        index_sheet = None
        try:
            result3=os.path.isfile(outputFilePath)
            file_ext,status1=self.__get_ext(outputFilePath)
            import csv

            if ( file_ext == '.xls' ):
                #checking the file exists and open the exists file.
                if result3==True:
                    rb = open_workbook(outputFilePath)
                    work_book = xl_copy(rb)
                else:
                    #creating the new file, while given file fails.
                    work_book = xlwt.Workbook(encoding="utf-8")

                log.debug('Input Sheet is :')
                log.debug(sheetname)
                try:
                    #adding the new sheet.
                    work_sheet = work_book.add_sheet(sheetname)

                except:
                    #geeting the existing sheet
                    work_sheet = work_book.get_sheet(sheetname)

                index_sheet = work_book.sheet_index(sheetname)

                log.debug('Input Sheet and file path while creating file :')
                log.debug(sheetname)
                log.debug(outputFilePath)
                for column in content.keys():
                    row = 0
                    max_col_width=2962
                    for value in content[column]:
                        work_sheet.write(row,int(column),value)
                        row+=1
                        #add the adjusting width to expand the column.
                        adjusted_width = len(value)*367
                        if (len(value)*367) > max_col_width:
                            max_col_width=(len(value)*367)
                        work_sheet.col(int(column)).width = max_col_width
                work_book.set_active_sheet(index_sheet)
                log.info("WorkSheet Saving : {}".format(outputFilePath))
                work_book.save(outputFilePath)
                del column,row,max_col_width,adjusted_width,value,work_book,work_sheet
                status=True

            elif( file_ext == '.xlsx' ):
                if result3==True:
                    #checking the file and open the existing file.
                    wb = openpyxl.load_workbook(outputFilePath)
                else:
                    #creating the new file.
                    wb = openpyxl.Workbook()

                log.debug(sheetname)
                sheet_names = wb.sheetnames
                if len(sheet_names)>0:
                    if sheetname in sheet_names:
                        index_sheet=sheet_names.index(sheetname)
                    else:
                        index_sheet = 0
                        #adding the New sheet.
                        wb.create_sheet(index=index_sheet, title=sheetname)

                log.debug('Input Sheet and file path while creating file :')
                log.debug(sheetname)
                log.debug(outputFilePath)
                #reading the existing sheet
                sheet = wb[sheetname]
                for column in content.keys():
                    row=1
                    max_col_width = 2.4
                    for value in content[column]:
                        sheet.cell(row=row, column=int(column)+1).value = value
                        row+=1
                        adjusted_width = (len(value) + 2) * 1.2
                        if adjusted_width > max_col_width:
                            max_col_width=adjusted_width
                        sheet.column_dimensions[get_column_letter(int(column)+1)].width = max_col_width
                wb.active=index_sheet
                log.info("WorkBook Saving: {}".format(outputFilePath))
                wb.save(outputFilePath)
                try:
                    del column,row,max_col_width,adjusted_width,value,sheet_names,wb
                except Exception as e :
                    log.error('some error : {}'.format(e))
                status=True

            elif( file_ext == '.csv'):
                if result3==True:
                    with open(outputFilePath, 'r') as csvfile:
                        csv_dict = [row for row in csv.DictReader(csvfile)]
                    if len(csv_dict) == 0:
                        outFile = open(outputFilePath, 'w', newline='')
                    else:
                        logger.print_on_console('Output file has old entries! Erasing the old data to store incoming result.')
                        outFile = open(outputFilePath, 'w+', newline='')
                else:
                    logger.print_on_console("File Does not exists, creating the file in specified path {}".format(outputFilePath))
                    outFile = open(outputFilePath, 'a', newline='')
                result=csv.writer(outFile)
                for key, value in content.items():
                    result.writerow(value)
                outFile.close()
                status=True
            elif(file_ext == '.txt'):
                if result3==True:
                    out = open(outputFilePath, "w")
                    logger.print_on_console('Output file has old entries! Erasing the old data to store incoming result.')
                else:
                    logger.print_on_console("File Does not exists, creating the file in specified path {}".format(outputFilePath))
                    out = open(outputFilePath, "w+")
                for key,value in content.items():
                    out.write(",".join(value)+'\n')
                out.close()
                status=True
        except Exception as e:
            err_msg='Writing to output file Failed'
            log.error(e)
        log.info('Status of write_result_file is ' + str(status) )
        try:
            del outputFilePath,sheetname,result3,index_sheet
        except Exception as e :
            log.error('some error : {}'.format(e))
        return status, err_msg

    def cell_by_cell_compare(self,input_val,*args):
        """
        def : cell_by_cell_compare
        purpose : cell by cell comparison between excel and csv files.
        param : input : <input_path1>;<sheet1(optional)>;<input_path2>;<sheet2(optional)>;<selective(optional- will print diff b/w mismatched text of each cell in provided input files )> output: output_path(optional)
        return : bool/ if output file mentioned then prints cell by cell to that file
        """
        status = TEST_RESULT_FAIL
        result = TEST_RESULT_FALSE
        err_msg = None
        output = OUTPUT_CONSTANT
        flg = False
        desc = None
        collect_content={}
        opt = False
        input_path1 = None
        sheetname1 = None
        input_path2 = None
        sheetname2 = None
        output_feild = None
        sheet_exist=[]
        input_filed_set = False
        dyn_var_opt = False
        con_var_opt = False
        log.debug('Comparing content cell by cell of given files ')
        try:
            import csv
            if (len(input_val)==5):
                input_path1 = input_val[0]
                sheetname1 = input_val[1]
                input_path2 = input_val[2]               
                extension1=os.path.splitext(input_path1)[1] if(input_path1 !=None) else None
                extension2=os.path.splitext(input_path2)[1] if(input_path2 !=None) else None
                sheetname2 = input_val[3] if (extension2!='.csv') else None
                if ( len(input_val) == 5 and input_val[4].strip().lower() == 'selective'):
                    logger.print_on_console("Selective Option Set to True")
                    opt = True
                    log.info("Selective option set True")
                input_filed_set = True
            elif (len(input_val)>=3):
                input_path1 = input_val[0]
                sheetname1 = input_val[1]
                input_path2 = input_val[2]
                extension1=os.path.splitext(input_path1)[1] if(input_path1 !=None) else None
                extension2=os.path.splitext(input_path2)[1] if(input_path2 !=None) else None
                sheetname2 = input_val[3] if (extension2!='.csv') else None
                input_filed_set = True
            else:
                input_filed_set = False

            if input_filed_set:
                if(str(args[0].split(";")[0]).startswith("{") and str(args[0].split(";")[0]).endswith("}")):
                    out_path=self.DV.get_dynamic_value(args[0].split(";")[0])
                    dyn_var_opt = True
                    if ( out_path ):
                        try:
                            if os.path.exists(out_path):
                                output_feild = out_path
                            log.info("Choosen the dynamic file path")
                        except Exception as e:
                            log.debug("File out path is not valid, setting to default")
                            output_feild = None
                        #logger.print_on_console("Choosen the dynamic file path")
                elif(str(args[0].split(";")[0]).startswith("_") and str(args[0].split(";")[0]).endswith("_")):
                    out_path=self.CV.get_constant_value(args[0].split(";")[0])
                    con_var_opt = True
                    if ( out_path ):
                        try:
                            if os.path.exists(out_path):
                                output_feild = out_path
                            else:
                                output_feild = out_path.split(";")[0]
                            log.info("Choosen the constant file path")
                        except Exception as e:
                            log.debug("File out path is not valid, setting to default")
                            output_feild = None
                else:
                    output_feild = args[0].split(";")[0] if args[0]!='' else None
                    log.info("Choosen the direct file path")
                    #logger.print_on_console("Choosen the direct file path")
            if(extension1 in ['.xlsx','.xls'] and extension2 in ['.xlsx','.xls']):# Indicates both files are excel(.xlsx, .xls)
                book1 = open_workbook(input_path1)
                book2 = open_workbook(input_path2)
                try:
                    if (sheetname1.strip() != ''):sheet1 = book1.sheet_by_name(sheetname1)
                    else :
                        sheet1 = book1.sheet_by_index(0)
                except Exception as e:
                    sheet_exist.append(str(e))
                try:
                    if (sheetname2.strip() != ''):sheet2 = book2.sheet_by_name(sheetname2)
                    else :
                        sheet2 = book2.sheet_by_index(0)
                except Exception as e:
                    #logger.print_on_console("Sheet Doesn't Exist: {}".format(str(e)))
                    sheet_exist.append(str(e))

                log.debug('verifying whether the files exists')
                result1=self.verify_file_exists(input_path1,'')
                result2=self.verify_file_exists(input_path2,'')

                if result1[1] == TEST_RESULT_TRUE and result2[1]==TEST_RESULT_TRUE and len(sheet_exist)==0:
                    log.info("taking the maxium row and column")
                    row_max=max(sheet1.nrows,sheet2.nrows)
                    col_max=max(sheet1.ncols,sheet2.ncols)
                    for rownum in range(int(row_max)):
                        for colnum in range(int(col_max)):
                            try:
                                c1=str(sheet1.cell(rownum, colnum).value)
                            except:
                                c1=None
                            try:
                                c2=str(sheet2.cell(rownum, colnum).value)
                            except:
                                c2=None
                            if (c1!=None and c2!=None) and (c1!='' and c2!=''):
                                if c1 != c2:
                                    out='Not Matched'
                                    desc="Not Matched, Cell Values {} and {}".format(c1,c2)
                                else:
                                    out="Matched"
                                    desc="Matched"
                            elif (c1==None or c1=='') and (c2== '' or c2==None):
                                out=''
                                desc=''
                            else:
                                out= 'Not Matched'
                                desc="Not Matched, Cell Values {} and {}".format(c1,c2)

                            if colnum not in collect_content.keys():
                                collect_content[colnum]=[]
                                if( opt == True ):
                                    collect_content[colnum].append(desc)
                                else: 
                                    collect_content[colnum].append(out)
                            else:
                                if( opt == True ):
                                    collect_content[colnum].append(desc)
                                else:
                                    collect_content[colnum].append(out)
                    try:
                        del rownum, colnum, desc, out, c1, c2 #deleting variables
                    except Exception as e :
                        log.error('some error : {}'.format(e))
                else:
                    if sheet_exist!=None:
                        err_msg="Provided sheet doesn't exist: {}".format(', '.join(sheet_exist))
                    else:
                        err_msg = 'Error : Validity of file path 1 is : ' + str(result1[1]) + 'and validity of file path 2 is : ' + str(result2[1]) + '. Please make sure file paths entered are correct'
            elif(extension1=='.csv' and extension2=='.csv'):#both files are csv
                i,j,x=0,0,0
                output1,output2=[],[]
                with open(input_path1,'r') as f1:
                    reader1=csv.reader(f1)
                    for row in reader1:
                        output1.append(row)            
                with open(input_path2,'r') as f2:
                    reader2=csv.reader(f2)
                    for row in reader2:
                        output2.append(row) 
                rc1=len(output1)
                rc2=len(output2)
                cc1=0
                cc2=0
                for i in range(len(output1)):
                    temp_1=len(output1[i])
                    if temp_1>cc1:
                        cc1=temp_1                
                for j in range(len(output2)):
                    temp_2=len(output2[j])
                    if temp_2>cc2:
                        cc2=temp_2
                row_max=max(rc1,rc2)
                col_max=max(cc1,cc2)
                for aa in range(int(row_max)): 
                    for bb in range(int(col_max)):
                        try:
                            val1=output1[aa][bb]
                        except:
                            val1=None
                        try:
                            val2=output2[aa][bb]
                        except:
                            val2=None
                        if (val1!=None and val2!=None) and (val1!='' and val2!=''):
                            if val1 != val2:
                                out='Not Matched'
                                desc="Not Matched, Cell Values {} and {}".format(val1,val2)
                            else:
                                out="Matched"
                                desc="Matched"
                        elif (val1==None or val1=='') and (val2== '' or val2==None):
                            out=''
                            desc=''
                        else:
                            out= 'Not Matched'
                            desc="Not Matched, Cell Values {} and {}".format(val1,val2)
                        if bb not in collect_content.keys():
                            collect_content[bb]=[]
                            if( opt == True ):
                                collect_content[bb].append(desc)
                            else: 
                                collect_content[bb].append(out)
                        else:
                            if( opt == True ):
                                collect_content[bb].append(desc)
                            else:
                                collect_content[bb].append(out)
            else:#one excel and one csv
                if(extension1=='.csv'):
                    file1=input_path1
                    file2=input_path2
                    sheetname1=input_val[3]
                    if(extension2 == '.xlsx'):
                        book1 = openpyxl.load_workbook(file2)
                        sheet1 = book1.get_sheet_by_name(sheetname1)
                    elif(extension2 =='.xls'):
                        book1 = open_workbook(file2)
                        sheet1 = book1.sheet_by_name(sheetname1)

                elif(extension2=='.csv'):
                    file1=input_path2
                    file2=input_path1
                    sheetname1=input_val[1]
                    if(extension1 == '.xlsx'):
                        book1 = openpyxl.load_workbook(file2)
                        sheet1=book1.get_sheet_by_name(sheetname1)
                    elif(extension1 == '.xls'):
                        book1 = open_workbook(file2)
                        sheet1 = book1.sheet_by_name(sheetname1)

                output1,output2=[],{}
                i,x,j=0,0,0
                if(file1):
                    with open(file1,'r') as f1:
                        reader1=csv.reader(f1)
                        for row in reader1: 
                            output1.append(row)
                    rc_csv=len(output1)
                    cc_csv=0
                    cc2=0
                    for i in range(len(output1)):
                        temp_1=len(output1[i])
                        if temp_1>cc_csv:
                            cc_csv=temp_1

                if(extension1=='.xlsx' or extension2=='.xlsx'):
                    min_row=sheet1.min_row
                    max_row=sheet1.max_row
                    max_col=sheet1.max_column
                    cell1=list(sheet1[min_row:max_row])

                    for eachcell in cell1:
                        output2[x]=[]
                        for i in range(len(eachcell)):
                            if eachcell[i].data_type=='n' and type(eachcell[i].value) == int:
                                output2[x].append(str(eachcell[i].value))
                            else:
                                output2[x].append(eachcell[i].value)
                        x+=1
                elif (extension1=='.xls' or extension2=='.xls'):
                    max_row=sheet1.nrows
                    max_col=sheet1.ncols
                    min_row=0
                    min_col=0
                    for row in range(min_row, max_row):
                        output2[x]=[]
                        for col in range(min_col,max_col):
                            cell = sheet1.cell(row,col)
                            cell_value = cell.value
                            if cell.ctype in (2,3) and int(cell_value) == cell_value:
                                cell_value = int(cell_value)
                            output2[x].append(str(cell_value))
                        x+=1
                row_max=max(rc_csv,max_row)
                col_max=max(cc_csv,max_col)
                for aa in range(int(row_max)): 
                    for bb in range(int(col_max)):
                        try:
                            val1=output1[aa][bb]
                        except:
                            val1=None
                        try:
                            val2=output2[aa][bb]
                        except:
                            val2=None
                        if (val1!=None and val2!=None) and (val1!='' and val2!=''):
                            if val1 != val2:
                                out='Not Matched'
                                desc="Not Matched, Cell Values {} and {}".format(val1,val2)
                            else:
                                out="Matched"
                                desc="Matched"
                        elif (val1==None or val1=='') and (val2== '' or val2==None):
                            out=''
                            desc=''
                        else:
                            out= 'Not Matched'
                            desc="Not Matched, Cell Values {} and {}".format(val1,val2)
                        if bb not in collect_content.keys():
                            collect_content[bb]=[]
                            if( opt == True ):
                                collect_content[bb].append(desc)
                            else: 
                                collect_content[bb].append(out)
                        else:
                            if( opt == True ):
                                collect_content[bb].append(desc)
                            else:
                                collect_content[bb].append(out)
            if( output_feild ):
                file_extension,status_get_ext = self.__get_ext(output_feild)
                if file_extension=='.csv':
                    cc={}
                    temp=0
                    #converting data present in collect_content to row wise data(write_result_file function uses write_row() for writing data to csv file)
                    for i in range(len(collect_content)):
                        temp_a=len(collect_content.get(i))
                        if temp_a>temp:
                            temp=temp_a
                    for t in range(temp):
                        cc[t]=[]
                    for i in range(len(collect_content)):
                        k=0
                        for j in range(len(collect_content.get(i))):
                            cc[k].append(collect_content.get(i)[j])
                            k=k+1
                    collect_content=cc
                sheet='CellByCellCompare_Result'
                if os.path.exists(output_feild):
                    logger.print_on_console( "Writing the output of cellByCellCompare to file ")
                    if file_extension in ['.xlsx','.xls']:
                        wb=open_workbook(output_feild)
                        if(sheet in wb._sheet_names):
                            msg='Output file has old entries! Erasing old data to store incoming result.'
                            logger.print_on_console(msg)
                            if file_extension == '.xlsx':
                                self.xlsx_obj.clear_content_xlsx(os.path.dirname(output_feild),os.path.basename(output_feild),sheet)
                            else:
                                self.xls_obj.clear_content_xls(os.path.dirname(output_feild),os.path.basename(output_feild),sheet)
                    flg, err_msg = self.write_result_file(output_feild, collect_content, sheet)
                else:
                    status_excel_create_file = False
                    file_extension,status_get_ext = self.__get_ext(output_feild)
                    if status_get_ext and file_extension is not None and file_extension in generic_constants.EXCEL_TYPES:
                        logger.print_on_console("File Does not exists, creating the file in specified path {}".format(output_feild))
                        status_excel_create_file,e_msg = self.dict[file_extension + '_create_file'](output_feild,sheet)
                    elif file_extension=='.csv':
                        status_excel_create_file=True
                    else:
                        err_msg = 'Warning! : Invalid file extension. Output file format should be either ".xls" ".csv" or ".xlxs".'
                    if (status_excel_create_file):
                        logger.print_on_console( "Writing the output of cellByCellCompare to file : " + str(output_feild) )
                        flg, err_msg = self.write_result_file(output_feild, collect_content, sheet)
                if ( flg ):
                    if opt==True:
                        comp_flg=True
                        if len(collect_content)!=0:
                            for each_key in collect_content.keys():
                                for each_match in collect_content[each_key]:
                                    if each_match.find("Not Matched")!=-1:
                                        comp_flg=False
                            if comp_flg!=False:
                                logger.print_on_console("Input files are same")
                            else:
                                logger.print_on_console("Input files are not same")
                            status = TEST_RESULT_PASS
                            result = TEST_RESULT_TRUE
                        else:
                            logger.print_on_console("Input files are empty")

                    else:
                        if len(collect_content)!=0:
                            status = TEST_RESULT_PASS
                            result = TEST_RESULT_TRUE
            else:
                log.debug("Output file not provided")
                if (collect_content):
                    if opt==True or dyn_var_opt or con_var_opt:
                        comp_flg=True
                        if len(collect_content)!=0:
                            for each_key in collect_content.keys():
                                for each_match in collect_content[each_key]:
                                    if each_match.find("Not Matched")!=-1:
                                        comp_flg=False
                            if comp_flg!=False:
                                logger.print_on_console("Input files are same")
                            else:
                                logger.print_on_console("Input files are not same")

                            status = TEST_RESULT_PASS
                            result = TEST_RESULT_TRUE
                        else:
                            logger.print_on_console("Input files are empty")
                    else:
                        logger.print_on_console("output variable or file path is not present")
                        status = TEST_RESULT_PASS
                        result = TEST_RESULT_TRUE
                    output = collect_content
                else :
                    err_msg = "Content is empty"
            if input_filed_set==False:
                err_msg="Invalid Inputs, Please provide the proper inputs"

            if ( err_msg ):
                log.error(err_msg)
                logger.print_on_console(err_msg)
        except Exception as e:
            err_msg = 'Error occured in compare content of two files ERR_MSG: ' + str(e)
            log.error(err_msg)
            logger.print_on_console(err_msg)
        try:
            del input_path1, sheetname1, input_path2, sheetname2, row_max, col_max, output_feild, collect_content, opt#deleting variables
        except Exception as e :
            log.error('some error : {}'.format(e))
        return status, result, output, err_msg

    def find_file_path(self,input,*args):
        """
        def : find_file_path
        purpose : Fetch the loaction of the input file
        param : filename, folderpath
        return : Location of the input file
        """
        try:
            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            err_msg=None
            log.debug('reading the inputs')
            output=[]
            if(len(input)==2):
                filename=input[0]
                folderpath=input[1]
                if filename.lower() == 'latest':
                    from pathlib import Path
                    path = Path(folderpath)
                    # Get a list of all files in the specified path
                    all_files = [f for f in path.iterdir() if f.is_file()]
                    # Sort the files by modification time (newest first)
                    sorted_files = sorted(all_files, key=lambda f: f.stat().st_mtime, reverse=True)
                    if sorted_files:
                        # Return the path to the latest downloaded file
                        log.info(f"The latest downloaded file is: {sorted_files[0]}")
                        logger.print_on_console('File location is fetched')
                        status = TEST_RESULT_PASS
                        methodoutput = TEST_RESULT_TRUE
                        output = [str(sorted_files[0]).replace('\\','/')]
                    else:
                        # Return None if no files are found
                        log.info('File does not exist inside '+folderpath)
                        logger.print_on_console('File does not exist inside '+folderpath)
                else:
                    for root, dirs, files in os.walk(folderpath):
                        if filename in files:
                            output.append(os.path.join(root,filename).replace('\\','/'))
                    if(output==[]):
                        log.info('File does not exist inside '+folderpath)
                        logger.print_on_console('File does not exist inside '+folderpath)
                    else:
                        log.info('File location is fetched')
                        logger.print_on_console('File location is fetched')
                        status = TEST_RESULT_PASS
                        methodoutput = TEST_RESULT_TRUE
            else:
                err_msg="Invalid number of inputs"
        except Exception as e:
            err_msg = 'Error occured in fetching file location'
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output,err_msg

    def selective_cell_compare(self,input,*args):
        """
        def : selective_cell_compare
        purpose : selective cell comparison between excel and csv files
        param : CSV/Excel file path1, sheetname (optional), selected cell range, CSV/Excel file path2, sheetname (optional), selected cell range,case(optional)
        return : bool/ if output file mentioned then prints cell by cell to that file

        """
        try:
            import csv
            status=TEST_RESULT_FAIL
            methodoutput=TEST_RESULT_FALSE
            err_msg=None
            output_res=OUTPUT_CONSTANT
            case = ''
            log.debug('reading the inputs')
            flag1=False
            dyn_var_opt = False
            con_var_opt = False
            input_valid = False
            extension1 = None
            extension2 = None
            if len(input)<=7:
                input_valid = True
                filepath1=input[0]
                filepath2=input[3]
                sheetname1=input[1]
                sheetname2=input[4]
                range1=input[2].split(':') if ':' in input[2] else None
                range2=input[5].split(':') if ':' in input[5] else None
                res1={}
                output_feild = None
                extension1=os.path.splitext(filepath1)[1] if(filepath1 !=None) else None
                extension2=os.path.splitext(filepath2)[1] if(filepath2 !=None) else None
            if (extension1==None or extension2==None) or not input_valid:
                err_msg=ERROR_CODE_DICT['ERR_INVALID_INPUT']
                log.error(err_msg)
                logger.print_on_console(err_msg)
                return status,methodoutput,output_res,err_msg
            if (range1==None or range2==None):
                err_msg="Invalid Range Delimiter! Please provide valid range"
                log.error(err_msg)
                logger.print_on_console(err_msg)
                return status,methodoutput,output_res,err_msg
            if(len(input)>6):
                if(isinstance(input[6], str) and input[6].lower() == 'ignorecase'):
                    case = input[6]
                else:
                    logger.print_on_console('Warning! : Case check input is invalid, Hence ignorecase has not been applied.')

            if(extension1=='.csv' and extension2=='.csv'):#both files are csv
                col11 = " ".join(re.findall("[a-zA-Z]+", range1[0]))
                col12 = " ".join(re.findall("[a-zA-Z]+", range1[1]))
                row11=int(" ".join(re.findall(r'[0-9]+', range1[0])))
                row12=int(" ".join(re.findall(r'[0-9]+', range1[1])))
                column11=self.convert.convertStringToInt(col11)
                column12=self.convert.convertStringToInt(col12)

                col21 = " ".join(re.findall("[a-zA-Z]+", range2[0]))
                col22 = " ".join(re.findall("[a-zA-Z]+", range2[1]))
                row21=int(" ".join(re.findall(r'[0-9]+', range2[0])))
                row22=int(" ".join(re.findall(r'[0-9]+', range2[1])))
                column21=self.convert.convertStringToInt(col21)
                column22=self.convert.convertStringToInt(col22)

                i,j,x=0,0,0
                output1,output2={},{}
                with open(filepath1,'r') as f1:
                    reader1=csv.reader(f1)
                    for row in islice(reader1,min(row11,row12)-1,max(row11,row12)): #row length
                        output1[i]=[]
                        for column in islice(row,min(column11,column12)-1,max(column11,column12)): #column length
                            output1[i].append(column)
                        i+=1

                with open(filepath2,'r') as f2:
                    reader2=csv.reader(f2)
                    for row in islice(reader2,min(row21,row22)-1,max(row21,row22)): #row length
                        output2[j]=[]
                        for column in islice(row,min(column21,column22)-1,max(column21,column22)): #column length
                            output2[j].append(column)
                        j+=1

                i=list(output2.keys())[0]
                count=0
                for aa in range(len(output1)):
                    res1[x]=[]
                    for bb in range(len(output2[i])):
                        if(case!=''):
                            if len(output1[aa]) != 0 and len(output2[aa]) != 0:
                                if (output1[aa][bb].lower()==output2[aa][bb].lower()):
                                    output='True'
                                    res1[x].append(output)
                                else:
                                    output='False'
                                    res1[x].append(output)
                            elif len(output1[aa]) == 0 or len(output2[aa]) == 0:
                                output='False'
                                res1[x].append(output)
                            else:
                                output='False'
                                res1[x].append(output)
                        else:
                            if len(output1[aa]) != 0 and len(output2[aa]) != 0:
                                if (output1[aa][bb]==output2[aa][bb]):
                                    output='True'
                                    res1[x].append(output)
                                else:
                                    output='False'
                                    res1[x].append(output)
                            elif len(output1[aa]) == 0 or len(output2[aa]) == 0:
                                output='False'
                                res1[x].append(output)
                            else:
                                output='False'
                                res1[x].append(output)
                    if len(output1[aa]) == 0 and len(output2[i]) == 0:
                        output='True'
                        res1[x].append(output)
                    count+=1
                    x+=1
                    i+=1
                    if(count>=len(output2)):
                        break

                if (len(output1)<len(output2)):
                    for i in range(len(output1)):
                        output2.pop(i)
                    xx=list(output2.keys())[0]
                    for k in range(len(output2)):
                        res1[x]=[]
                        for j in range(len(output2[xx])):
                            output='False'
                            res1[x].append(output)
                        if len(output2[xx]) == 0:
                            output='False'
                            res1[x].append(output)
                        x+=1
                        xx+=1
                elif(len(output1)>len(output2)):
                    for i in range(len(output2)):
                        output1.pop(i)
                    xx=list(output1.keys())[0]
                    for k in range(len(output1)):
                        res1[x]=[]
                        for j in range(len(output1[xx])):
                            output='False'
                            res1[x].append(output)
                        if len(output1[xx]) == 0:
                            output='False'
                            res1[x].append(output)
                        x+=1
                        xx+=1

            elif(extension1=='.xlsx' and extension2=='.xlsx'):# Indicates both files are excel(.xlsx)
                book1 = openpyxl.load_workbook(filepath1)
                book2 = openpyxl.load_workbook(filepath2)
                #verify whether file exists or not
                result1=self.verify_file_exists(filepath1,'')
                result2=self.verify_file_exists(filepath2,'')

                sheet1=book1.get_sheet_by_name(sheetname1)
                sheet2=book2.get_sheet_by_name(sheetname2)

                cell1=list(sheet1[range1[0]:range1[1]])
                cell2=list(sheet2[range2[0]:range2[1]])

                x,i=0,0
                for aa in range(len(cell1)):
                    res1[x]=[]
                    for bb in range(len(cell2[i])):
                        if(case!=''):
                            if (str(cell1[aa][bb].value).lower()==str(cell2[aa][bb].value).lower()):
                                output='True'
                                res1[x].append(output)
                            else:
                                output='False'
                                res1[x].append(output)
                        else:
                            if (str(cell1[aa][bb].value)==str(cell2[aa][bb].value)):
                                output='True'
                                res1[x].append(output)
                            else:
                                output='False'
                                res1[x].append(output)
                    x+=1
                    i+=1
                    if(i>=len(cell2)):
                        break

                if (len(cell1)<len(cell2)):
                    for i in range(len(cell1)):
                        del cell2[0]
                    for k in range(len(cell2)):
                        res1[x]=[]
                        for m in range(len(cell2[0])):
                            output='False'
                            res1[x].append(output)
                        x+=1
                elif(len(cell1)>len(cell2)):
                    for i in range(len(cell2)):
                        del cell1[0]
                    for k in range(len(cell1)):
                        res1[x]=[]
                        for m in range(len(cell1[0])):
                            output='False'
                            res1[x].append(output)
                        x+=1

            elif(extension1=='.xls' and extension2=='.xls'):
                book1 = open_workbook(filepath1)
                book2 = open_workbook(filepath2)
                sheet1 = book1.sheet_by_name(sheetname1)
                sheet2 = book2.sheet_by_name(sheetname2)

                #verify whether file exists or not
                result1=self.verify_file_exists(filepath1,'')
                result2=self.verify_file_exists(filepath2,'')

                col11 = " ".join(re.findall("[a-zA-Z]+", range1[0]))
                col12 = " ".join(re.findall("[a-zA-Z]+", range1[1]))
                row11=int(" ".join(re.findall(r'[0-9]+', range1[0])))
                row12=int(" ".join(re.findall(r'[0-9]+', range1[1])))
                column11=self.convert.convertStringToInt(col11)
                column12=self.convert.convertStringToInt(col12)

                col21 = " ".join(re.findall("[a-zA-Z]+", range2[0]))
                col22 = " ".join(re.findall("[a-zA-Z]+", range2[1]))
                row21=int(" ".join(re.findall(r'[0-9]+', range2[0])))
                row22=int(" ".join(re.findall(r'[0-9]+', range2[1])))
                column21=self.convert.convertStringToInt(col21)
                column22=self.convert.convertStringToInt(col22)

                output1,output2={},{}
                x,j,z=0,0,0
                for row in range(row11-1, row12):
                        output1[x]=[]
                        for col in range(column11-1,column12):
                            cell = sheet1.cell(row,col)
                            cell_value = cell.value
                            if cell.ctype in (2,3) and int(cell_value) == cell_value:
                                cell_value = int(cell_value)
                            output1[x].append(str(cell_value))
                        x+=1

                for row in range(row21-1, row22):
                        output2[z]=[]
                        for col in range(column21-1,column22):
                            cell = sheet2.cell(row,col)
                            cell_value = cell.value
                            if cell.ctype in (2,3) and int(cell_value) == cell_value:
                                cell_value = int(cell_value)
                            output2[z].append(str(cell_value))
                        z+=1

                i=list(output2.keys())[0]
                count=0
                for aa in range(len(output1)):
                    res1[j]=[]
                    for bb in range(len(output2[i])):
                        if(case!=''):
                            if (str(output1[aa][bb]).lower()==str(output2[aa][bb]).lower()):
                                output='True'
                                res1[j].append(output)
                            else:
                                output='False'
                                res1[j].append(output)
                        else:
                            if (str(output1[aa][bb])==str(output2[aa][bb])):
                                output='True'
                                res1[j].append(output)
                            else:
                                output='False'
                                res1[j].append(output)
                    aa+=1
                    j+=1
                    count+=1
                    if(count>=len(output2)):
                        break

                if (len(output1)<len(output2)):
                    for i in range(len(output1)):
                        output2.pop(i)
                    xx=list(output2.keys())[0]
                    for k in range(len(output2)):
                        res1[j]=[]
                        for m in range(len(output2[xx])):
                            output='False'
                            res1[j].append(output)
                        j+=1
                elif(len(output1)>len(output2)):
                    for i in range(len(output2)):
                        output1.pop(i)
                    xx=list(output1.keys())[0]
                    for k in range(len(output1)):
                        res1[j]=[]
                        for m in range(len(output1[xx])):
                            output='False'
                            res1[j].append(output)
                        j+=1

            elif((extension1=='.xls' and extension2=='.xlsx')or(extension1=='.xlsx' and extension2=='.xls')):
                #verify whether file exists or not
                result1=self.verify_file_exists(filepath1,'')
                result2=self.verify_file_exists(filepath2,'')
                output1,output2={},{}
                x,y,j=0,0,0

                if(extension1=='.xlsx'):
                    col11 = " ".join(re.findall("[a-zA-Z]+", range2[0]))
                    col12 = " ".join(re.findall("[a-zA-Z]+", range2[1]))
                    row11=int(" ".join(re.findall(r'[0-9]+', range2[0])))
                    row12=int(" ".join(re.findall(r'[0-9]+', range2[1])))
                    column11=self.convert.convertStringToInt(col11)
                    column12=self.convert.convertStringToInt(col12)

                    book1 = openpyxl.load_workbook(filepath1)
                    sheet1=book1.get_sheet_by_name(sheetname1)
                    book2 = open_workbook(filepath2)
                    sheet2 = book2.sheet_by_name(sheetname2)
                    cell1=list(sheet1[range1[0]:range1[1]])

                elif(extension2=='.xlsx'):
                    col11 = " ".join(re.findall("[a-zA-Z]+", range1[0]))
                    col12 = " ".join(re.findall("[a-zA-Z]+", range1[1]))
                    row11=int(" ".join(re.findall(r'[0-9]+', range1[0])))
                    row12=int(" ".join(re.findall(r'[0-9]+', range1[1])))
                    column11=self.convert.convertStringToInt(col11)
                    column12=self.convert.convertStringToInt(col12)

                    book1 = openpyxl.load_workbook(filepath2)
                    sheet1=book1.get_sheet_by_name(sheetname2)
                    book2 = open_workbook(filepath1)
                    sheet2 = book2.sheet_by_name(sheetname1)
                    cell1=list(sheet1[range2[0]:range2[1]])

                for eachcell in cell1:
                    output1[x]=[]
                    for i in range(len(eachcell)):
                        output1[x].append(eachcell[i].value)
                    x+=1

                for row in range(row11-1, row12):
                    output2[y]=[]
                    for col in range(column11-1,column12):
                        cell = sheet2.cell(row,col)
                        cell_value = cell.value
                        if cell.ctype in (2,3) and int(cell_value) == cell_value:
                            cell_value = int(cell_value)
                        output2[y].append(str(cell_value))
                    y+=1

                i=list(output2.keys())[0]
                count=0
                for aa in range(len(output1)):
                    res1[j]=[]
                    for bb in range(len(output2[i])):
                        if(case!=''):
                            if (str(output1[aa][bb]).lower()==str(output2[aa][bb]).lower()):
                                output='True'
                                res1[j].append(output)
                            else:
                                output='False'
                                res1[j].append(output)
                        else:
                            if (str(output1[aa][bb])==str(output2[aa][bb])):
                                output='True'
                                res1[j].append(output)
                            else:
                                output='False'
                                res1[j].append(output)
                    aa+=1
                    j+=1
                    count+=1
                    if(count>=len(output2)):
                        break

                if (len(output1)<len(output2)):
                    for i in range(len(output1)):
                        output2.pop(i)
                    xx=list(output2.keys())[0]
                    for k in range(len(output2)):
                        res1[j]=[]
                        for m in range(len(output2[xx])):
                            output='False'
                            res1[j].append(output)
                        j+=1
                elif(len(output1)>len(output2)):
                    for i in range(len(output2)):
                        output1.pop(i)
                    xx=list(output1.keys())[0]
                    for k in range(len(output1)):
                        res1[j]=[]
                        for m in range(len(output1[xx])):
                            output='False'
                            res1[j].append(output)
                        j+=1


            else:#one excel and one csv
                if(extension1=='.csv'):
                    file1=filepath1
                    file2=filepath2
                    range1=range1
                    range2=range2
                    sheetname1=input[4]
                    if(extension2=='.xlsx'):
                        book1 = openpyxl.load_workbook(file2)
                        sheet1=book1.get_sheet_by_name(sheetname1)
                    elif(extension2=='.xls'):
                        book1 = open_workbook(file2)
                        sheet1 = book1.sheet_by_name(sheetname1)

                elif(extension2=='.csv'):
                    file1=filepath2
                    file2=filepath1
                    range1=range2
                    range2=range1
                    sheetname1=input[1]
                    if(extension1=='.xlsx'):
                        book1 = openpyxl.load_workbook(file2)
                        sheet1=book1.get_sheet_by_name(sheetname1)
                    elif(extension1=='.xls'):
                        book1 = open_workbook(file2)
                        sheet1 = book1.sheet_by_name(sheetname1)

                col11 = " ".join(re.findall("[a-zA-Z]+", range1[0]))
                col12 = " ".join(re.findall("[a-zA-Z]+", range1[1]))
                row11=int(" ".join(re.findall(r'[0-9]+', range1[0])))
                row12=int(" ".join(re.findall(r'[0-9]+', range1[1])))
                column11=self.convert.convertStringToInt(col11)
                column12=self.convert.convertStringToInt(col12)
                output1,output2={},{}
                i,x,j=0,0,0
                if(file1):
                    with open(file1,'r') as f1:
                        reader1=csv.reader(f1)
                        for row in islice(reader1,min(row11,row12)-1,max(row11,row12)): #row length
                            output1[i]=[]
                            for column in islice(row,min(column11,column12)-1,max(column11,column12)): #column length
                                output1[i].append(column)
                            i+=1

                #verify whether file exists or not
                result1=self.verify_file_exists(file1,'')
                result2=self.verify_file_exists(file2,'')

                # sheet1=book1.get_sheet_by_name(sheetname1)
                if(extension1=='.xlsx' or extension2=='.xlsx'):
                    cell1=list(sheet1[range2[0]:range2[1]])

                    for eachcell in cell1:
                        output2[x]=[]
                        for i in range(len(eachcell)):
                            output2[x].append(eachcell[i].value)
                        x+=1
                elif(extension1=='.xls' or extension2=='.xls'):
                    for row in range(row11-1, row12):
                        output2[x]=[]
                        for col in range(column11-1,column12):
                            cell = sheet1.cell(row,col)
                            cell_value = cell.value
                            if cell.ctype in (2,3) and int(cell_value) == cell_value:
                                cell_value = int(cell_value)
                            output2[x].append(str(cell_value))
                        x+=1


                i=list(output2.keys())[0]
                count=0
                for aa in range(len(output1)):
                    res1[j]=[]
                    for bb in range(len(output2[i])):
                        if(case!=''):
                            if (str(output1[aa][bb]).lower()==str(output2[aa][bb]).lower()):
                                output='True'
                                res1[j].append(output)
                            else:
                                output='False'
                                res1[j].append(output)
                        else:
                            if (str(output1[aa][bb])==str(output2[aa][bb])):
                                output='True'
                                res1[j].append(output)
                            else:
                                output='False'
                                res1[j].append(output)
                    aa+=1
                    j+=1
                    count+=1
                    if(count>=len(output2)):
                        break

                if (len(output1)<len(output2)):
                    for i in range(len(output1)):
                        output2.pop(i)
                    xx=list(output2.keys())[0]
                    for k in range(len(output2)):
                        res1[j]=[]
                        for m in range(len(output2[xx])):
                            output='False'
                            res1[j].append(output)
                        j+=1
                elif(len(output1)>len(output2)):
                    for i in range(len(output2)):
                        output1.pop(i)
                    xx=list(output1.keys())[0]
                    for k in range(len(output1)):
                        res1[j]=[]
                        for m in range(len(output1[xx])):
                            output='False'
                            res1[j].append(output)
                        j+=1
                        
            if(str(args[0].split(";")[0]).startswith("{") and str(args[0].split(";")[0]).endswith("}")):
                out_path = self.DV.get_dynamic_value(args[0].split(";")[0])
                dyn_var_opt = True
                if(out_path):
                    output_feild = out_path
            elif(str(args[0].split(";")[0]).startswith("_") and str(args[0].split(";")[0]).endswith("_")):
                out_path=self.CV.get_constant_value(args[0])
                con_var_opt = True
                if(out_path):
                    output_feild = out_path
            else:
                output_feild = args[0].split(";")[0]
            if res1:
                status = TEST_RESULT_PASS
                methodoutput = TEST_RESULT_TRUE
            if(output_feild):
                file_extension,status_get_ext = self.__get_ext(output_feild)
                sheet='selectiveCellCompare'
                result=os.path.isfile(output_feild)

                msg='Output file has old entries! Erasing old data to store incoming result.'
                if status_get_ext and file_extension is not None and file_extension in generic_constants.SELECTIVE_CELL_FILE_TYPES:
                    logger.print_on_console( "Writing the output of selectiveCellCompare to file ")
                    if(file_extension=='.xlsx'):
                        if result==True:
                            wb=openpyxl.load_workbook(output_feild)
                        else:
                            logger.print_on_console("File Does not exists, creating the file in specified path {}".format(output_feild))
                            wb=openpyxl.Workbook()

                        if(sheet in wb.sheetnames):
                            logger.print_on_console(msg)
                            self.xlsx_obj.clear_content_xlsx(os.path.dirname(output_feild),os.path.basename(output_feild),sheet)
                            wb=openpyxl.load_workbook(output_feild)
                        else:
                            wb.create_sheet(index=0, title=sheet)

                        row,col=1,1
                        getSheet=wb[sheet]
                        for key, value in res1.items():
                            for i in range(len(res1[key])):
                                getSheet.cell(row, col , value[i])
                                col += 1
                            col = 1
                            row+=1
                        wb.save(output_feild)
                        wb.close()
                        flag1=True

                    elif(file_extension=='.xls'):
                        if result==True:
                            rb = open_workbook(output_feild)
                            wb = xl_copy(rb)
                            getSheet = wb.get_sheet(sheet)
                        else:
                            logger.print_on_console("File Does not exists, creating the file in specified path {}".format(output_feild))
                            wb = xlwt.Workbook(encoding="utf-8")
                        try:
                            if(getSheet):
                                logger.print_on_console(msg)
                                self.xls_obj.clear_content_xls(os.path.dirname(output_feild),os.path.basename(output_feild),sheet)
                                rb = open_workbook(output_feild)
                                wb = xl_copy(rb)
                                getSheet = wb.get_sheet(sheet)
                        except:
                            getSheet = wb.add_sheet(sheet)

                        row,col=0,0
                        for key, value in res1.items():
                            for i in range(len(res1[key])):
                                getSheet.write(row, col , value[i])
                                col += 1
                            col = 0
                            row+=1
                        wb.save(output_feild)
                        flag1=True

                    #for csv and txt
                    if(file_extension=='.csv' or file_extension=='.txt'):
                        flag1, err_msg = self.write_result_file(output_feild, res1, sheet)

                    if flag1 is False:
                        log.error(err_msg)
                        logger.print_on_console(err_msg)
                        status=TEST_RESULT_FAIL
                        methodoutput=TEST_RESULT_FALSE
                    else:
                        log.info('Compared cells between the mentioned files')
                        logger.print_on_console('Compared cells between the mentioned files')
                        status = TEST_RESULT_PASS
                        methodoutput = TEST_RESULT_TRUE
                else:
                    err_msg = 'Warning! : Invalid file extension(Supports only .xlsx, .xls, .txt, .csv).'
            elif not dyn_var_opt or not con_var_opt:
                err_msg='Output field is empty'
                log.error(err_msg)
                logger.print_on_console(err_msg)
        except Exception as err_msg:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,res1,err_msg

    def get_page_count(self,filePath):
        """
        def : get_page_count
        purpose : get page count of .pdf files
        param : filePath
        return : pagecount [int]
        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        err_msg=None
        pageCount=None
        try:
            if filePath != None and filePath != '':
                file_ext,res=self.__get_ext(filePath)
                if file_ext == '.pdf':
                    res,pageCount,err_msg = self.dict[file_ext+'_get_page_count'](filePath)
                else:
                    err_msg = file_ext + ' File type is not supported.'
               
                if res and err_msg == None:
                    status=TEST_RESULT_PASS
                    methodoutput=TEST_RESULT_TRUE
                    info_msg="Page Count is "
                    logger.print_on_console(info_msg + str(pageCount))
            else:
                err_msg = 'Error in File Path.'

        except Exception as e:
            # err_msg=err_msg=generic_constants.ERR_MSG1+'getting line number of'+generic_constants.ERR_MSG2
            err_msg = 'Error occured while fetching page count.'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)

        return status,methodoutput,pageCount,err_msg

    def get_delimited_value(self, filePath, row, col, delimiter):
        """
        def : get_delimited_value
        purpose : To fecth value from file based on row and column separated by delimiter
        param : filePath, row, col, delimiter
        return : value
        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        err_msg=None
        output=None
        try:
            row = int(row)
            col = int(col)
            if filePath != None and filePath != '':
                with open(filePath, "r") as file1:
                    read_content = file1.readlines()
                if row != None and row != '':
                    if row <= len(read_content):
                        if delimiter != None and delimiter != '':
                            content = read_content[row].split(delimiter)
                            if col != None and col != '':
                                if col <= len(content):
                                    value = content[col-1]
                                    logger.print_on_console(f"Value: {value}")
                                    status=TEST_RESULT_PASS
                                    methodoutput=TEST_RESULT_TRUE
                                    output = value
                                else:
                                    err_msg = 'column length is out of index.'
                            else:
                                err_msg = 'colum number is empty.'
                        else:
                            err_msg = 'delimiter is empty.'
                    else:
                        err_msg = 'row length is out of index.'
                else:
                    err_msg = 'row number is empty.'
            else:
                err_msg = 'Error in File Path.'

        except Exception as e:
            err_msg = 'Error occured while fetching value.'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)

        return status,methodoutput,output,err_msg
