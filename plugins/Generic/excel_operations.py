#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      sushma.p
#
# Created:     05-10-2016
# Copyright:   (c) sushma.p 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------

from copy import copy 
from distutils.log import ERROR
from encodings import search_function
from operator import index
from tkinter.tix import Tree
import logger
import generic_constants
import xlwt
import openpyxl
import xlsxwriter
import xlutils
import datetime
import xlrd
import os
import sys
from xlrd import open_workbook
from openpyxl import load_workbook
from constants import *
import logging
from constants import *
log = logging.getLogger('excel_operations.py')
import core_utils
from xlrd.sheet import ctype_text
import itertools
import csv
import platform
if SYSTEM_OS == "Windows":
    import win32com.client
    import win32api
if SYSTEM_OS == 'Darwin' or SYSTEM_OS =='Linux':
    import pandas
    from pandas import ExcelWriter

class ExcelFile:

    """The instantiation operation __init__ creates an empty object of the class ExcelFile when it is instantiated"""
    def __init__(self):

        self.excel_path=None
        self.sheetname=None
        self.xls_obj=ExcelXLS()
        self.xlsx_obj=ExcelXLSX()
        self.csv_obj=ExcelCSV()

        self.dict={
        'delete_row_.xls':self.xls_obj.delete_row_xls,
        'read_cell_.xls':self.xls_obj.read_cell_xls,
        'write_cell_.xls':self.xls_obj.write_cell_xls,
        'clear_cell_.xls':self.xls_obj.clear_cell_xls,
        'get_rowcount_.xls':self.xls_obj.get_rowcount_xls,
        'get_colcount_.xls':self.xls_obj.get_colcount_xls,
        'copy_workbook_.xls':self.xls_obj.copy_workbook_xls,

        'delete_row_.xlsx':self.xlsx_obj.delete_row_xlsx,
        'read_cell_.xlsx':self.xlsx_obj.read_cell_xlsx,
        'write_cell_.xlsx':self.xlsx_obj.write_cell_xlsx,
        'clear_cell_.xlsx':self.xlsx_obj.clear_cell_xlsx,
        'get_rowcount_.xlsx':self.xlsx_obj.get_rowcount_xlsx,
        'get_colcount_.xlsx':self.xlsx_obj.get_colcount_xlsx,
        'copy_workbook_.xlsx':self.xlsx_obj.copy_workbook_xlsx,

        'delete_row_.csv':self.csv_obj.delete_row_csv,
        'read_cell_.csv':self.csv_obj.read_cell_csv,
        'write_cell_.csv':self.csv_obj.write_cell_csv,
        'clear_cell_.csv':self.csv_obj.clear_cell_csv,
        'get_rowcount_.csv':self.csv_obj.get_rowcount_csv,
        'get_colcount_.csv':self.csv_obj.get_colcount_csv,

        }


    def open_and_save_file(self,input_path):
        #win32 part opening and closing of file to claculate and save values of the formula
        #This is to support the feature of   simulataneous writing and reading to a file
        excelobj= object_creator()
        excel=excelobj.excel_object()
        if excel!=None:
            try:
                excel.DisplayAlerts = False
                excel_file = excel.Workbooks.Open(input_path)
                excel_file.Close(True)
            except Exception as e:
                log.debug("Excel COM object error",e)


    def __get_ext(self,input_path):
        """
        def : __get_ext(private method)
        purpose : returns the extension/type of the given file and checks if it is valid type
        param : input_path
        return : Returns Status and file type

        """
        err_msg=None
        try:
            filename,file_ext=os.path.splitext(input_path)
            status=True
            log.info('File type is :')
            log.info(file_ext)
            return file_ext,status,err_msg

        except Exception as e:
            log.error(e)
            err_msg=generic_constants.INVALID_FILE_FORMAT
            ##logger.print_on_console(err_msg)
        return '',False,err_msg


    def set_excel_path(self,input_path,*args):
        """
        def : set_excel_path
        purpose : sets the given excel file path and sheet name for excel operations
        param : input_path,sheetname
        return : bool

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        output = OUTPUT_CONSTANT
        err_msg = None
        sheetname = args[0] if args else None
        logger.print_on_console(generic_constants.INPUT_IS+input_path+' '+(sheetname or ''))
        if input_path != None  and input_path != '':
            if os.path.isfile(input_path):
                file_ext,res,err_msg=self.__get_ext(input_path)
                if file_ext in generic_constants.EXCEL_TYPES and sheetname != None  and sheetname != '':
                    log.info('File type is :'+str(file_ext))
                    logger.print_on_console('File type is :'+str(file_ext))
                    self.excel_path=input_path
                    try:
                        if file_ext.endswith('x'):
                            wb=openpyxl.load_workbook(self.excel_path,data_only=True)
                            sheet=wb.get_sheet_by_name(sheetname)
                            del wb,sheet
                        else:
                            wb = open_workbook(self.excel_path)
                            sheet = wb.sheet_by_name(sheetname)
                            del wb,sheet
                        self.sheetname=sheetname
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
                    except KeyError:
                        err_msg="Worksheet "+sheetname+" does not exist."
                        self.excel_path=None
                        self.sheetname=None
                elif file_ext == '.csv':
                    log.info('File type is :'+str(file_ext))
                    logger.print_on_console('File type is :'+str(file_ext))
                    self.excel_path=input_path
                    status=TEST_RESULT_PASS
                    methodoutput=TEST_RESULT_TRUE
                else:
                    err_msg="Error: Please check your inputs"
            else:
                err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_FOUND_EXCEPTION']
        else:
            err_msg=generic_constants.INVALID_INPUT
        if err_msg != None:
            log.error(err_msg)
            logger.print_on_console(err_msg)

        return status,methodoutput,output,err_msg


    def delete_row(self,row,*args):
        """
        def : delete_row
        purpose : calls the respective method to delete the given row of excel
                  file set by the user
        param : row
        return : Returns Bool

        """
        output = OUTPUT_CONSTANT
        err_msg = None
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        try:
            if self.excel_path != None:
                file_ext,res,err_msg=self.__get_ext(self.excel_path)
                if res:
                    myfile = open(self.excel_path, "a+") # or "a+", whatever you need
                    myfile.close()
                    res,err_msg=self.dict['delete_row_'+file_ext](int(row),self.excel_path,self.sheetname,*args)
                    if res:
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
                        info_msg="Specified row(s) are deleted"
                        logger.print_on_console(info_msg)                        
            else:
                err_msg=generic_constants.FILE_PATH_NOT_SET
        except IOError:
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
        except ValueError as e:
            err_msg=ERROR_CODE_DICT['ERR_NUMBER_FORMAT_EXCEPTION']
        except Exception as e:
            err_msg='Error occured in deleting row'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output,err_msg


    def read_cell(self,row,col):
        """
        def : read_cell
        purpose : calls the respective method to read the given cell of excel
                  file set by the user
        param : row,col
        return : Returns cell value

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        output=None
        err_msg = None
        try:
            if self.excel_path != None:
                file_ext,res,err_msg=self.__get_ext(self.excel_path)
                if res:
                    info_msg=generic_constants.INPUT_IS+self.excel_path+' '+(self.sheetname or '')
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    info_msg='Row is '+str(row)+' and col is '+str(col)
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    try:
                        row=int(row)
                        col=int(col)
                    except Exception as e:
                        row=int(row)
                        col=self.convertStringToInt(col)
                    if row>0 and col>0:
                        res,value,err_msg=self.dict['read_cell_'+file_ext](row,col,self.excel_path,self.sheetname)
                    #872 added unicode support for info (Himanshu)
                        try:
                            info_msg='cell value is '+str(value)
                        except:
                            info_msg='cell value is '+value
                        logger.print_on_console(info_msg)
                        log.info(info_msg)
                        if res:
                            status=TEST_RESULT_PASS
                            methodoutput=TEST_RESULT_TRUE
                    #872 return without str conversion if unicode (Himanshu)
                            if not(type(value) is str):
                                output=str(value)                  ##Value returned must be integer so that it will display 20 instead of 20L
                                log.info(info_msg)
                            else:
                                output=value
                    else:
                        err_msg=ERROR_CODE_DICT["ERR_ROW_COLUMN"]
            else:
                err_msg=generic_constants.FILE_PATH_NOT_SET
        except ValueError as e:
            err_msg=ERROR_CODE_DICT['ERR_NUMBER_FORMAT_EXCEPTION']
        except Exception as e:
            err_msg='Error occured in reading excel cell value'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output,err_msg


    def write_cell(self,row,col,value,*args):
        """
        def : write_cell
        purpose : calls the respective method to write to the given cell of excel
                  file set by the user
        param : row,col,value
        return : Returns Bool

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        output = OUTPUT_CONSTANT
        err_msg = None

        #Support for special languages (#872) decode if unicode(special language) (Himanshu)
        coreutilsobj=core_utils.CoreUtils()
        value=coreutilsobj.get_UTF_8(value)

        try:
            if self.excel_path != None:
                file_ext,res,err_msg=self.__get_ext(self.excel_path)
                if res:
                    info_msg=generic_constants.INPUT_IS+self.excel_path+' '+(self.sheetname or '')
                    log.info(info_msg)
                    #983(Sakshi) - str required when value is an int because of the error - cannot concatenate string and int.
                    if(isinstance(value,int)):
                        info_msg='Row is '+str(row)+' col is '+str(col)+' and Value: '+str(value)
                    else:
                        #872 removed str from value to support unicode (Himanshu)
                        info_msg='Row is '+str(row)+' col is '+str(col)+' and Value: '+value
                    log.info(info_msg)
                    ##logger.print_on_console(info_msg)
                    try:
                        row=int(row)
                        col=int(col)
                    except Exception as e:
                        row=int(row)
                        if col.isalpha():
                            col=self.convertStringToInt(col)
                        else:
                            raise ValueError
                    if 0<row<=1048576 and 0<col<=16384:                     
                    # if row>0 and col>0:
                        res,err_msg=self.dict['write_cell_'+file_ext](row,col,value,self.excel_path,self.sheetname,*args)
                        if res:
                            status=TEST_RESULT_PASS
                            methodoutput=TEST_RESULT_TRUE
                    else:
                        err_msg=ERROR_CODE_DICT["ERR_ROW_COL_OUT_OF_RANGE"]
            else:
                err_msg=generic_constants.FILE_PATH_NOT_SET
        except ValueError as e:
            err_msg=ERROR_CODE_DICT['ERR_NUMBER_FORMAT_EXCEPTION']
        except IOError:
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
        except Exception as e:
            err_msg='Error occured in writing excel cell value'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output,err_msg


    def clear_cell(self,row,col):
        """
        def : clear_cell
        purpose : calls the respective method to clear the given cell of excel
                  file set by the user
        param : row,col
        return : Returns Bool

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        output = OUTPUT_CONSTANT
        err_msg = None
        try:
            if self.excel_path != None:
                file_ext,res,err_msg=self.__get_ext(self.excel_path)
                if res:
                    info_msg=generic_constants.INPUT_IS+self.excel_path+' '+(self.sheetname or '')
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    info_msg='Row is '+str(row)+' and col is '+str(col)
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    try:
                        row=int(row)
                        col=int(col)
                    except Exception as e:
                        row=int(row)
                        col=self.convertStringToInt(col)
                    if row>0 and col>0:
                        res,err_msg=self.dict['clear_cell_'+file_ext](row,col,self.excel_path,self.sheetname)
                        if res:
                            status=TEST_RESULT_PASS
                            methodoutput=TEST_RESULT_TRUE
                    else:
                        err_msg=ERROR_CODE_DICT["ERR_ROW_COLUMN"]
            else:
                err_msg=generic_constants.FILE_PATH_NOT_SET
        except ValueError as e:
            err_msg=ERROR_CODE_DICT['ERR_NUMBER_FORMAT_EXCEPTION']
        except IOError:
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
        except Exception as e:
            err_msg='Error occured in clearing excel cell value'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output,err_msg


    def get_rowcount(self,col_number=None):
        """
        def : get_rowcount
        purpose : calls the respective method to get the number of rows of excel
                  file set by the user
        param : None
        return : Returns row count(int)

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        row_count=None
        err_msg = None
        try:
            if self.excel_path != None:
                file_ext,res,err_msg=self.__get_ext(self.excel_path)
                if res:
                    info_msg=generic_constants.INPUT_IS+self.excel_path+' '+(self.sheetname or '')
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    res,row_count,err_msg=self.dict['get_rowcount_'+file_ext](self.excel_path,self.sheetname,col_number)
                    logger.print_on_console('Row count is:'+str(row_count))
                    if res:
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
            else:
                err_msg=generic_constants.FILE_PATH_NOT_SET
        except Exception as e:
            err_msg='Error occured in getting row count of excel'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,row_count,err_msg


    def get_colcount(self,row_number=None):
        """
        def : get_colcount
        purpose : calls the respective method to get the number of columns of excel
                  file set by the user
        param : None
        return : Returns col count(int)

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        col_count=None
        err_msg = None
        try:
            if self.excel_path != None:
                file_ext,res,err_msg=self.__get_ext(self.excel_path)
                if res:
                    info_msg=generic_constants.INPUT_IS+self.excel_path+' '+(self.sheetname or '')
                    logger.print_on_console(info_msg)
                    log.info(info_msg)
                    res,col_count,err_msg=self.dict['get_colcount_'+file_ext](self.excel_path,self.sheetname,row_number)
                    logger.print_on_console('Column count is:'+str(col_count))
                    if res:
                        status=TEST_RESULT_PASS
                        methodoutput=TEST_RESULT_TRUE
            else:
                err_msg=generic_constants.FILE_PATH_NOT_SET
        except Exception as e:
            err_msg='Error occured in getting column count of excel'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,col_count,err_msg


    def clear_excel_path(self,*args):
        """
        def : clear_excel_path
        purpose : celars the excel path
        param : None
        return : bool

        """
        status=TEST_RESULT_FAIL
        methodoutput=TEST_RESULT_FALSE
        output = OUTPUT_CONSTANT
        err_msg = None
        if self.excel_path != None and self.excel_path != '':
            self.excel_path=None
            self.sheetname=None
            status=TEST_RESULT_PASS
            methodoutput=TEST_RESULT_TRUE
        else:
            err_msg=generic_constants.FILE_PATH_NOT_SET
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,output,err_msg

    
    def copy_workbook(self,filePath1,filePath2,*args):
        """
        def : copy_workbook
        purpose : calls the respective method to copy workbook from filePath1 to filePath2
        param : filePath1,filePath2,option
        return : Returns Bool

        """
        # output = OUTPUT_CONSTANT
        err_msg = None
        status = TEST_RESULT_FAIL
        methodoutput = TEST_RESULT_FALSE
        option = '0'
        if len(args) > 0 and args[0] != '':
            option = None if args[0] not in ['0','1','2'] else args[0]
        logger.print_on_console('Copy from workbook '+filePath1 + ' to ' + filePath2)
        try:
            if option != None:
                if filePath1 != None and filePath2 != None:
                    file_ext1,res1,err_msg=self.__get_ext(filePath1)
                    if err_msg is None:
                        file_ext2,res2,err_msg=self.__get_ext(filePath2)
                    if err_msg is None and (file_ext1 not in ['.xls','.xlsx'] or file_ext2 not in ['.xls','.xlsx']):
                        err_msg = 'Invalid file format'
                    
                    if err_msg is None:
                        if file_ext1 == file_ext2 and res1 and res2:
                            myfile = open(filePath1, "r")
                            myfile.close()
                            myfile = open(filePath2, "r")
                            myfile.close()
                            res,err_msg=self.dict['copy_workbook_'+file_ext1](filePath1,filePath2,option)
                            if res:
                                status=TEST_RESULT_PASS
                                methodoutput=TEST_RESULT_TRUE
                                info_msg="Successfully Copied"
                                logger.print_on_console(info_msg)
                        else:
                            err_msg = 'Error in copying workbooks as file extensions are different.'
                else:
                    err_msg=generic_constants.FILE_PATH_NOT_SET
            else:
                err_msg = 'Option is Invalid.'
        except IOError as e:
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE'] + ' / ' + ERROR_CODE_DICT['ERR_FILE_NOT_FOUND_EXCEPTION']
        except ValueError as e:
            err_msg=ERROR_CODE_DICT['ERR_NUMBER_FORMAT_EXCEPTION']
        except Exception as e:
            err_msg='Error occured in copying workbooks.'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,methodoutput,err_msg

    def convertStringToInt(self,col):
        num=0
        import string
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

    def getSheetTrueName(self,name):
        if(name[-1] == ')'):
            name = name[::-1]
            num = ''
            index = 0
            for elem in range(1,len(name)):
                if(name[elem] != '('):
                    num+=name[elem]
                else:
                    index = elem
                    break
            
            name = name[::-1]
            num = num[::-1]
            if(num == ''):
                num = '0'
            if(num.isdigit()):
                if(index+1 == len(name)):
                    return '',True,int(num)
                # if(index+1 == ' '):
                #     index+=1
                return name[:-1*(index+1)],True,int(num)
            else:
                return name,False,1

        return name,False,1

class ExcelXLS:

    def __init__(self):
        self.excel_obj=''
        self.cell_type={'string' : 'General',
        'formula': 'f',
        'number': '0.00',
        'date':'n',
        'boolean': 'b',
        'null':'n',
        'inline': 'inlineStr',
        'error': 'e',
        'formula_cache': 'str'}


    def __load_workbook_xls(self,inputpath,sheetname):
        """
        def : __load_workbook_xls(private method)
        purpose : loads the given .xls file
        param : inputpath,sheetname
        return : list

        """
        ##book = open_workbook(inputpath,formatting_info=True)
        book = open_workbook(inputpath)
        sheet=book.sheet_by_name(sheetname)
        last_row=-1
        last_col=-1
        if sheet.nrows>0:
            last_row=sheet.nrows-1
            row=sheet.row(last_row)
            last_col=len(row)-1

        return book,sheet,last_row,last_col,sheet.nrows


    def __write_to_cell_xls(self,input_path,book,sheetname,row,col,value,*args):
        """
        def : __write_to_cell_xls(private method)
        purpose : writes to the the cell in the given row and column of .xls file
        param : input_path,book,sheetname,row,col,value
        return : bool

        """

        from xlutils.copy import copy
        from xlwt import easyxf,XFStyle
        workbook = copy(book)
        sheets=book.sheet_names()
        status=False
        err_msg=None
        flag=False
        self.excel_obj=ExcelFile()
        try:
            sheetnum=sheets.index(sheetname)
            s = workbook.get_sheet(sheetnum)
            if len(args)>0 and args[0] is not None:
                type=args[0].lower()
                if type in list(self.cell_type.keys()):
                    type=self.cell_type[type]
                    if type=='0.00':
                        flag=True
                    	#Setting cell format to number
                        style=XFStyle()
                        style.num_format_str=type
                        value=int(value)
                        s.write(int(row),int(col),value,style)
                    elif type in ['f','b']:
                        flag=True
                    	#Evaluating formula
                        value=xlwt.ExcelFormula.Formula(value)
                        ##value='='+value
                        s.write(int(row),int(col),value)
                    else:
                        flag=True
                        ##value=str(value)
                        s.write(int(row),int(col),value)
                    workbook.save(input_path)
                    status= True
                else:
                    logger.print_on_console('Cell Type not supported')
                    log.info('Cell Type not supported')
            else:
                flag=True
                status= True
                ##value=str(value)
                s.write(int(row),int(col),value)


            workbook.save(input_path)

        except IOError:
            err_msg=err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
            log.error(err_msg)
        except Exception as e:
            log.error(e)
            err_msg='Error writing to excel cell of .xls file'
        workbook.save(input_path)
        if SYSTEM_OS == 'Windows' and flag and status:
            self.excel_obj.open_and_save_file(input_path)
        return status,err_msg


    def get_linenumber_xls(self,input_path,sheetname,content):
        """
        def : get_linenumber_xls
        purpose : get the line number where content is present in given .xls file
        param  : input_path,sheetname,content
        return : Returns line number (list)

        """
        status=False
        line_number=[]
        err_msg=None
        try:
            # logger.print_on_console(generic_constants.INPUT_IS+input_path+' '+sheetname)
            log.info(generic_constants.INPUT_IS+input_path+' '+sheetname)
            book = open_workbook(input_path)
            sheet = book.sheet_by_name(sheetname)
            for col in range(sheet.ncols):
                indices = [i for i, x in enumerate(sheet.col_values(col)) if x == content]
                if indices != None:
                    for i in indices:
                        i = i+1
                        line_number.append(i)
                        status=True
            # log.info(line_number)
            # logger.print_on_console('Line numbers:')
            # logger.print_on_console(line_number)
        except Exception as e:
           err_msg='Error getting line number in .xls'
           log.error(e)
           logger.print_on_console(err_msg)
        ##log.info('Status is '+str(status))
        return status,line_number,err_msg


    def replace_content_xls(self,input_path,sheetname,existingcontent,replacecontent,*args):
        """
        def : replace_content_xls
        purpose : replace the 'existingcontent' by 'replacecontent' in given sheet of
                  given .xls file
        param : input_path,sheetname,existingcontent,replacecontent
        return : bool

        """
        logger.print_on_console(generic_constants.INPUT_IS+input_path+' '+sheetname)
        log.info(generic_constants.INPUT_IS+input_path+' '+sheetname)
        status=False
        err_msg=None
        log.debug('Replacing content of .xls files')
        try:
            from xlutils.copy import copy
            book = open_workbook(input_path)
            sheet = book.sheet_by_name(sheetname)
            workbook = copy(book)
            sheets=book.sheet_names()
            sheetnum=sheets.index(sheetname)
            s = workbook.get_sheet(sheetnum)
            log.debug('Existing content '+str(existingcontent))
            log.debug('Replacing content '+str(replacecontent))
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    if existingcontent == sheet.cell(row,col).value:
                        log.debug('Replacing at '+str(col))
                        s.write(row,col,replacecontent)
            workbook.save(input_path)
            status=True
        except Exception as e:
           err_msg='Error occured in replacing content of .xls files'
           log.error(e)
        log.info('Status is '+str(status))
        return status,err_msg


    def delete_row_xls(self,row,excel_path,sheetname,*args):
        """
        def : delete_row_xls
        purpose : deletes the given row of both .xls and .xlsx files
        param : row
        return : bool

        """
        status=False
        err_msg=None
        self.excel_obj=ExcelFile()
        log.debug('Deleting a row of .xls file')
        excel_file=None
        try:
            if SYSTEM_OS == 'Darwin':
                writer = ExcelWriter(excel_path)
                excel_data = pandas.read_excel(excel_path)
                if row<=len(excel_data.index):
                    excel_data = excel_data.drop(excel_data.index[[row - 1]])
                    excel_data.to_excel(writer, sheetname, index=False)
                    status=True
                else:
                    err_msg=ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
                writer.save()
            elif SYSTEM_OS == 'Windows':
                excelobj= object_creator()
                excel=excelobj.excel_object()
                excel.DisplayAlerts = False
                excel_file = excel.Workbooks.Open(excel_path)
                if not(excel_file.ReadOnly):
                    sheet = excel.Sheets(sheetname)
                    last_row=sheet.UsedRange.Rows.Count
                    if len(args)>0:
                        rows_to_delete=[int(value) for value in args]
                        rows_to_delete.append(row)
                        rows_to_delete.sort(reverse=True)
                        for row_number in rows_to_delete:
                            if row_number<=last_row:
                                sheet.Rows(row_number).Delete()
                                status=True
                        if not status:
                            err_msg=ERROR_CODE_DICT["ERR_MULTIROW_DOESN'T_EXIST"]
                        
                    else:       
                        if row<=last_row:
                            sheet.Rows(row).Delete()
                            status=True
                        else:
                            err_msg=ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
                else:
                    err_msg='Excel is Read only'
            else:
                writer = pandas.ExcelWriter(excel_path)
                excel_data = pandas.read_excel(excel_path,header=None)
                if row<=len(excel_data.index):
                    excel_data = excel_data.drop(excel_data.index[[row - 1]])
                    excel_data.to_excel(writer, sheetname, index=False,header=False)
                    writer.save()
                    status=True    
                else:
                    err_msg=ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
        except Exception as e:
            err_msg='Error occured in deleting row of excel file'
            log.error(e)
        finally:
            if excel_file!=None:
                excel_file.Close(True)


        if err_msg!=None:
            log.error(err_msg)
        log.info('Status is '+str(status))
        return status,err_msg


    def compare_content_xls(self,input_path1,sheetname1,input_path2,sheetname2,*args):
        """
        def : compare_content_xls
        purpose : compares the data of given sheets of 2 different excel files
        param : input_path1,input_path2,Sheet1,Sheet2
        return : bool

        """
        status=False
        err_msg=None
        x = False
        log.debug('Comparing content of .xls files')
        try:
            ##import pandas as pd
            ##file1=pd.ExcelFile(input_path1)
            ##data1=file1.parse(sheetname1)
            ##log.debug('File content1'+str(data1))
            ##file2=pd.ExcelFile(input_path2)
            ##data2=file2.parse(sheetname2)
            ##log.debug('File content2'+str(data2))
            ##status= (str(data1)==str(data2))
            from itertools import zip_longest
            book1 = open_workbook(input_path1)
            book2 = open_workbook(input_path2)
            sheet1 = book1.sheet_by_name(sheetname1) if sheetname1!='' else None
            sheet2 = book2.sheet_by_name(sheetname2) if sheetname2!='' else None
            if (sheet1==None or sheet2==None):
                err_msg=ERROR_CODE_DICT['ERR_INVALID_INPUT']
                return status,err_msg
            for rownum in range(max(sheet1.nrows, sheet2.nrows)):
                if rownum < sheet1.nrows and rownum < sheet2.nrows:
                    row_rb1 = sheet1.row_values(rownum)
                    row_rb2 = sheet2.row_values(rownum)

                    for colnum, (c1, c2) in enumerate(zip_longest(row_rb1, row_rb2)):
                        if c1 != c2:
                            log.debug('Row ',str(rownum+1),' Col ',str(colnum+1),' cell value 1 ',c1,' cell value 2 ',c2)
                            x = True
                            break
                    if(x):
                        status = False
                    else:
                        status=True
                else:
                    log.error('Row ',str(rownum+1),' is missing')
                    err_msg='File contents are not same'
                    status=False
                    break


        except Exception as e:
            err_msg='Error occured in compare content of .xls files'
            log.error(e)
        log.info('Status is '+str(status))
        return status,err_msg


    def verify_content_xls(self,input_path,sheetname,content,*args):
        """
        def : verify_content_xls
       purpose : verify whether the 'existingcontent' is present in given sheet of
                  given .xls file
        param : input_path,sheetname,content
        return : bool

        """
        status=False
        err_msg=None
        res,line_number,err_msg=self.get_linenumber_xls(input_path,sheetname,content)
        if res and line_number != None:
            status=True
        return status,err_msg


    def delete_sheet_xls(self,inputpath,sheetname):
        """
        def : delete_sheet_xls
        purpose : Deleting the sheet from .xls file
        param  : inputpath,sheetname
        return : nothing (if it fails then through the exception)

        """
        excelobj= object_creator()
        excel=excelobj.excel_object()
        excel.DisplayAlerts = False
        excel_file = excel.Workbooks.Open(inputpath)
        sheet = excel.Sheets(sheetname)
        sheet.Delete()
        excel_file.Close(True)


    def create_sheet_xls(self,inputpath,sheetname):
        """
        def : create_sheet_xls
        purpose : creating a new sheet without removing existing data.
        param : inputpath,sheetname
        return : nothing (if it fails then through the exception)

        """
        from xlutils.copy import copy
        xl_ref=xlrd.open_workbook(inputpath)
        xlw_ref=copy(xl_ref)
        xlw_ref.add_sheet(sheetname)
        xlw_ref.save(inputpath)


    def clear_content_xls(self,inputpath,filename,sheetname):
        """
        def : clear_content_xls
        purpose : removes the given sheet from the given .xls file
        param : input_path,sheetname
        return : bool

        """
        status=False
        err_msg=None
        info_msg=generic_constants.INPUT_IS+inputpath+' filename: '+filename+' Sheetname: '+sheetname
        logger.print_on_console(info_msg)
        log.info(info_msg)
        if inputpath.endswith('/'):
            inputpath=inputpath+filename
        else:
            inputpath=inputpath+'/'+filename
        try:
            if sheetname in xlrd.open_workbook(inputpath).sheet_names():
                import random,string
                dummy_sheet_name =''.join([random.choice(string.ascii_letters + string.digits) for n in range(10)])
                self.create_sheet_xls(inputpath,dummy_sheet_name)
                self.delete_sheet_xls(inputpath,sheetname)
                self.create_sheet_xls(inputpath,sheetname)
                self.delete_sheet_xls(inputpath,dummy_sheet_name)
                status=True
            else:
                err_msg='File/Sheet does not exist to clear'
                log.error(err_msg)
        except PermissionError as ex:
            err_msg = generic_constants.FILE_OPENED
            log.error(ex)
        except Exception as e:
            err_msg='File/Sheet does not exist to clear'
            log.error(e)
        return status,err_msg


    def read_cell_xls(self,row,col,excel_path,sheetname,*args):
        """
        def : read_cell_xls
        purpose : reads the cell value in the given row and column of .xls file
        param : input_path,sheetname
        return : cell value

        """
        status=False
        value=None
        err_msg=None
        try:
            log.debug('Writing to the file'+generic_constants.INPUT_IS+excel_path+' '+sheetname)
            book = open_workbook(excel_path)
            sheet = book.sheet_by_name(sheetname)
            if row<=sheet.nrows:
                if col<=sheet.ncols:
                    cell=sheet.cell(row-1,col-1)
                    value=cell.value
                    if cell.ctype==3:
                        value=datetime.datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))
                        value=value.date()
                    #print cell.ctype
                    status=True
                    log.info(value)
                else:
                    err_msg=ERROR_CODE_DICT["ERR_COL_DOESN'T_EXIST"]
            else:
                err_msg=ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
        except Exception as e:
            err_msg='Error occured in read cell of .xls file'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
        return status,value,err_msg


    def clear_cell_xls(self,row,col,excel_path,sheetname):
        """
        def : clear_cell_xls
        purpose : clears the given cell of .xls file set by the user
        param : row,col
        return : bool

        """
        status=False
        err_msg=None
        log.debug(generic_constants.INPUT_IS+excel_path+' '+sheetname)
        from xlutils.copy import copy
        book = open_workbook(excel_path)
        workbook = copy(book)
        sheets=book.sheet_names()
        try:
            if row>0 and col>0:
                sheetnum=sheets.index(sheetname)
                s = workbook.get_sheet(sheetnum)
                s.write(row-1,col-1,'')
                workbook.save(excel_path)
                status=True
            else:
                logger.print_on_console('Row/Col should be greater than 0')
        except Exception as e:
            err_msg='Error occurred in clearing cell of .xls file'
        workbook.save(excel_path)
        return status,err_msg


    def write_cell_xls(self,row,col,value,excel_path,sheetname,*args):
        """
        def : write_cell_xls
        purpose : calls the private methods to write given value to the cell of given
                row or col of .xls file
        param : row,col,value
        return : bool

        """
        #loads the xls workbook
        workook_info=self.__load_workbook_xls(excel_path,sheetname)
        #writes to the cell in given row,col
        status=False
        if(int(row)>0 and int(col)>0):
            row=int(row)-1
            col=int(col)-1
            status,err_msg=self.__write_to_cell_xls(excel_path,workook_info[0],sheetname,row,col,value,*args)
        return status,err_msg


    def get_rowcount_xls(self,excel_path,sheetname,column_number):
        """
        def : get_rowcount_xls
        purpose : get the number of rows of .xls excel file set by the user
        param : None
        return : Returns row count(int)

        """
        status=False
        row_count=None
        err_msg=None
        try:
            log.debug('Fetching row count of '+generic_constants.INPUT_IS+excel_path+' '+sheetname)
            book = open_workbook(excel_path)
            sheet = book.sheet_by_name(sheetname)
            if column_number == None or column_number == '':
                row_count=sheet.nrows
                status = True
            else:
                column_number = int(column_number) - 1
                if column_number < sheet.ncols:
                    row_count = len(sheet.col(column_number))
                    for i in reversed(sheet.col(column_number)):
                        if i.value == '': row_count -= 1
                        else: break
                else: row_count=0
                status = True
        except ValueError as e:
            err_msg = ERROR_CODE_DICT['ERR_COL_NUMBER']
            log.error(e)
            logger.print_on_console(err_msg)
        except Exception as e:
            err_msg='Error getting the row count of .xls file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,row_count,err_msg


    def get_colcount_xls(self,excel_path,sheetname,row_number):
        """
        def : get_colcount_xls
        purpose : get the number of columns of .xls excel file set by the user
        param : None
        return : Returns row count(int)

        """
        status=False
        col_count=None
        err_msg=None
        log.debug('Fetching col count of '+generic_constants.INPUT_IS+excel_path+' '+sheetname)
        try:
            book = open_workbook(excel_path)
            sheet = book.sheet_by_name(sheetname)
            if row_number == '' or row_number == None:
                col_count=sheet.ncols
                status = True
            else:
                row_number = int(row_number) - 1
                if row_number < sheet.nrows:     
                    col_count = len(sheet.row(row_number))
                    for j in reversed(sheet.row(row_number)):
                        if j.value == '': col_count -= 1
                        else: break
                else: col_count=0        
                status = True
        except ValueError as e:
            err_msg = ERROR_CODE_DICT['ERR_ROW_NUMBER']
            log.error(e)
            logger.print_on_console(err_msg)

        except Exception as e:
            err_msg='Error getting the col count of .xls file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,col_count,err_msg


    def write_to_file_xls(self,input_path,sheetname,content,*args):
        """
        def : write_to_file_xls
        purpose : writes the given content to the given .xls file
        param : inputpath,sheetname,content,('newline'-optional param)
        return : bool

        """
        self.excel_obj=ExcelFile()
        status=False
        err_msg=None
        coreutilsobj=core_utils.CoreUtils()
        sheetname=coreutilsobj.get_UTF_8(sheetname)
        log.debug(generic_constants.INPUT_IS+input_path+' '+sheetname)
        try:
            #loads the xls workbook
            workbook_info=self.__load_workbook_xls(input_path,sheetname)
            row=workbook_info[2]
            col=workbook_info[3]
            len_args=len(args)
            ##print row,col
            ##if len(args)>0 and args[0].lower() == 'newline'  :
                ##col=0
                ##if workbook_info[3] != -1:
                    ##row=row+1
                    ##col=workbook_info[3]
                ##else:
                    ##row=0
            ##elif workbook_info[3]==-1:
                ##row=col=0
            ##else:
                ##col+=1
            ###writes to the cell in given row,col
            ##status,err_msg=self.__write_to_cell_xls(input_path,workbook_info[0],sheetname,row,col,content,args)
            if row==-1 and col==-1:
                row=0
                col=0
                ##print row, col, content
                status,err_msg=self.__write_to_cell_xls(input_path,workbook_info[0],sheetname,row,col,content)
            else:
                if len(args)>0 and args[len_args-1].lower() == 'newline':
                    row=row+1
                    col=0
                    status,err_msg=self.__write_to_cell_xls(input_path,workbook_info[0],sheetname,row,col,content)
                else:
                    col=col+1
                    status,err_msg=self.__write_to_cell_xls(input_path,workbook_info[0],sheetname,row,col,content)
            ##print "content ",row,col
            if len(args)>0 and args[len_args-1].lower() == 'newline':

                for i in range(0,(len_args-1)):
                    workbook_info=self.__load_workbook_xls(input_path,sheetname)
                    row=row+1
                    col=0
                    ##print "start",i,row,col
                    status,err_msg=self.__write_to_cell_xls(input_path,workbook_info[0],sheetname,row,col,args[i])
            else:
                for i in range(0,(len_args)):
                    workbook_info=self.__load_workbook_xls(input_path,sheetname)
                    row=row
                    col=col+1
                #writes to the cell in given row,col
                    ##print "start",i,row,col
                    status,err_msg=self.__write_to_cell_xls(input_path,workbook_info[0],sheetname,row,col,args[i])

        except Exception as e:
            err_msg='Error occurred writing to .xls file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,err_msg


    def create_file_xls(self,excel_path,sheet_name,*args):
        """
        def : create_file_xls
        purpose : creates an xls file using provided path
        param : excel_path,sheet_name
        return : bool

        """
        status=False
        err_msg=None
        try:
            wb = xlwt.Workbook()
            wb.add_sheet(sheet_name)
            wb.save(excel_path)
            status = True
        except Exception as e:
            log.error(e)
            err_msg='Error occurred creating .xls file '
            logger.print_on_console(err_msg)
        return status,err_msg

    def copy_workbook_xls(self,filePath1,filePath2,option):
        """
        def : copy_workbook_xls
        purpose : calls copy_workbook_xls to copy from one workbook to another
        param : filePath1,filePath2,option
        return : bool
        """
        status=False
        err_msg=None
        self.excel_obj=ExcelFile()
        try:
            wb1=open_workbook(filePath1)
            wb2=open_workbook(filePath2)
            sheetNames1 = wb1.sheet_names()
            sheetNames2 = wb2.sheet_names()
            copyWorkbook1 = xlutils.copy.copy(wb1)

            if(option == '0'):
                index = -1
                for sheet in sheetNames2:
                    index+=1
                    if sheet.upper() not in (name.upper() for name in sheetNames1):
                        sheetData = wb2.sheet_by_index(index)
                        newSheet = copyWorkbook1.add_sheet(sheet)

                        for row in range(sheetData.nrows):
                            for col in range(sheetData.ncols):
                                newSheet.write(row,col,sheetData.cell_value(row,col))

            if(option == '1'):
                copyWorkbook2 = xlutils.copy.copy(wb2)
                index = -1
                for sheet in sheetNames1:
                    index+=1
                    newSheetName = None
                    if sheet.upper() not in (name.upper() for name in sheetNames2):
                        newSheetName = sheet
                    else:
                        trueName,check,num = self.excel_obj.getSheetTrueName(sheet)
                        if(check == False):
                            trueName+=" "
                        
                        num+=1
                        searchName = trueName+'('+str(num)+')'
                        while(searchName.upper() in (name.upper() for name in sheetNames2)):
                            num+=1
                            searchName = trueName+'('+str(num)+')'
                        newSheetName = searchName

                    sheetNames2.append(newSheetName)
                    sheetData = wb1.sheet_by_index(index)
                    newSheet = copyWorkbook2.add_sheet(newSheetName)

                    for row in range(sheetData.nrows):
                        for col in range(sheetData.ncols):
                            newSheet.write(row,col,sheetData.cell_value(row,col))

                copyWorkbook2.save(filePath2)

            if option == '0' or option == '2':
                copyWorkbook1.save(filePath2)
            status = True
        except Exception as e:
            err_msg = 'Error occured While copying'
            log.error(e)
        return status,err_msg

    def get_page_count_xls(self,filePath):
        """
        def : get_page_count
        purpose : get page count of .xls file
        param : filePath
        return : pagecount [int]
        """
        status=False
        err_msg=None
        count = None
        try:
            wb = open_workbook(filePath)
            count = len(wb.sheet_names())
            status = True
        except Exception as e:
            log.error(e)
            err_msg = 'Error occured fetching the page count.'
        return status,count,err_msg


class ExcelXLSX:

    def __init__(self):


        self.excel_obj=''

        self.cell_type={'string' : 'General',
        'formula': 'f',
        'number': '0.00',
        'date':'n',
        'boolean': 'b',
        'null':'n',
        'inline': 'inlineStr',
        'error': 'e',
        'formula_cache': 'str'}

        self.date_formats={'[$-14009]dd\ mmmm\ yyyy;@':'%d %B %Y',
            'mm-dd-yy':'%d-%m-%Y',
            '[$-14009]d\.m\.yy;@':'%d.%m.%y',
            'dd\/mm\/yyyy':'%d/%m/%y',
            '[$-F800]dddd\,\ mmmm\ dd\,\ yyyy':'%d %B %Y',
            '[$-14009]dd/mm/yyyy;@':'%d-%m-%Y',
            '[$-14009]dd/mm/yy;@':'%d-%m-%y',
            '[$-14009]d/m/yy;@':'%d-%m-%y',
            '[$-14009]yyyy/mm/dd;@':'%Y-%m-%d',
            '[$-14009]d\ mmmm\ yyyy;@':'%d %B %Y',
            'mmm/yyyy':'%b-%Y',
            }


    def __load_workbook_xlsx(self,inputpath,sheetname):
        """
        def : __load_workbook_xlsx(private method)
        purpose : loads the given .xlsx file
        param : inputpath,sheetname
        return : list

        """
        book=load_workbook(inputpath)
        sheet=book.get_sheet_by_name(sheetname)
        row_list=list(sheet.iter_rows())
        row=row_list[0]
        cell=row[0]
        if sheet.max_row==1 and cell.value is None:
            last_row=1
            last_col=1
        else:
            last_row=sheet.max_row
            row=row_list[len(row_list)-1]
            last_col=len(row)+1

        return book,sheet,last_row,last_col,sheet.max_row


    def __write_to_cell_xlsx(self,input_path,book,sheetname,row,col,value,*args):
        """
        def : __write_to_cell_xlsx(private method)
        purpose : writes to the the cell in the given row and column of .xlsx file
        param : input_path,book,sheetname,row,col,value
        return : bool

        """
        status=False
        self.excel_obj=ExcelFile()
        err_msg=None
        flag=False
        try:
           sheet=book.get_sheet_by_name(sheetname)
           cell=sheet.cell(row=row,column=col)
           if len(args)>0 and args[0] is not None:
            type=args[0].lower()
            if type in list(self.cell_type.keys()):
                type=self.cell_type[type]
                if type=='0.00':
                    cell.number_format=type
                    value=int(value)
                    flag=True
                elif type=='f':
                    cell.number_format='General'
                    if not value.startswith('='):
                        value='='+value
                    flag=True
                elif type=='b':          
                    cell.number_format='General'  
                    if value.lower() in ['true','false']:
                        # value=value.lower()
                        flag=True
                    else:
                        flag=False
                        status=False
                        raise ValueError
                        # err_msg="Boolean input must be provided for boolean type"
                elif type=='General':
                    cell.number_format='@'
                    flag=True
                else:
                    cell.number_format='General'
                    flag=True
                    ##value=str(value)
                # cell.value=value
                # status=True
                if flag:
                    cell.value=value
                    status=True
            else:
                logger.print_on_console('Cell Type not supported')
           else:
                cell.value=value
                status=True
                flag = True
        except ValueError as e:
            err_msg=ERROR_CODE_DICT['ERR_NUMBER_FORMAT_EXCEPTION']
        except IOError:
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']
            log.error(err_msg)
        except Exception as e:
            log.error(e)
            err_msg='Error writing to excel cell of .xlsx file'
            # logger.print_on_console(err_msg)
        book.save(input_path)
        if SYSTEM_OS == 'Windows' and flag and status:
            self.excel_obj.open_and_save_file(input_path)
        return status,err_msg


    def __get_formatted_date(self,val,fmt):
        from datetime import datetime
        if fmt in list(self.date_formats.keys()):
            return val.strftime(self.date_formats[fmt])
        return None

    def __convertStringToInt(self,col):
        num=0
        import string
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num
    def get_linenumber_xlsx(self,input_path,sheetname,content,abs_flag=False):
        """
        def : get_linenumber_xlsx
        purpose : get the line number where content is present in given .xlsx file
        param  : input_path,sheetname,content
        return : Returns line number (list)

        """
        status=False
        line_number=None
        err_msg=None
        log.info(generic_constants.INPUT_IS+input_path+' '+sheetname)
        try:
            book = load_workbook(input_path,data_only=True)
            sheet = book.get_sheet_by_name(sheetname)
            i=0
            value=[]
            if not abs_flag:
                for row in sheet.iter_rows():
                    i+=1
                    for cell in row:
                        if  cell.internal_value is not None and content == cell.internal_value:
                            value.append(i)
                            status=True
                line_number=value
            else:
                file_content=[]
                for row in sheet.iter_rows():
                    for cell in row:
                        if  cell.internal_value is not None:
                            file_content.append(cell.internal_value)
                full_file_content=''.join(map(str,file_content))
                del file_content
                content=content.replace("\t","")
                if content.replace(" ","")==full_file_content.replace(" ",""):
                    line_number=sheet.max_row
                    status=True
                    del full_file_content
        except Exception as e:
           err_msg='Error getting line number in .xlsx'
           log.error(e)
           logger.print_on_console(err_msg)
        return status,line_number,err_msg


    def replace_content_xlsx(self,input_path,sheetname,existingcontent,replacecontent,*args):
        """
        def : replace_content_xlsx
        purpose : replace the 'existingcontent' by 'replacecontent' in given sheet of
                  given .xlsx file
        param : input_path,sheetname,existingcontent,replacecontent
        return : bool

        """
        logger.print_on_console(generic_constants.INPUT_IS+input_path+' '+sheetname)
        status=False
        try:
            book = load_workbook(input_path,data_only=True)
            sheet = book.get_sheet_by_name(sheetname)
            for row in sheet.iter_rows():
                for cell in row:
                    if existingcontent ==  str(cell.value):
                        cell.value=replacecontent
                        status=True
            book.save(input_path)

        except Exception as e:
           log.error(e)
           logger.print_on_console(e)
        return status


    def delete_row_xlsx(self,row,excel_path,sheetname,*args):
        """
        def : delete_row_xlsx
        purpose : calls delete_row_xls to delete a row in given .xlsx file
        param : row
        return : bool

        """
        #same functionality as compare_content_xlsx
        obj=ExcelXLS()
        return obj.delete_row_xls(row,excel_path,sheetname,*args)


    def compare_content_xlsx(self,input_path1,sheetname1,input_path2,sheetname2,*args):
        """
        def : compare_content_xlsx
        purpose : calls compares the data of given sheets of 2 different excel files
        param : input_path1,input_path2,Sheet1,Sheet2
        return : bool

        """
        #logger.print_on_console(generic_constants.INPUT_IS+input_path1+' '+input_path2)
        obj=ExcelXLS()
        #same functionality as compare_content_xlsx
        status=obj.compare_content_xls(input_path1,sheetname1,input_path2,sheetname2)
        return status


    def verify_content_xlsx(self,input_path,sheetname,content,*args):
        """
        def : verify_content_xlsx
        purpose : verify whether the 'existingcontent' is present in given sheet of
                  given .xlsx file
        param : input_path,sheetname,content
        return : bool

        """
        status=False
        abs_flag=True if args[-1].lower()=='abs' else False
        res,line_number,err_msg=self.get_linenumber_xlsx(input_path,sheetname,content,abs_flag)
        if res and line_number != None:
            status=True
        return status,err_msg


    def clear_content_xlsx(self,inputpath,filename,sheetname):
        """
        def : clear_content_xlsx
        purpose : removes the given sheet from the given .xlsx file
        param : input_path,sheetname
        return : bool

        """
        status=False
        err_msg=None
        info_msg=generic_constants.INPUT_IS+inputpath+' filename: '+filename+' Sheetname: '+sheetname
        logger.print_on_console(info_msg)
        log.info(info_msg)
        inputpath=inputpath+'/'+filename
        try:
            import random,string
            dummy_sheet_name =''.join([random.choice(string.ascii_letters + string.digits) for n in range(10)])
            book=load_workbook(inputpath)
            sheet=book.get_sheet_by_name(sheetname)
            book.create_sheet(dummy_sheet_name)
            book.remove(sheet)
            book.active=len(book.sheetnames)-1
            book.save(inputpath)
            book.create_sheet(sheetname)
            book.save(inputpath)
            sheet=book.get_sheet_by_name(dummy_sheet_name)
            book.remove(sheet)
            book.save(inputpath)
            status=True
        except PermissionError as ex:
            err_msg = generic_constants.FILE_OPENED
            log.error(ex)
        except Exception as e:
            err_msg='Error occured in clearing the content of excel sheet'
            log.error(e)
        return status,err_msg


    def read_cell_xlsx(self,row,col,excel_path,sheetname):
        """
        def : read_cell_xlsx
        purpose : reads the cell value in the given row and column of .xls file
        param : input_path,sheetname
        return : cell value

        """
        status=False
        value=None
        err_msg=None
        wb=None
        try:
            log.debug('Writing to the file'+generic_constants.INPUT_IS+excel_path+' '+sheetname)
            wb=openpyxl.load_workbook(excel_path,data_only=True)
            sheet=wb.get_sheet_by_name(sheetname)
            value=None
            if row<=sheet.max_row:
                if col<=sheet.max_column:
                    cell=sheet.cell(row=row,column=col)
                    value=cell.value

                    if cell.is_date == True and self.__get_formatted_date(value,cell.number_format) is not None :
                        value= self.__get_formatted_date(value,cell.number_format)
                    elif cell.number_format == '0%':    ## added elif statements to detrmine whether the number is percentage and diaplay it as percentage
                        value=value*100
                        value=str(value)+'%'
                    status=True
                else:
                    err_msg=ERROR_CODE_DICT["ERR_COL_DOESN'T_EXIST"]
            else:
                err_msg=ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
        except Exception as e:
            err_msg='Error occured in read cell of .xlsx file'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,value,err_msg


    def clear_cell_xlsx(self,row,col,excel_path,sheetname):
        """
        def : clear_cell_xls
        purpose : clears the given cell of .xls file set by the user
        param : row,col
        return : bool

        """
        status=False
        err_msg=None
        log.debug(generic_constants.INPUT_IS+excel_path+' '+sheetname)
        workbook=openpyxl.load_workbook(excel_path)
        try:
            sheet=workbook.get_sheet_by_name(sheetname)
            cell=sheet.cell(row=row,column=col,value=None)
            cell.value=None
            workbook.save(excel_path)
            status=True
        except Exception as e:
            err_msg='Error occurred in clearing cell of .xlsx file'
        workbook.save(excel_path)
        return status,err_msg


    def write_cell_xlsx(self,row,col,value,excel_path,sheetname,*args):
        """
        def : write_cell_xlsx
        purpose : calls the private methods to write given value to the cell of given
                row or col of .xlsx file
        param : row,col,value
        return : bool

        """
        status=False
        err_msg=None
        try:
            #loads the xls workbook
            workbook_info=self.__load_workbook_xlsx(excel_path,sheetname)
            #writes to the cell in given row,col
            if not(workbook_info[0].read_only):
                status,err_msg=self.__write_to_cell_xlsx(excel_path,workbook_info[0],sheetname,int(row),int(col),value,*args)
            else:
                err_msg = 'Excel is readonly'
                logger.print_on_console(err_msg)
                log.info(err_msg )
        except IOError:
            err_msg=ERROR_CODE_DICT['ERR_FILE_NOT_ACESSIBLE']        
        except Exception as e:
            err_msg = 'Error writing to excel cell of .xlsx file'
            log.error(e)
            # logger.print_on_console(err_msg)
        return status,err_msg


    def get_rowcount_xlsx(self,excel_path,sheetname,col_number):
        """
        def : get_rowcount_xlsx
        purpose : get the number of rows of .xlsx excel file set by the user
        param : None
        return : Returns row count(int)

        """
        status=False
        row_count=None
        err_msg = None
        
        try:
            log.debug('Fetching row count of '+generic_constants.INPUT_IS+excel_path+' '+sheetname)
            book=openpyxl.load_workbook(excel_path)
            sheet = book.get_sheet_by_name(sheetname)       
            if col_number == None or col_number=='':
                if(sheet.max_row==1 and sheet.max_column==1 and sheet.cell(1,1).value is None):
                    row_count=0
                    status = True
                else:
                    row_count=sheet.max_row
                    status = True
            else:
                # when we need to return the row count of a speific column 
                col_number=int(col_number)
                if 1 <= col_number <= 16384:
                    try:
                        col = openpyxl.utils.get_column_letter(col_number)
                        row_count = max(i.row for i in sheet[col] if i.value is not None)
                    except ValueError:
                        row_count = 0
                    status=True    
                else:
                    err_msg = ERROR_CODE_DICT["ERR_COL_DOESN'T_EXIST"]
        except ValueError as e:
            err_msg = ERROR_CODE_DICT["ERR_COL_NUMBER"]
            log.error(e)
            logger.print_on_console(err_msg)
            
        except Exception as e:
            err_msg='Error getting the row count of .xlsx file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,row_count,err_msg


    def get_colcount_xlsx(self,excel_path,sheetname,row_number):
        """
        def : get_colcount_xls
        purpose : get the number of columns of .xls excel file set by the user
        param : None
        return : Returns row count(int)

        """
        status=False
        col_count=None
        err_msg = None
        
        try:
            log.debug('Fetching col count of '+generic_constants.INPUT_IS+excel_path+' '+sheetname)
            book=openpyxl.load_workbook(excel_path)
            sheet=book.get_sheet_by_name(sheetname)
            if row_number == None or row_number=='':
                if(sheet.max_row==1 and sheet.max_column==1 and sheet.cell(1,1).value is None):
                    col_count=0
                    status = True
                else:
                    col_count=sheet.max_column
                    status = True
            else:
                row_number=int(row_number)
                if 1<= row_number <=1048576:
                    try:
                        col_count = max(self.__convertStringToInt(i.column) for i in sheet[row_number] if i.value is not None)
                    except ValueError:
                        col_count = 0
                    status=True
                else:
                    err_msg = ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
        except ValueError as e:
            err_msg = ERROR_CODE_DICT["ERR_ROW_NUMBER"]
            log.error(e)
            logger.print_on_console(err_msg)
        except Exception as e:
            err_msg='Error getting the col count of .xlsx file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,col_count,err_msg


    def write_to_file_xlsx(self,input_path,sheetname,content,*args):
        """
        def : write_to_file_xlsx
        purpose : writes the given content to the given .xlsx file
        param : inputpath,sheetname,content,('newline'-optional param)
        return : bool

        """

        status=False
        err_msg=None
        coreutilsobj=core_utils.CoreUtils()
        sheetname=coreutilsobj.get_UTF_8(sheetname)
        log.debug(generic_constants.INPUT_IS+input_path+' '+sheetname)
        try:
            #loads the xlsx workbook

            workbook_info=self.__load_workbook_xlsx(input_path,sheetname)
            row=workbook_info[2]
            col=workbook_info[3]

            len_args=len(args)
            ##print row,col
            ##print args,'dfvgbjnml,',row,col
            ##status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,content)
            if len(args)>0 and args[len_args-1].lower() == 'newline':
                if row == 1 and col == 1:
                    status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,content)
                else:
                    row=row+1
                    col=1
                    status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,content)
                for i in range(0,(len_args-1)):
                    row=row+1
                    col=1
                    ##print row,col
                    status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,args[i])
            else:
                status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,content)
                for i in range(0,(len_args)):
                    col=col+1
                    ##print row,col
                    status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,args[i])
        except Exception as e:
            log.error(e)
            err_msg='Error occurred writing to .xlsx file '
            logger.print_on_console(err_msg)
        return status,err_msg


    def create_file_xlsx(self,excel_path,sheet_name,*args):
        """
        def : create_file_xlsx
        purpose : creates an xlsx file using provided path
        param : excel_path,sheet_name
        return : bool

        """
        status=False
        err_msg=None
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            wb.save(excel_path)
            status = True
        except Exception as e:
            log.error(e)
            err_msg='Error occurred creating .xlsx file '
            logger.print_on_console(err_msg)
        return status,err_msg

    def copy_workbook_xlsx(self,filePath1,filePath2,option):
        """
        def : copy_workbook_xlsx
        purpose : calls copy_workbook_xlsx to copy from one workbook to another
        param : filePath1,filePath2,option
        return : bool
        """
        status=False
        err_msg=None
        self.excel_obj=ExcelFile()
        try:
            wb1=openpyxl.load_workbook(filePath1)
            wb2=openpyxl.load_workbook(filePath2)
            sheetNames1 = wb1.sheetnames
            sheetNames2 = wb2.sheetnames

            if(option == '2'):
                for sheets in sheetNames2:
                    destSheet = wb2.get_sheet_by_name(sheets)
                    wb2.remove(destSheet)

            for sheet in sheetNames1:
                sheetData = wb1.get_sheet_by_name(sheet)
                destSheet = None
                if sheet.upper() not in (name.upper() for name in wb2.sheetnames):
                    wb2.create_sheet(sheet)
                    destSheet = wb2.get_sheet_by_name(sheet)

                elif(option == '0'):
                    existingName = None
                    for name in wb2.sheetnames:
                        if(name.upper() == sheet.upper()):
                            existingName = name
                            break

                    destSheet = wb2.get_sheet_by_name(existingName)
                    destSheet.delete_rows(1,destSheet.max_row)

                elif(option == '1'):
                    trueName,check,num = self.excel_obj.getSheetTrueName(sheet)
                    if(check == False):
                        trueName+=" "
                    
                    num+=1
                    searchName = trueName+'('+str(num)+')'
                    while(searchName.upper() in (name.upper() for name in wb2.sheetnames)):
                        num+=1
                        searchName = trueName+'('+str(num)+')'
                    wb2.create_sheet(searchName)
                    destSheet = wb2.get_sheet_by_name(searchName)

                for row in sheetData:
                    for cell in row:
                        destSheet[cell.coordinate].value = cell.value


            wb2.save(filePath2)
            status = True
        except Exception as e:
            log.error(e)
            err_msg='Error occurred While copying Workbook'
        return status,err_msg

    def get_page_count_xlsx(self,filePath):
        """
        def : get_page_count
        purpose : get page count of .xlsx file
        param : filePath
        return : pagecount [int]
        """
        status=False
        err_msg=None
        count = None
        try:
            wb = openpyxl.load_workbook(filePath)
            count = len(wb.sheetnames)
            status = True
        except Exception as e:
            log.error(e)
            err_msg = 'Error occured fetching the page count.'
        return status,count,err_msg



class ExcelCSV:

    def __init__(self):


        self.excel_obj=''

        self.cell_type={'string' : 'General',
        'formula': 'f',
        'number': '0.00',
        'date':'n',
        'boolean': 'b',
        'null':'n',
        'inline': 'inlineStr',
        'error': 'e',
        'formula_cache': 'str'}

        self.date_formats={'[$-14009]dd\ mmmm\ yyyy;@':'%d %B %Y',
            'mm-dd-yy':'%d-%m-%Y',
            '[$-14009]d\.m\.yy;@':'%d.%m.%y',
            'dd\/mm\/yyyy':'%d/%m/%y',
            '[$-F800]dddd\,\ mmmm\ dd\,\ yyyy':'%d %B %Y',
            '[$-14009]dd/mm/yyyy;@':'%d-%m-%Y',
            '[$-14009]dd/mm/yy;@':'%d-%m-%y',
            '[$-14009]d/m/yy;@':'%d-%m-%y',
            '[$-14009]yyyy/mm/dd;@':'%Y-%m-%d',
            '[$-14009]d\ mmmm\ yyyy;@':'%d %B %Y',
            'mmm/yyyy':'%b-%Y',
            }


    def __get_formatted_date(self,val,fmt):
        from datetime import datetime
        if fmt in list(self.date_formats.keys()):
            return val.strftime(self.date_formats[fmt])
        return None


    def get_linenumber_csv(self,input_path,sheetname,content):
        """
        def : get_linenumber_xlsx
        purpose : get the line number where content is present in given .xlsx file
        param  : input_path,sheetname,content
        return : Returns line number (list)

        """
        status=False
        line_number=None
        err_msg=None
        logger.print_on_console(generic_constants.INPUT_IS+input_path+' '+sheetname)
        try:
            book = load_workbook(input_path,data_only=True)
            sheet = book.get_sheet_by_name(sheetname)
            i=0
            value=[]
            for row in sheet.iter_rows():
                i+=1
                for cell in row:
                    if  cell.internal_value is not None and content == cell.internal_value:
                        value.append(i)
                        status=True
            line_number=value
        except Exception as e:
           err_msg='Error getting line number in .xlsx'
           log.error(e)
           logger.print_on_console(err_msg)
        return status,line_number,err_msg


    def replace_content_csv(self,input_path,sheetname,existingcontent,replacecontent,*args):
        """
        def : replace_content_xlsx
        purpose : replace the 'existingcontent' by 'replacecontent' in given sheet of
                  given .xlsx file
        param : input_path,sheetname,existingcontent,replacecontent
        return : bool

        """
        logger.print_on_console(generic_constants.INPUT_IS+input_path+' '+sheetname)
        status=False
        try:
            book = load_workbook(input_path,data_only=True)
            sheet = book.get_sheet_by_name(sheetname)
            for row in sheet.iter_rows():
                for cell in row:
                    if existingcontent ==  str(cell.value):
                        cell.value=replacecontent
                        status=True
            book.save(input_path)

        except Exception as e:
           log.error(e)
           logger.print_on_console(e)
        return status


    def delete_row_csv(self,row,excel_path,sheetname):
        """
        def : delete_row_xlsx
        purpose : calls delete_row_xls to delete a row in given .xlsx file
        param : row
        return : bool

        """
        status=False
        err_msg=None
        log.debug(generic_constants.INPUT_IS+excel_path)
        sheet = []
        try:
            with open(excel_path) as f:
                reader = csv.reader(f)
                for _row in reader:
                    sheet.append(_row)
                row_count = len(sheet)
                col_count = len(sheet[0]) if row_count>0 else 0
                if row<=row_count:
                    del sheet[row-1]
                else:
                    err_msg=ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
            with open(excel_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(sheet)
            del sheet
            status=True
        except Exception as e:
            err_msg='Error occurred in clearing cell of .csv file'
        return status,err_msg


    def compare_content_csv(self,input_path1,input_path2,*args): # To compare the content for verifying data for csv files
        """
        def : compare_content_csv
        purpose : compares the data of given sheets of 2 different excel files
        param : input_path1,input_path2,Sheet1,Sheet2
        return : bool

        """
        status=False
        err_msg=None
        x = False
        log.debug('Comparing content of .csv files')
        try:
            data_file1=[]
            data_file2=[]
            x = False
            status = False
            if len(args)>0 and input_path2=='':
                input_path2=args[0]
            try :
                with open(input_path1 , 'rt') as file1:
                    reader =csv.reader(file1)
                    for i in reader:
                        data_file1.append(i)
                with open(input_path2 ,'rt') as file2:
                    reader =csv.reader(file2)
                    for i in reader:
                        data_file2.append(i)
            except Exception as e:
                logger.print_on_console(e)
            finally:
                file1.close()
                file2.close()
            try:
                for file1_row, file2_row in zip(data_file1, data_file2):
                    if file1_row == file2_row:
                        x = True
                        break
                if(x):
                    status = True
                else:
                    status = False

            except Exception as e:
                log.error(e)
                err_msg = str(e)

        except Exception as e:
            err_msg='Error occured in compare content of .csv files'
            log.error(e)

        log.info('Status is '+str(status))
        return status,err_msg


    def verify_content_csv(self,input_path,sheetname,content,*args):
        """
        def : verify_content_xlsx
        purpose : verify whether the 'existingcontent' is present in given sheet of
                  given .xlsx file
        param : input_path,sheetname,content
        return : bool

        """
        status=False
        res,line_number,err_msg=self.get_linenumber_xlsx(input_path,sheetname,content)
        if res and line_number != None:
            status=True
        return status,err_msg


    def clear_content_csv(self,inputpath,filename,sheetname):
        """
        def : clear_content_xlsx
        purpose : removes the given sheet from the given .xlsx file
        param : input_path,sheetname
        return : bool

        """
        status=False
        err_msg=None
        info_msg=generic_constants.INPUT_IS+inputpath+' filename: '+filename+' Sheetname: '+sheetname
        logger.print_on_console(info_msg)
        log.info(info_msg)
        inputpath=inputpath+'/'+filename
        try:
            import random,string
            dummy_sheet_name =''.join([random.choice(string.ascii_letters + string.digits) for n in range(10)])
            book=load_workbook(inputpath)
            sheet=book.get_sheet_by_name(sheetname)
            book.create_sheet(dummy_sheet_name)
            book.remove(sheet)
            book.active=len(book.sheetnames)-1
            book.save(inputpath)
            book.create_sheet(sheetname)
            book.save(inputpath)
            sheet=book.get_sheet_by_name(dummy_sheet_name)
            book.remove(sheet)
            book.save(inputpath)
            status=True
        except Exception as e:
            err_msg='Error occured in clearing the content of excel sheet'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,err_msg


    def read_cell_csv(self,row,col,excel_path,sheetname,*args):
        """
        def : read_cell_csv
        purpose : reads the cell value in the given row and column of .csv file
        param : input_path,sheetname
        return : cell value

        """
        status=False
        value=None
        err_msg=None
        sheet = []
        try:
            log.debug('Writing to the file'+generic_constants.INPUT_IS+excel_path)
            with open(excel_path) as f:
                reader = csv.reader(f, delimiter=',')
                for _row in reader:
                    sheet.append(_row)
                row_count = len(sheet)
                col_count = len(sheet[0]) if row_count>0 else 0
                if row<=row_count:
                    if col<=col_count:
                        value = (sheet[row-1][col-1])
                        status=True
                    else:
                        err_msg=ERROR_CODE_DICT["ERR_COL_DOESN'T_EXIST"]
                else:
                    err_msg=ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
            del sheet
        except Exception as e:
            err_msg='Error occured in read cell of .csv file'
            log.error(e)
        if err_msg!=None:
            log.error(err_msg)
            logger.print_on_console(err_msg)
        return status,value,err_msg


    def clear_cell_csv(self,row,col,excel_path,sheetname):
        """
        def : clear_cell_csv
        purpose : clears the given cell of .csv file set by the user
        param : row,col
        return : bool

        """
        status=False
        err_msg=None
        log.debug(generic_constants.INPUT_IS+excel_path)
        sheet = []
        try:
            with open(excel_path) as f:
                reader = csv.reader(f)
                for _row in reader:
                    sheet.append(_row)
                row_count = len(sheet)
                col_count = len(sheet[0]) if row_count>0 else 0
                if row<=row_count:
                    if col<=col_count:
                        sheet[row-1][col-1] = ''
                    else:
                        err_msg=ERROR_CODE_DICT["ERR_COL_DOESN'T_EXIST"]
                else:
                    err_msg=ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
            with open(excel_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(sheet)
            del sheet
            status=True
        except Exception as e:
            err_msg='Error occurred in clearing cell of .csv file'
        return status,err_msg


    def write_cell_csv(self,row,col,value,excel_path,sheetname,*args):
        """
        def : write_cell_csv
        purpose : calls the private methods to write given value to the cell of given
                row or col of .csv file
        param : row,col,value
        return : bool

        """
        status=False
        err_msg=None
        sheet = []
        try:
            with open(excel_path) as f:
                reader = csv.reader(f)
                for _row in reader:
                    sheet.append(_row)
                row_count = len(sheet)
                col_count = len(sheet[0]) if row_count>0 else 0
                if row<=row_count:
                    if col<=col_count:
                        sheet[row-1][col-1] = value
                    else:
                        err_msg=ERROR_CODE_DICT["ERR_COL_DOESN'T_EXIST"]
                else:
                    err_msg=ERROR_CODE_DICT["ERR_ROW_DOESN'T_EXIST"]
            with open(excel_path, 'w', newline='') as f:
                writer = csv.writer(f)
                # for row in sheet:
                writer.writerows(sheet)
            del sheet
            status=True
        except Exception as e:
            err_msg = 'Error writing to excel cell of .csv file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,err_msg


    def get_rowcount_csv(self,excel_path,sheetname,*args):
        """
        def : get_rowcount_csv
        purpose : get the number of rows of .csv excel file set by the user
        param : None
        return : Returns row count(int)

        """
        status=False
        row_count=None
        err_msg=None
        try:
            log.debug('Fetching row count of '+generic_constants.INPUT_IS+excel_path)
            with open(excel_path) as f:
                row_count = sum(1 for row in f)
            status=True
        except Exception as e:
            err_msg='Error getting the row count of .csv file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,row_count,err_msg


    def get_colcount_csv(self,excel_path,sheetname,*args):
        """
        def : get_colcount_csv
        purpose : get the number of columns of .csv excel file set by the user
        param : None
        return : Returns row count(int)

        """
        status=False
        col_count=None
        err_msg=None
        try:
            log.debug('Fetching col count of '+generic_constants.INPUT_IS+excel_path)
            with open(excel_path) as f:
                col_count = sum(1 for col in f.readline().split(','))
            status=True
        except Exception as e:
            err_msg='Error getting the col count of .csv file'
            log.error(e)
            logger.print_on_console(err_msg)
        return status,col_count,err_msg


    def write_to_file_csv(self,input_path,sheetname,content,*args):
        """
        def : write_to_file_xlsx
        purpose : writes the given content to the given .xlsx file
        param : inputpath,sheetname,content,('newline'-optional param)
        return : bool

        """

        status=False
        err_msg=None
        coreutilsobj=core_utils.CoreUtils()
        sheetname=coreutilsobj.get_UTF_8(sheetname)
        log.debug(generic_constants.INPUT_IS+input_path+' '+sheetname)
        try:
            #loads the xlsx workbook

            workbook_info=self.__load_workbook_xlsx(input_path,sheetname)
            row=workbook_info[2]
            col=workbook_info[3]

            len_args=len(args)
            ##print row,col
            ##print args,'dfvgbjnml,',row,col
            ##status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,content)
            if len(args)>0 and args[len_args-1].lower() == 'newline':
                if row == 1 and col == 1:
                    status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,content)
                else:
                    row=row+1
                    col=1
                    status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,content)
                for i in range(0,(len_args-1)):
                    row=row+1
                    col=1
                    ##print row,col
                    status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,args[i])
            else:
                status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,content)
                for i in range(0,(len_args)):
                    col=col+1
                    ##print row,col
                    status,err_msg=self.__write_to_cell_xlsx(input_path,workbook_info[0],sheetname,row,col,args[i])
        except Exception as e:
            log.error(e)
            err_msg='Error occurred writing to .xlsx file '
            logger.print_on_console(err_msg)
        return status,err_msg


    def create_file_csv(self,excel_path,sheet_name,*args):
        """
        def : create_file_csv
        purpose : creates an csv file using provided path
        param : excel_path,sheet_name
        return : bool

        """
        status=False
        err_msg=None
        try:
            # from openpyxl import Workbook
            # wb = Workbook()
            # ws = wb.active
            # ws.title = sheet_name
            # wb.save(excel_path)
            with open(excel_path, 'w') as f:
                pass
            status = True
        except Exception as e:
            log.error(e)
            err_msg='Error occurred creating .csv file '
            logger.print_on_console(err_msg)
        return status,err_msg



class object_creator:
    def excel_object(self):
        excel = None
        try:
            import pythoncom
            pythoncom.CoInitialize()
            try:
                excel = win32com.client.dynamic.Dispatch("Excel.Application")
            except Exception as e:
                log.info("Error in dynamic.Dispatch proceeding with gencache.EnsureDispatc, err_msg : ",e)
                excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
        except Exception as e:
            log.info("EnsureDispatch and Dispatch failed, the error is : ",e)
        return excel
