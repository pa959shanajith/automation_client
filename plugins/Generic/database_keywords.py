#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      rakesh.v
#
# Created:     20-10-2016
# Copyright:   (c) rakesh.v 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------

##prerequistes - Respective database drivers should be avialable in client machine
import csv
import file_operations
import excel_operations
import xlwt
from xlutils.copy import copy as xl_copy
import openpyxl
from openpyxl.utils import get_column_letter
import xlrd
from xlrd import open_workbook
import generic_constants
import os
import logger
from encryption_utility import AESCipher
import logging
import dynamic_variable_handler
import constant_variable_handler
try:
    import pyodbc
    import jaydebeapi
    import jpype
except:
    pass

from constants import *
log = logging.getLogger('database_keywords.py')


class DatabaseOperation():

    def __init__(self):
        self.DV = dynamic_variable_handler.DynamicVariables()
        self.CV = constant_variable_handler.ConstantVariables()

    def processException(self, e):
        etype = type(e)
        ae = None
        if etype == pyodbc.OperationalError:
            if hasattr(e, 'args') and len(e.args) > 1: ae = e.args[1].split(';')[0]
            err_msg = ERROR_CODE_DICT["ERR_DB_CONNECTION"]
        elif etype in [pyodbc.IntegrityError, pyodbc.InterfaceError, pyodbc.InternalError]:
            if hasattr(e, 'args') and len(e.args) > 1: ae = e.args[1].split(';')[0]
            err_msg = ERROR_CODE_DICT["ERR_DB_OPS"]
        elif etype in [pyodbc.DataError, pyodbc.ProgrammingError, pyodbc.NotSupportedError]:
            if hasattr(e, 'args') and len(e.args) > 1: ae = e.args[1].split(';')[0]
            err_msg = ERROR_CODE_DICT["ERR_DB_QUERY"]
        ##22200: Invalid sheet error for verifydata
        elif etype == xlrd.biffh.XLRDError:
            err_msg = ERROR_CODE_DICT["ERR_INVALID_SHEET"]

        ##to handle mdb files exception     
        elif etype == pyodbc.Error:
            if hasattr(e, 'args') and len(e.args) > 1: ae = e.args[1].split(';')[-1]
            if ('password' in str(ae)):
                err_msg=generic_constants.INVALID_INPUT
            elif('file' in str(ae)):
                err_msg="Invalid DB file."
            else:
                err_msg = ERROR_CODE_DICT['ERR_DB_OPS']
        elif etype == jpype._jvmfinder.JVMNotFoundException:
            if hasattr(e, 'args'):
                ae = e.args[0].split('.')[2].strip()
            err_msg = ERROR_CODE_DICT['ERR_JAVA_NOT_FOUND']
        else:
            err_msg = ERROR_CODE_DICT['ERR_DB_OPS']
        logger.print_on_console(err_msg) 
        log.error(err_msg)
        if ae is not None:
            # logger.print_on_console(ae) 
            log.error(ae)
            err_msg = ae
        log.error(e)
        return err_msg

    def mongoquery(self, query,q_type):
        import json
        from bson.json_util import dumps,loads
        flag=False
        query=json.loads(query)
        if q_type == 'insert_many':
            self.coll.insert_many(query)
        elif q_type == 'insert_one':
            self.coll.insert_one(query)
        elif q_type == 'update_one':
            self.coll.update_one(query)
        elif q_type == 'update_many':
            self.coll.update_many(query)
        elif q_type == 'delete_one':
            self.coll.delete_one(query)
        elif q_type == 'delete_many':
            self.coll.delete_many(query)
        elif q_type == 'find_one':
            res = self.coll.find_one(query)
            flag=True
        elif q_type == 'find':
            res = self.coll.find(query)
            flag=True
        if flag:
            json_rows = loads(dumps(res))
            columns = [k for k,v in sorted(json_rows[0].items())]
            columns=tuple(columns)
            rows=[]
            for i in json_rows:
                row=[]
                for k,v in sorted(i.items()):
                    row.append(str(v))
                row=tuple(row)
                rows.append(row)
            return rows,columns

    def redisquery(self, cnxn, query, q_type):
        if q_type == 'keys':
            keys=cnxn.keys(query)
            rows=[]
            for i in keys:
                key=i
                val=cnxn.get(i)
                if isinstance(key,bytes):
                    key=key.decode('utf-8')
                if isinstance(val,bytes):
                    val=val.decode('utf-8')
                rows.append(tuple([key,val]))
            columns=tuple(['key','value'])
            return rows,columns
        if q_type=='get':
            res=cnxn.get(query)
            row=[query,res]
        elif q_type=='set':
            val=query.split(',')
            if len(val)>=2:
                val[1]=','.join(val[1:])
                cnxn.set(val[0],val[1])
        

    def runQuery(self, ip , port , userName , password, dbName, query, dbtype, *args):
        """
        def : runQuery
        purpose : Executes the query
        param : database type, IP, port number , database name, username , password , query
        return : boolean
        """
        status=generic_constants.TEST_RESULT_FAIL
        result=generic_constants.TEST_RESULT_FALSE
        verb = OUTPUT_CONSTANT
        err_msg=None
        cursor = None
        try:
            cnxn = self.connection(dbtype, ip , port , dbName, userName , password, args)
            if cnxn is not None:
                dbtype=int(dbtype)
                ## execute query in cassandra
                if dbtype == 11:
                    cnxn.execute(query)
                    status=generic_constants.TEST_RESULT_PASS
                    result=generic_constants.TEST_RESULT_TRUE
                    cnxn=cnxn.shutdown()
                ## execute query in mongo
                elif dbtype == 9:
                    db=cnxn[dbName]
                    if len(args)==2:
                        self.coll=db[args[0]]
                        self.mongoquery(query,args[1])
                        status=generic_constants.TEST_RESULT_PASS
                        result=generic_constants.TEST_RESULT_TRUE
                ## execute query in redis
                elif dbtype == 10:
                    self.redisquery(cnxn,query,args[0])
                    status=generic_constants.TEST_RESULT_PASS
                    result=generic_constants.TEST_RESULT_TRUE
                else:
                    cursor = cnxn.cursor()
                    statement = ['create','update','insert','drop','delete','CREATE','UPDATE','INSERT','DROP','DELETE']
                    if any(x in query for x in statement ):
                        log.debug('Inside IF condition')
                        cursor.execute(query)
                        chk_var=['update','UPDATE','delete','DELETE']
                        if(any(x in query for x in chk_var) and cursor.rowcount==0):
                            status=generic_constants.TEST_RESULT_FAIL
                            result=generic_constants.TEST_RESULT_FALSE
                        else:
                            cnxn.commit()
                            status=generic_constants.TEST_RESULT_PASS
                            result=generic_constants.TEST_RESULT_TRUE
                    else:
                        log.debug('Inside else condition')
                        cursor.execute(query)
                        status=generic_constants.TEST_RESULT_PASS
                        result=generic_constants.TEST_RESULT_TRUE
        except Exception as e:
            err_msg = self.processException(e)
        finally:
            if cursor is not None: cursor.close()
            if cnxn is not None: cnxn.close()
        return status,result,verb,err_msg

    def secureRunQuery(self, ip , port , userName , password, dbName, query, dbtype, *args):
        """
        def : secureRunQuery
        purpose : Executes the query
        param : database type, IP, port number , database name, username , password , query
        return : boolean
        """
        status=generic_constants.TEST_RESULT_FAIL
        result=generic_constants.TEST_RESULT_FALSE
        verb = OUTPUT_CONSTANT
        err_msg=None
        try:
            encryption_obj = AESCipher()
            decrypted_password = encryption_obj.decrypt(password)
            status,result,verb,err_msg=self.runQuery(ip,port,userName,decrypted_password,dbName,query,dbtype,args)
        except Exception as e:
            err_msg = self.processException(e)
        return status,result,verb,err_msg


    def getData(self, ip , port , userName , password, dbName, query, dbtype, *args):
        """
        def : getData
        purpose : Executes the query and gets the data
        param : database type, IP, port number , database name, username , password , query
        return : data
        """
        status=generic_constants.TEST_RESULT_FAIL
        result=generic_constants.TEST_RESULT_FALSE
        value = None
        err_msg=None
        details = []
        res=[]
        res.append(ip)
        res.append(port)
        res.append(userName)
        res.append(password)
        res.append(dbName)
        res.append(query)
        res.append(dbtype) 
        cnxn=None
        cursor=None
        try:
            cnxn = self.connection(dbtype, ip , port , dbName, userName , password, args)
            if cnxn is not None:
                dbtype=int(dbtype)
                if dbtype == 13:
                    res.append((args[0],args[1]))
                ## execute query in cassandra 
                if dbtype == 11: 
                    cnxn.execute(query)
                    status=generic_constants.TEST_RESULT_PASS
                    result=generic_constants.TEST_RESULT_TRUE
                    cnxn=cnxn.shutdown()
                ## execute query in mongo
                elif dbtype == 9:
                    db=cnxn[dbName]
                    if len(args)>=2:
                        self.coll=db[args[0]]
                        rows,columns=self.mongoquery(query,args[1])
                        status=generic_constants.TEST_RESULT_PASS
                        result=generic_constants.TEST_RESULT_TRUE
                        res.append(args[0])
                        res.append(args[1])
                ## execute query in redis
                elif dbtype == 10:
                    rows,columns=self.redisquery(cnxn,query,args[0])
                    res.append(args[0])
                    status=generic_constants.TEST_RESULT_PASS
                    result=generic_constants.TEST_RESULT_TRUE
                ##added to check if getalltables is not used with any other dbtype
                else:
                    cursor = cnxn.cursor()
                    if (query.lower() =='getalltables' and dbtype!='3'):
                        err_msg = ERROR_CODE_DICT["ERR_DB_QUERY"]
                        logger.print_on_console(err_msg)
                        log.error(err_msg)
                    else:
                        if query.lower() !='getalltables':
                            ##added to check if given query is valid
                            cursor.execute(query)
                            status=generic_constants.TEST_RESULT_PASS
                            result=generic_constants.TEST_RESULT_TRUE
                details=res
        except Exception as e:
            err_msg = self.processException(e)
        finally:
            if cnxn is not None: cnxn.close()
        return status,result,details,err_msg


    def fetchData(self,input_val,*args):
        value = None
        cursor = None
        if(len(input_val)>= 8):
            try:
                cnxn = self.connection(input_val[6], input_val[0] , input_val[1] , input_val[4], input_val[2] , input_val[3],input_val[7])
                dbtype=int(input_val[6])
                if cnxn is not None:
                    import re
                    data = re.findall(r'\[([^]]*)\]',input_val[-1])
                    ##if condition added to resolve invalid input in displaydynamicvariable when more then two indexes were entered
                    if(len(data)==2):
                        row = data[0]
                        col = data[1]
                        row = int(row)
                        col = int(col)
                        row_no = row - 1
                        col_no = col - 1
                        if dbtype == 11: 
                            from cassandra.query import dict_factory
                            import uuid
                            cnxn.row_factory = dict_factory
                            query_res = cnxn.execute(input_val[5])
                            columns = tuple(query_res.column_names)
                            json_rows = query_res.current_rows
                            rows=[]
                            for i in json_rows:
                                row=[]
                                for k,v in i.items():
                                    if isinstance(v,uuid.UUID):
                                        row.append(str(v))
                                    else:
                                        row.append(v)
                                row=tuple(row)
                                rows.append(row)
                            value=rows
                            cnxn=cnxn.shutdown()
                        ## execute query in mongo
                        elif dbtype == 9:
                            db=cnxn[input_val[4]]
                            self.coll=db[input_val[7]]
                            rows,columns=self.mongoquery(input_val[5],input_val[8])
                            value=rows
                        ## execute query in redis
                        elif dbtype == 10:
                            rows,columns=self.redisquery(cnxn,input_val[5],input_val[7])
                            value=rows
                        else:
                            cursor = cnxn.cursor()
                            ##fetching table names
                            if (input_val[5].lower()=='getalltables'):
                                rows=[]
                                tables=[]
                                for table in list(cursor.tables()):
                                    if table[3]=='TABLE':
                                        tables.append(table[2])
                                rows.append(tables)
                                value = rows[col_no][row_no]
                            else:
                                cursor.execute(input_val[5])
                                rows = cursor.fetchall()
                        value = rows[row_no][col_no]
                        ##if condition added to fix issue:304-Generic : getData keyword:  Actual data  is not getting stored in dynamic variable instead "null" is stored.
                        ##changes done by jayashree.r
                    if value == None:
                        value = 'None'
                    log.info('Value obtained :')
                    log.info(value)
            except Exception as e:
                self.processException(e)
            finally:
                if cursor is not None: cursor.close()
                if cnxn is not None: cnxn.close()
            return value

    def secureGetData(self, ip , port , userName , password, dbName, query, dbtype,*args):
        """
        def : secureGetData
        purpose : Executes the query and gets the data
        param : database type, IP, port number , database name, username , password , query
        return : data
        """
        status=generic_constants.TEST_RESULT_FAIL
        result=generic_constants.TEST_RESULT_FALSE
        verb = OUTPUT_CONSTANT
        err_msg=None
        try:
            encryption_obj = AESCipher()
            decrypted_password = encryption_obj.decrypt(password)
            temp = args
            status,result,value,err_msg=self.getData(ip,port,userName,decrypted_password,dbName,query,dbtype,temp)
        except Exception as e:
            err_msg = self.processException(e)
        return status,result,value,err_msg

    def verifyData(self, ip , port , userName , password, dbName, query, dbtype,inp_file,inp_sheet,*args):
        """
        def : verifyData
        purpose : Executes the query and compares the data with data in files
        param : database type, IP, port number , database name, username , password , query , Filename , sheetname
        return : bool
        """
        status=generic_constants.TEST_RESULT_FAIL
        result=generic_constants.TEST_RESULT_FALSE
        verb = OUTPUT_CONSTANT
        err_msg=None
        cursor=None
        sheet=None
        try:
            ext = self.get_ext(inp_file)
            if ext == '.xls':
                file_path=self.create_file_xls()
            elif ext == '.xlsx':
                file_path=self.create_file_xlsx()
            else:
                file_path=self.create_file_csv()

            cnxn = self.connection(dbtype, ip , port , dbName, userName , password,args)
            if cnxn is not None:
                dbtype=int(dbtype)
                ## execute query in cassandra
                if dbtype == 11: 
                    from cassandra.query import dict_factory
                    import uuid
                    cnxn.row_factory = dict_factory
                    query_res = cnxn.execute(query)
                    columns = tuple(query_res.column_names)
                    json_rows = query_res.current_rows
                    rows=[]
                    for i in json_rows:
                        row=[]
                        for k,v in i.items():
                            if isinstance(v,uuid.UUID):
                                row.append(str(v))
                            else:
                                row.append(v)
                        row=tuple(row)
                        rows.append(row)
                    cnxn=cnxn.shutdown()
                ## execute query in mongo
                elif dbtype == 9:
                    db=cnxn[dbName]
                    if len(args)>=2:
                        self.coll=db[args[0]]
                        rows,columns=self.mongoquery(query,args[1])
                ## execute query in redis
                elif dbtype == 10:
                    rows,columns=self.redisquery(cnxn,query,args[0])
                else: 
                    cursor = cnxn.cursor()
                    cursor.execute(query)
                    rows = cursor.fetchall()
                    columns = [column[0] for column in cursor.description]
                verify = os.path.isfile(inp_file)
                if(verify== True and ext in[".xls",".xlsx"]):
                    book = open_workbook(inp_file)
                    sheet = book.sheet_by_name(inp_sheet)
                if (verify == True and (type(sheet)==xlrd.sheet.Sheet or sheet == None)):
                    ext = self.get_ext(file_path)
                    if (ext == '.xls'):
                        work_book = xlwt.Workbook(encoding="utf-8")
                        work_sheet = work_book.add_sheet(generic_constants.DATABASE_SHEET)
                        i=0
                        j=0
                        for x in columns:
                            work_sheet.write(i,j,x)
                            j=j+1
                        b=0
                        k=1
                        for row in rows:
                            l=0
                            for y in row:
                                try:
                                    work_sheet.write(k,l,y)
                                except Exception as e:
                                    if 'character' in str(e):
                                        i = str(e).index('character')
                                        err_new=str(e)[:i+len('character')]
                                        if err_new==generic_constants.UNICODE_ERR:
                                            newrow=[]
                                            for ele in row:
                                                if type(ele) == str:
                                                     ele = ele.encode('utf-8')
                                                     newrow.append(ele)
                                                else:
                                                    newrow.append(ele)
                                        row = tuple(newrow)
                                        work_sheet.write(k,l,y)
                                l+=1
                            k+=1
                        work_book.save(file_path)
                        obj = file_operations.FileOperations()
                        output = obj.compare_content(file_path,generic_constants.DATABASE_SHEET,inp_file,inp_sheet)
                        if output[1] == "True":
                            status=generic_constants.TEST_RESULT_PASS
                            result=generic_constants.TEST_RESULT_TRUE
                        else:
                            status=generic_constants.TEST_RESULT_FAIL
                            result=generic_constants.TEST_RESULT_FALSE
                    elif (ext == '.csv'):
                        try:
                            columns = [column[0] for column in cursor.description]
                            path = file_path
                            with open(path,'w') as csvfile:
                                writer = csv.writer(csvfile,lineterminator='\n')
                                writer.writerow(columns)

                                for row in rows:
                                    try:
                                        writer.writerow(row)
                                        num=num+1
                                    except Exception as e:
                                        if 'character' in str(e):
                                            i = str(e).index('character')
                                            err_new=str(e)[:i+len('character')]
                                            if err_new==generic_constants.UNICODE_ERR:
                                                newrow=[]
                                                for ele in row:
                                                    if type(ele) == str:
                                                         ele = ele.encode('utf-8')
                                                         newrow.append(ele)
                                                    else:
                                                        newrow.append(ele)
                                            row = tuple(newrow)
                                            writer.writerow(row)
                            csvfile.close()
                        except Exception as e:
                            err_msg = "Error while writing to csv"
                            log.error(err_msg)
                            log.error(e)
                        obj = file_operations.FileOperations()
                        output = obj.compare_content(file_path,inp_file)
                        if output[1] == "True":
                            status=generic_constants.TEST_RESULT_PASS
                            result=generic_constants.TEST_RESULT_TRUE
                        else:
                            status=generic_constants.TEST_RESULT_FAIL
                            result=generic_constants.TEST_RESULT_FALSE
                    ## added to resolve file compare issue with data fetched from exportdata keyword
                    elif (ext =='.xlsx'):
                        #Creating the New File.
                        work_book = openpyxl.Workbook()
                        #adding the New Sheet.
                        index_sheet = 0
                        work_book.create_sheet(index=index_sheet, title=generic_constants.DATABASE_SHEET)
                        i=1
                        j=1
                        work_sheet = work_book[generic_constants.DATABASE_SHEET]
                        for x in columns:
                            work_sheet.cell(row=i, column=j).value = x
                            j=j+1
                            try:
                                max_col_width = 2.4
                                adjusted_width = (len(str(x)) + 2) * 1.2
                                if adjusted_width > max_col_width:
                                    max_col_width=adjusted_width
                                work_sheet.column_dimensions[get_column_letter(j)].width = max_col_width
                            except Exception as e:
                                log.info("Setting the dimensions for the column: {}".format(e))
                        try:
                            del x,columns,j,i,max_col_width,adjusted_width
                        except Exception as e :
                            log.error('some error : {}'.format(e))
                        k=2
                        for row in rows:
                            l=1
                            for y in row:
                                try:
                                    work_sheet.cell(row=k, column=l).value = y   
                                except Exception as e:
                                    if 'character' in str(e):
                                        i = str(e).index('character')
                                        err_new=str(e)[:i+len('character')]
                                        if err_new==generic_constants.UNICODE_ERR:
                                            newrow=[]
                                            for ele in row:
                                                if type(ele) == str:
                                                    ele = ele.encode('utf-8')
                                                    newrow.append(ele)
                                                else:
                                                    newrow.append(ele)
                                            del ele
                                        row = tuple(newrow)
                                        work_sheet.cell(row=k, column=l).value = y
                                l+=1
                                try:
                                    max_col_width = 2.4
                                    adjusted_width = (len(str(y)) + 2) * 1.2
                                    if adjusted_width > max_col_width:
                                        max_col_width=adjusted_width
                                    work_sheet.column_dimensions[get_column_letter(l)].width = max_col_width
                                except Exception as e:
                                    log.info("Setting the dimensions for the rows: {}".format(e))
                            k+=1
                        work_book.save(file_path)
                        obj = file_operations.FileOperations()
                        output = obj.compare_content(file_path,generic_constants.DATABASE_SHEET,inp_file,inp_sheet)
                        if output[1] == "True":
                            status=generic_constants.TEST_RESULT_PASS
                            result=generic_constants.TEST_RESULT_TRUE
                        else:
                            status=generic_constants.TEST_RESULT_FAIL
                            result=generic_constants.TEST_RESULT_FALSE

                    else:
                        err_msg = ERROR_CODE_DICT['ERR_INVALID_INPUT']
                        logger.print_on_console(err_msg)
                        log.info(err_msg)
                else:
                    logger.print_on_console(generic_constants.FILE_NOT_EXISTS)
        except Exception as e:
            err_msg = self.processException(e)
        finally:
            os.remove(file_path)
            if cursor is not None: cursor.close()
            if cnxn is not None: cnxn.close()
        return status,result,verb,err_msg

    def secureVerifyData(self, ip , port , userName , password, dbName, query, dbtype,inp_file,inp_sheet, *args):
        """
        def : secureVerifyData
        purpose : Executes the query and compares the data with data in files
        param : database type, IP, port number , database name, username , password , query , Filename , sheetname
        return : bool
        """
        status=generic_constants.TEST_RESULT_FAIL
        result=generic_constants.TEST_RESULT_FALSE
        verb = OUTPUT_CONSTANT
        err_msg=None
        try:
            encryption_obj = AESCipher()
            decrypted_password = encryption_obj.decrypt(password)
            status,result,verb,err_msg=self.verifyData(ip,port,userName,decrypted_password,dbName, query, dbtype,inp_file,inp_sheet,args)
        except Exception as e:
            err_msg = self.processException(e)
        return status,result,verb,err_msg

    def exportData(self, ip , port , userName , password, dbName, query, dbtype, *args):
        """
        def : exportData
        purpose : Executes the query and exports the data to excel or csv file
        param : database type, IP, port number , database name, username , password , query , Filename , sheetname
        return : bool
        """
        status=generic_constants.TEST_RESULT_FAIL
        result=generic_constants.TEST_RESULT_FALSE
        verb = OUTPUT_CONSTANT
        err_msg=None
        cursor = None
        try:
            cnxn = self.connection(dbtype, ip , port , dbName, userName , password,args)
            if cnxn is not None:
                dbtype=int(dbtype)
                ## execute query in cassandra
                if dbtype == 11: 
                    from cassandra.query import dict_factory
                    import uuid
                    cnxn.row_factory = dict_factory
                    query_res = cnxn.execute(query)
                    columns = tuple(query_res.column_names)
                    json_rows = query_res.current_rows
                    rows=[]
                    for i in json_rows:
                        row=[]
                        for k,v in i.items():
                            if isinstance(v,uuid.UUID):
                                row.append(str(v))
                            else:
                                row.append(v)
                        row=tuple(row)
                        rows.append(row)
                    cnxn=cnxn.shutdown()
                ## execute query in mongo
                elif dbtype == 9:
                    db=cnxn[dbName]
                    if len(args)>=2:
                        self.coll=db[args[0]]
                        rows,columns=self.mongoquery(query,args[1])
                        args=args[2:]
                ## execute query in redis
                elif dbtype == 10:
                    if len(args)>=2:
                        rows,columns=self.redisquery(cnxn,query,args[0])
                        args=args[1:]
                else:
                    cursor = cnxn.cursor()
                    try:
                        cursor.execute(query)
                    except Exception as e:
                        err_msg = "Invalid Query"
                        log.error(err_msg)
                        # logger.print_on_console(err_msg)    
                    rows = cursor.fetchall()
                    columns = [column[0] for column in cursor.description]
                ##logic for output col reading
                if dbtype == 13:
                    args=args[2:]
                if type(args) is tuple:
                    if str(args).startswith("(("):
                        args=args[0]
                    else:
                        args=args
                if (args[0].startswith("{")):
                    inp_path = self.DV.get_dynamic_value(args[0])
                    if inp_path!=None:
                        if len(inp_path.split(';'))>1:
                            fields=inp_path.split(";")[0]
                            inp_sheet=inp_path.split(";")[1]
                        else:
                            fields=inp_path.split(";")[0]
                            inp_sheet=None
                    else:
                        fields = None
                        inp_sheet=None
                elif (args[0].startswith("_")) and (args[0].endswith("_")):
                    inp_path = self.CV.get_constant_value(args[0])
                    if inp_path!=None:
                        if len(inp_path.split(';'))>1:
                            fields = None
                            inp_sheet=None
                        elif len(inp_path.split(';'))==1 and len(args)>1:
                            fields=inp_path.split(";")[0]
                            inp_sheet=args[1]                       
                        else:
                            fields=inp_path.split(";")[0]
                            inp_sheet=None
                    else:
                        fields = None
                        inp_sheet=None
                else:
                    if len(args)>1:
                        if (args[0].startswith('|') and args[0].endswith('|')):
                            from util_operations import UtilOperations
                            fields=UtilOperations().staticFetch(0, args[0])
                            inp_sheet = args[1]
                        else:
                            fields = args[0]
                            inp_sheet=args[1]
                    else:
                        if args[0].startswith('|') and args[0].endswith('|'):
                            from util_operations import UtilOperations
                            fields = UtilOperations().staticFetch(0, args[0])
                            inp_sheet = None
                        else:
                            fields=args[0]
                            inp_sheet=None
                ext = self.get_ext(fields)
                if (ext == '.xls'):
                    verify = os.path.isfile(fields)
                    try:
                        if(verify == False):
                            try:
                                #creating the New file.
                                work_book = xlwt.Workbook(encoding="utf-8") #Changed code
                                if(inp_sheet is None or inp_sheet == ''):
                                    inp_sheet = generic_constants.DATABASE_SHEET
                                try:
                                    work_sheet = work_book.add_sheet(inp_sheet)
                                except:
                                    work_sheet = work_book.get_sheet(inp_sheet)
                                    
                                index_sheet = work_book.sheet_index(inp_sheet)
                                work_book.set_active_sheet(index_sheet)
                                log.debug('Input Sheet and file path while creating file :')
                                log.debug(inp_sheet)
                                log.debug(fields)
                            except Exception as e:
                                err_msg = ERROR_CODE_DICT["ERR_FILE_WRITE"]
                                # logger.print_on_console(err_msg)
                                log.error(err_msg)
                                log.error(e)
                        else:
                            try:
                                #opening and copy the exisiting file.
                                rb = open_workbook(fields)
                                work_book = xl_copy(rb)
                                if(inp_sheet is None or inp_sheet == ''):
                                    inp_sheet = generic_constants.DATABASE_SHEET
                                    log.debug('Input Sheet is :')
                                    log.debug(inp_sheet)
                                    #work_book = xlwt.Workbook(encoding="utf-8")
                                    #changed the above line , becoz it removes the existing sheets.
                                try:
                                    work_sheet = work_book.add_sheet(inp_sheet)
                                except:
                                    work_sheet = work_book.get_sheet(inp_sheet)
                                    
                                index_sheet = work_book.sheet_index(inp_sheet)
                                work_book.set_active_sheet(index_sheet)
                            except Exception as e:
                                err_msg = ERROR_CODE_DICT["ERR_FILE_WRITE"]
                                # logger.print_on_console(err_msg)
                                log.error(err_msg)
                                log.error(e)
                        i=0
                        j=0
                        for x in columns:
                            work_sheet.write(i,j,x)
                            j=j+1
                        del i,j,x,columns
                        k=1
                        for row in rows:
                            l=0
                            for y in row:
                                try:
                                    work_sheet.write(k,l,y)
                                except Exception as e:
                                    if 'character' in str(e):
                                        i = str(e).index('character')
                                        err_new=str(e)[:i+len('character')]
                                        if err_new==generic_constants.UNICODE_ERR:
                                            newrow=[]
                                            for ele in row:
                                                if type(ele) == str:
                                                     ele = ele.encode('utf-8')
                                                     newrow.append(ele)
                                                else:
                                                    newrow.append(ele)
                                            del ele
                                        row = tuple(newrow)
                                        work_sheet.write(k,l,y)
                                l+=1
                            k+=1
                        work_book.save(fields)
                        try:
                            del k,l,row,rows,y
                            del work_book,work_sheet,index_sheet,rb
                        except Exception as e :
                            log.error('some error : {}'.format(e))
                    except Exception as e:
                        err_msg = ERROR_CODE_DICT["ERR_FILE_WRITE"]
                        logger.print_on_console(err_msg)
                        log.error(err_msg)
                        log.error(e)
                    if err_msg == None:
                        status=generic_constants.TEST_RESULT_PASS
                        result=generic_constants.TEST_RESULT_TRUE

                elif(ext == '.xlsx'):
                    verify = os.path.isfile(fields)
                    try:
                        if(verify == False):
                            try:
                                #Creating the New File.
                                work_book = openpyxl.Workbook()
                                
                                if(inp_sheet is None or inp_sheet == ''):
                                    inp_sheet = generic_constants.DATABASE_SHEET
                                    log.debug('Input Sheet is :')
                                    log.debug(inp_sheet)
                                    #adding the New Sheet.
                                sheet_names = work_book.sheetnames
                                if len(sheet_names)>0:
                                    if inp_sheet in sheet_names:
                                        index_sheet=sheet_names.index(inp_sheet)
                                    else:
                                        index_sheet = 0
                                        work_book.create_sheet(index=index_sheet, title=inp_sheet)
                                log.debug('Input Sheet and file path while creating file :')
                                log.debug(inp_sheet)
                                log.debug(fields)
                            except Exception as e:
                                err_msg = ERROR_CODE_DICT["ERR_FILE_WRITE"]
                                # logger.print_on_console(err_msg)
                                log.error(err_msg)
                                log.error(e)
                        else:
                            try:
                                #Opening the Exisiting File.
                                work_book = openpyxl.load_workbook(fields)
                                if(inp_sheet is None or inp_sheet == ''):
                                    inp_sheet = generic_constants.DATABASE_SHEET
                                    log.debug('Input Sheet is :')
                                    log.debug(inp_sheet)
                                
                                sheet_names = work_book.sheetnames
                                if len(sheet_names)>0:
                                    if inp_sheet in sheet_names:
                                        index_sheet=sheet_names.index(inp_sheet)
                                    else:
                                        index_sheet = 0
                                        work_book.create_sheet(index=index_sheet, title=inp_sheet)

                            except Exception as e:
                                err_msg = ERROR_CODE_DICT["ERR_FILE_WRITE"]
                                # logger.print_on_console(err_msg)
                                log.error(err_msg)
                                log.error(e)
                        i=1
                        j=1
                        work_sheet = work_book[inp_sheet]
                        
                        for x in columns:
                            work_sheet.cell(row=i, column=j).value = x
                            j=j+1
                            try:
                                max_col_width = 2.4
                                adjusted_width = (len(str(x)) + 2) * 1.2
                                if adjusted_width > max_col_width:
                                    max_col_width=adjusted_width
                                work_sheet.column_dimensions[get_column_letter(j)].width = max_col_width
                            except Exception as e:
                                log.info("Setting the dimensions for the column: {}".format(e))
                        try:
                            del x,columns,j,i,max_col_width,adjusted_width
                        except Exception as e :
                            log.error('some error : {}'.format(e))
                        k=2
                        for row in rows:
                            l=1
                            for y in row:
                                try:
                                    work_sheet.cell(row=k, column=l).value = y
                                    
                                except Exception as e:
                                    if 'character' in str(e):
                                        i = str(e).index('character')
                                        err_new=str(e)[:i+len('character')]
                                        if err_new==generic_constants.UNICODE_ERR:
                                            newrow=[]
                                            for ele in row:
                                                if type(ele) == str:
                                                     ele = ele.encode('utf-8')
                                                     newrow.append(ele)
                                                else:
                                                    newrow.append(ele)
                                            del ele
                                        row = tuple(newrow)
                                        work_sheet.cell(row=k, column=l).value = y
                                l+=1
                                try:
                                    max_col_width = 2.4
                                    adjusted_width = (len(str(y)) + 2) * 1.2
                                    if adjusted_width > max_col_width:
                                        max_col_width=adjusted_width
                                    work_sheet.column_dimensions[get_column_letter(l)].width = max_col_width
                                except Exception as e:
                                    log.info("Setting the dimensions for the rows: {}".format(e))
                            k+=1
                        
                        work_book.active=index_sheet
                        work_book.save(fields)
                        try:
                            del row,y,l,k,max_col_width,adjusted_width
                            del work_book,work_sheet,index_sheet
                        except Exception as e :
                            log.error('some error : {}'.format(e))
                    except Exception as e:
                        err_msg = ERROR_CODE_DICT["ERR_FILE_WRITE"]
                        logger.print_on_console(err_msg)
                        log.error(err_msg)
                        log.error(e)
                    if err_msg == None:
                        status=generic_constants.TEST_RESULT_PASS
                        result=generic_constants.TEST_RESULT_TRUE

                elif(ext == '.csv'): #Code changed to write csv files
                    path = fields
                    with open(path,'w') as csvfile:
                        writer = csv.writer(csvfile,lineterminator='\n')
                        writer.writerow(columns)
                        for row in rows:
                            try:
                                writer.writerow(row)
                            except Exception as e:
                                if 'character' in str(e):
                                    i = str(e).index('character')
                                    err_new=str(e)[:i+len('character')]
                                    if err_new==generic_constants.UNICODE_ERR:
                                        newrow=[]
                                        for ele in row:
                                            if type(ele) == str:
                                                 ele = ele.encode('utf-8')
                                                 newrow.append(ele)
                                            else:
                                                newrow.append(ele)
                                        del ele
                                    row = tuple(newrow)
                                    writer.writerow(row)
                    csvfile.close()
                    del row,columns
                    status=generic_constants.TEST_RESULT_PASS
                    result=generic_constants.TEST_RESULT_TRUE
                else:
                    err_msg = ERROR_CODE_DICT['ERR_INVALID_INPUT']
                    logger.print_on_console(err_msg)
                    log.error(err_msg)
        except Exception as e:
            err_msg = self.processException(e)
        finally:
            if cursor is not None: cursor.close()
            if cnxn is not None: cnxn.close()
        # if err_msg!=None:
        #     logger.print_on_console(err_msg)
        return status,result,verb,err_msg

    def secureExportData(self, ip , port , userName , password, dbName, query, dbtype,*args):
        """
        def : secureExportData
        purpose : Executes the query and exports the data to excel or csv file
        param : database type, IP, port number , database name, username , password , query , Filename , sheetname
        return : bool
        """
        status=generic_constants.TEST_RESULT_FAIL
        result=generic_constants.TEST_RESULT_FALSE
        verb = OUTPUT_CONSTANT
        err_msg=None
        try:
            encryption_obj = AESCipher()
            decrypted_password = encryption_obj.decrypt(password)
            out_col = args
            status,result,verb,err_msg=self.exportData(ip,port,userName,decrypted_password,dbName, query, dbtype, out_col,args)
        except Exception as e:
            err_msg = self.processException(e)
        return status,result,verb,err_msg

    def connection(self,dbtype, ip , port , dbName, userName , password, *args):
        """
        def : connection
        purpose : connecting to database based on database type given
        param : database type, IP, port number , database name, username , password , query
        return : connection object
        """
        dbNumber = {4:'{SQL Server}',5:'{Microsoft ODBC for Oracle}'}#,2:'{IBM DB2 ODBC DRIVER}'}
        err_msg=None
        if not(dbtype == ''): 
            try:
                dbtype= int(dbtype)
                if dbtype == 2:
                    import ibm_db
                    cnxn = ibm_db.connect('database=%s;hostname=%s;port=%s;protocol=TCPIP;uid=%s;pwd=%s'%(dbName,ip,port,userName,password), '', '')
                    import ibm_db_dbi
                    self.cnxn = ibm_db_dbi.Connection(cnxn)
                elif dbtype == 6:
                    import cx_Oracle
                    path = os.path.normpath(os.environ["AVO_ASSURE_HOME"] + "/Lib/instantclient")
                    os.environ["PATH"] += os.pathsep + path
                    host = str(ip)+":"+str(port)+"/"+dbName
                    self.cnxn = cx_Oracle.connect(userName, password, host, encoding="UTF-8")
                elif dbtype == 7:             
                    jar_path = os.path.normpath(os.environ["AVO_ASSURE_HOME"] + "/Lib/DB/jtds-1.3.1.jar")
                    self.cnxn = jaydebeapi.connect('net.sourceforge.jtds.jdbc.Driver', 'jdbc:jtds:sybase://%s:%s/%s' % (ip, port, dbName), [userName, password], jar_path)    
                ##added to connect with mdb files mdb files using pyodbc 
                elif dbtype == 3:
                    if(userName == '' and port == '' and dbName == ''):
                        filename, file_ext=os.path.splitext(ip)
                        if file_ext in ['.mdb','.accdb']:
                            self.cnxn = pyodbc.connect('driver={Microsoft Access Driver (*.mdb, *.accdb)};PORT=%s;DATABASE=%s;UID=%s;PWD=%s;DBQ=%s' % ( port, dbName, userName ,password,ip ) )
                        else:
                            err_msg=generic_constants.INVALID_FILE_FORMAT
                            logger.print_on_console(err_msg)
                            log.error(err_msg)
                    else:
                        err_msg=generic_constants.INVALID_INPUT
                        logger.print_on_console(err_msg)
                        log.error(err_msg)
                elif dbtype == 8:
                    import pymysql
                    self.cnxn = pymysql.connect(host=ip,user=userName,password=password,db=dbName)
                elif dbtype == 9:
                    from pymongo import MongoClient
                    self.cnxn = MongoClient(ip,int(port),username=userName,password=password,authSource=dbName)
                elif dbtype == 10:
                    import redis
                    self.cnxn = redis.StrictRedis(host=ip,port=int(port),password=password,db=int(dbName))
                elif dbtype == 11:
                    from cassandra.cluster import Cluster
                    from cassandra.auth import PlainTextAuthProvider
                    self.cnxn = Cluster([ip],port=port,auth_provider=PlainTextAuthProvider(username=userName, password=password)).connect()
                    self.cnxn.set_keyspace(dbName)
                elif dbtype == 12:
                    import sqlite3
                    self.cnxn = sqlite3.connect(ip)
                elif dbtype == 13:
                    import snowflake.connector
                    self.cnxn = snowflake.connector.connect(user=userName,password=password,account=ip,warehouse=args[0][0],database=dbName,schema=args[0][1])
                elif dbtype == 14:
                    import psycopg2
                    self.cnxn = psycopg2.connect(host=ip,port=port,user=userName,password=password,database=dbName)
                else:
                    self.cnxn = pyodbc.connect('driver=%s;SERVER=%s;PORT=%s;DATABASE=%s;UID=%s;PWD=%s' % ( dbNumber[dbtype], ip, port, dbName, userName ,password ) )
                return self.cnxn
            except Exception as e:
                self.processException(e)
        else:
            err_msg=generic_constants.INVALID_INPUT
            logger.print_on_console(err_msg)
            log.error(err_msg)
        return None

    def get_ext(self,input_path):
        """
        def : __get_ext
        purpose : returns the file type and verifies if it is valid file type
        param : input_path
        return : bool,filetype
        """
        status=False
        try:
            filename, file_ext=os.path.splitext(input_path)
            if file_ext in generic_constants.FILE_TYPES:
                status=True
            else:
                logger.print_on_console(generic_constants.INVALID_FILE_FORMAT)
            return file_ext
        except:
            pass
        return '',status

    def create_file_xls(self): #Method to create xls file
        """
        def : create_file
        purpose : creates a xls file inside generic folder directory
        param :
        return : path
        """
        path = None
        try:
            wb = xlwt.Workbook()
            ws = wb.add_sheet(generic_constants.DATABASE_SHEET)
            maindir = os.environ["AVO_ASSURE_HOME"]
            path = maindir +'/plugins/Generic' + generic_constants.DATABASE_FILE_XLS
            wb.save(path)
        except Exception as e:
            log.error(e)
        return path

    def create_file_xlsx(self): #Method to create xlsx file
        """
        def : create_file
        purpose : creates a xlsx file inside generic folder directory
        param :
        return : path
        """
        path=None
        try:
            wb = xlwt.Workbook()
            ws = wb.add_sheet(generic_constants.DATABASE_SHEET)
            maindir = os.environ["AVO_ASSURE_HOME"]
            path = maindir +'/plugins/Generic' + generic_constants.DATABASE_FILE_XLSX
            wb.save(path)
        except Exception as e:
            log.error(e)
        return path

    def create_file_csv(self): #Method to create csv file
        """
        def : create_file
        purpose : creates a csv file inside generic folder directory
        param :
        return : path
        """
        path=None
        try:
            maindir = os.environ["AVO_ASSURE_HOME"]
            path = maindir +'/plugins/Generic' + generic_constants.DATABASE_FILE_CSV
            with open(path,'wb') as file:
                pass
            file.close()
        except Exception as e:
            log.error(e)
        return path


##obj = DatabaseOperation()
##obj.runQuery('10.44.10.54','1433','version20_test','version2.0_Test','Version20_TestDB',"select * from Persons",'4')
##obj.secureRunQuery('10.44.10.54','1433','version20_test','451MwSMdBUUKiSlqb85IKvYYAt8ag1cO0JE5cEDdotg=','Version20_TestDB',"select * from Persons",'4')
##obj.getData('10.44.10.54','1433','version20_test','version2.0_Test','Version20_TestDB',"select * from Persons",'4')
##obj.secureGetData('10.44.10.54','1433','version20_test','O0v/vLLjfD+hR22U9lzTifYYAt8ag1cO0JE5cEDdotg=','Version20_TestDB',"select * from Persons",'4')
##obj.exportData('4','10.44.10.54','1433','Version20_TestDB','version20_test','version2.0_Test',"select * from Persons",'D:\db6.xls','Sheet1')
##obj.secureExportData('10.44.10.54','1433','version20_test','O0v/vLLjfD+hR22U9lzTifYYAt8ag1cO0JE5cEDdotg=','Version20_TestDB',"select * from Persons",'4','D:\db6.xls','Sheet1')
##obj.exportData('10.44.10.54','1433','version20_test','version2.0_Test','Version20_TestDB',"select * from Persons",'4','D:\db6.xls','Sheet1')
##obj.verifyData('4','10.44.10.54','1433','Version20_TestDB','version20_test','version2.0_Test',"select * from Persons",'D:\db5.xls','Sheet1')
##obj.secureVerifyData('10.44.10.54','1433','version20_test','O0v/vLLjfD+hR22U9lzTifYYAt8ag1cO0JE5cEDdotg=','Version20_TestDB',"select * from Persons",'4','D:\db6.xls','Sheet1')